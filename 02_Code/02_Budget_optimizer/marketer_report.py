"""Build a marketer-facing forecast/optimizer report.

This report is intentionally different from technical audit exports. It hides
backend columns and answers the business workflow:

1. What campaign brief was uploaded?
2. What scenarios were evaluated?
3. What do we expect from each scenario?
4. What did adaptive scenario 6 try and what is the best reliable allocation?
5. Which plan should a marketer choose, and how much should they trust it?
"""

from __future__ import annotations

import argparse
import json
import math
import re
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

PYMC_CODE_DIR = Path(__file__).resolve().parents[1] / "01_PyMC"
if str(PYMC_CODE_DIR) not in sys.path:
    sys.path.insert(0, str(PYMC_CODE_DIR))

from mmm_core.forecast_engine import _compile_decision_policy
from mmm_core.model_package import sha256_file

SCENARIOS = [
    {
        "scenario_code": "S01_AS_IS",
        "candidate_marker": "__scenario1_current_plan",
        "scenario_name": "Сценарий 1. Как загрузили",
        "plain_description": "Ничего не меняем: бюджет, каналы, гео и даты остаются как в исходной спецификации маркетолога.",
    },
    {
        "scenario_code": "S02_EQUAL_CELL",
        "candidate_marker": "__scenario2_equal_cell_split",
        "scenario_name": "Сценарий 2. Ровно по всем связкам",
        "plain_description": "Общий бюджет кампании делится поровну между всеми исходными связками geo x channel.",
    },
    {
        "scenario_code": "S03_KEEP_CHANNEL_EQUAL_GEO",
        "candidate_marker": "__scenario3_channel_balanced",
        "scenario_name": "Сценарий 3. Каналы как были, гео ровно",
        "plain_description": "Сохраняем бюджет каждого канала, но внутри канала делим его поровну между его гео.",
    },
    {
        "scenario_code": "S04_KEEP_GEO_EQUAL_CHANNEL",
        "candidate_marker": "__scenario4_geo_balanced",
        "scenario_name": "Сценарий 4. Гео как были, каналы ровно",
        "plain_description": "Сохраняем бюджет каждого гео, но внутри гео делим его поровну между доступными каналами.",
    },
    {
        "scenario_code": "S05_SUPPORT_SAFE",
        "candidate_marker": "__scenario5_",
        "scenario_name": "Сценарий 5. Осторожный сценарий",
        "plain_description": "Сначала ищем полный план с минимальным риском, затем расширяем допустимый диапазон. Частичный вариант показывается только когда весь бюджет надежно распределить невозможно.",
    },
]


@dataclass(frozen=True)
class ReportPaths:
    model_run_dir: Path
    flighting_path: Path
    optimizer_output_dir: Path
    output_xlsx: Path
    run_id: str | None


def _parse_args() -> ReportPaths:
    parser = argparse.ArgumentParser(description="Build marketer-facing MMM forecast/optimizer report.")
    parser.add_argument("--model-run-dir", required=True)
    parser.add_argument("--flighting-path", required=True)
    parser.add_argument("--optimizer-output-dir", required=True)
    parser.add_argument("--output-xlsx", required=True)
    parser.add_argument("--run-id", default=None)
    args = parser.parse_args()
    return ReportPaths(
        model_run_dir=Path(args.model_run_dir).expanduser().resolve(),
        flighting_path=Path(args.flighting_path).expanduser().resolve(),
        optimizer_output_dir=Path(args.optimizer_output_dir).expanduser().resolve(),
        output_xlsx=Path(args.output_xlsx).expanduser().resolve(),
        run_id=args.run_id,
    )


def _read_csv(path: Path) -> pd.DataFrame:
    if not path.exists():
        raise FileNotFoundError(path)
    return pd.read_csv(path)


def _first_existing(directory: Path, suffix: str) -> Path | None:
    matches = sorted(directory.glob(f"*{suffix}"))
    return matches[0] if matches else None


def _safe_run_id(value: Any) -> str:
    return str(value or "").strip().replace("/", "_").replace("::", "__").replace(" ", "_")


def _campaign_prepare_card(paths: ReportPaths) -> dict[str, Any]:
    if paths.run_id:
        card_path = paths.optimizer_output_dir / f"{_safe_run_id(paths.run_id)}_campaign_prepare_card.json"
    else:
        matches = sorted(paths.optimizer_output_dir.glob("*_campaign_prepare_card.json"))
        card_path = matches[0] if matches else Path("__missing__")
    if not card_path.exists():
        return {}
    return json.loads(card_path.read_text(encoding="utf-8"))


def _source_campaign_context(paths: ReportPaths) -> pd.DataFrame:
    card = _campaign_prepare_card(paths)
    adapter = card.get("source_adapter_audit") or {}
    campaigns = adapter.get("campaigns") or []
    return pd.DataFrame(campaigns)


def _campaign_support_bounds(paths: ReportPaths) -> pd.DataFrame:
    support_path = paths.model_run_dir / "historical_campaign_support_bounds.csv"
    return _read_csv(support_path) if support_path.exists() else pd.DataFrame()


def _million(value: Any) -> float:
    value = float(value or 0.0)
    return value / 1_000_000.0


def _full_list(values: pd.Series) -> str:
    """Serialize a complete machine-readable set without presentation truncation."""

    return ", ".join(
        str(value)
        for value in sorted(set(values.dropna().astype(str)))
        if str(value)
    )


def _campaign_cells(flighting: pd.DataFrame) -> pd.DataFrame:
    df = flighting.copy()
    df["date"] = pd.to_datetime(df["date"]).dt.date
    return (
        df.groupby(["campaign_name", "segment", "geo", "channel"], dropna=False)
        .agg(
            budget_rub=("budget_rub", "sum"),
            start_date=("date", "min"),
            end_date=("date", "max"),
            active_days=("date", "nunique"),
            creative_name=("creative_name", _full_list),
        )
        .reset_index()
    )


def _scenario_code_from_candidate(candidate_name: Any) -> str | None:
    text = str(candidate_name or "")
    for scenario in SCENARIOS:
        if scenario["candidate_marker"] in text:
            return str(scenario["scenario_code"])
    return None


def _campaign_total_rows(finalist_summary: pd.DataFrame) -> pd.DataFrame:
    total = finalist_summary[finalist_summary["channel"].eq("__TOTAL__")].copy()
    campaign_total = total[total["segment"].eq("__ALL__")]
    return campaign_total if not campaign_total.empty else total


def _result_rows_from_optimizer(
    finalist_summary: pd.DataFrame,
    candidate_scores: pd.DataFrame,
) -> pd.DataFrame:
    """Build scenarios 1-5 from optimizer artifacts without rerunning forecast."""
    total = _campaign_total_rows(finalist_summary)
    total["scenario_code"] = total["candidate_name"].map(_scenario_code_from_candidate)
    total = total[total["scenario_code"].notna()].copy()
    rows: list[dict[str, Any]] = []
    scenario_lookup = {s["scenario_code"]: s for s in SCENARIOS}
    for (source_campaign, candidate_name, scenario_code), sub in total.groupby(
        ["source_campaign_name", "candidate_name", "scenario_code"],
        dropna=False,
    ):
        scenario = scenario_lookup[scenario_code]
        score_rows = candidate_scores[candidate_scores["candidate_name"].eq(candidate_name)]
        row: dict[str, Any] = {
            "campaign_name": source_campaign,
            "scenario_no": scenario_code[:3],
            "scenario_name": scenario["scenario_name"],
            "scenario_description": scenario["plain_description"],
            "candidate_name": candidate_name,
            "budget_mln_rub": _million(sub["spend_rub"].max()),
        }
        requested_budget = (
            float(score_rows["requested_budget_rub"].iloc[0])
            if not score_rows.empty and "requested_budget_rub" in score_rows
            else float(sub["spend_rub"].max())
        )
        allocated_budget = (
            float(score_rows["allocated_budget_rub"].iloc[0])
            if not score_rows.empty and "allocated_budget_rub" in score_rows
            else float(sub["spend_rub"].max())
        )
        unallocated_budget = max(requested_budget - allocated_budget, 0.0)
        row["requested_budget_mln_rub"] = _million(requested_budget)
        row["allocated_budget_mln_rub"] = _million(allocated_budget)
        row["unallocated_budget_mln_rub"] = _million(unallocated_budget)
        row["allocated_budget_share"] = allocated_budget / requested_budget if requested_budget > 0 else 0.0
        row["allocation_share"] = row["allocated_budget_share"]
        row["support_limit_policy"] = (
            str(score_rows["support_limit_policy"].iloc[0])
            if not score_rows.empty and "support_limit_policy" in score_rows
            else "unconstrained_benchmark"
        )
        support_column = "spend_support_warnings_n" if "spend_support_warnings_n" in sub else "support_warnings_n"
        support_warnings = int(pd.to_numeric(sub[support_column], errors="coerce").fillna(0).max())
        row["support_warnings_n"] = support_warnings
        row["elevated_support_warnings_n"] = (
            int(score_rows["elevated_support_warnings_n"].iloc[0])
            if not score_rows.empty and "elevated_support_warnings_n" in score_rows
            else int(pd.to_numeric(sub.get("support_elevated_n"), errors="coerce").fillna(0).max())
            if "support_elevated_n" in sub
            else 0
        )
        row["strong_support_warnings_n"] = (
            int(score_rows["strong_support_warnings_n"].iloc[0])
            if not score_rows.empty and "strong_support_warnings_n" in score_rows
            else int(pd.to_numeric(sub.get("support_strong_n"), errors="coerce").fillna(0).max())
            if "support_strong_n" in sub
            else 0
        )
        row["hard_support_warnings_n"] = int(score_rows["hard_support_warnings_n"].iloc[0]) if not score_rows.empty else 0
        policy_violations = int(score_rows["policy_violations_n"].iloc[0]) if not score_rows.empty else 0
        risk_policy_violations = (
            int(score_rows["risk_policy_violations_n"].iloc[0])
            if not score_rows.empty and "risk_policy_violations_n" in score_rows
            else 0
        )
        row["policy_violations_n"] = policy_violations + risk_policy_violations
        row["policy_violation_codes"] = str(score_rows["policy_violation_codes"].iloc[0]) if not score_rows.empty else "OK"
        for column in [
            "scenario_kind",
            "scenario_variant",
            "scenario_feasibility_status",
            "full_allocation_impossible_reason",
            "limiting_constraints",
            "within_support_budget_rub",
            "within_support_share",
            "controlled_extrapolation_budget_rub",
            "controlled_extrapolation_share",
            "high_risk_budget_rub",
            "high_risk_share",
            "within_support_cells_n",
            "controlled_extrapolation_cells_n",
            "high_risk_cells_n",
        ]:
            row[column] = (
                score_rows[column].iloc[0]
                if not score_rows.empty and column in score_rows
                else np.nan
            )
        if scenario_code == "S05_SUPPORT_SAFE" and row.get("scenario_variant") == "safe_partial":
            row["scenario_name"] = "Сценарий 5. Безопасно распределяемая часть"
            row["scenario_description"] = (
                "Весь бюджет нельзя разместить в пределах утвержденной границы риска. "
                "Показана рассчитанная часть и явный остаток для ручного решения."
            )
        elif scenario_code == "S05_SUPPORT_SAFE":
            row["scenario_name"] = "Сценарий 5. Осторожный полный план"
        for target, prefix in [("turnover_per_user", "rto")]:
            t = sub[sub["target"].eq(target)]
            if t.empty:
                continue
            rec = t.iloc[0]
            row[f"{prefix}_p10_mln"] = _million(rec["total_effect_p10"]) if prefix != "orders" else float(rec["total_effect_p10"])
            row[f"{prefix}_p50_mln"] = _million(rec["total_effect_p50"]) if prefix != "orders" else float(rec["total_effect_p50"])
            row[f"{prefix}_p90_mln"] = _million(rec["total_effect_p90"]) if prefix != "orders" else float(rec["total_effect_p90"])
            row[f"{prefix}_roas_p50"] = rec.get("roas_p50", "")
            row[f"{prefix}_allowed"] = rec.get("allowed_use_counts", "")
            row[f"{prefix}_risk"] = rec.get("risk_level_counts", "")
            row[f"{prefix}_optimizer_use"] = rec.get("optimizer_use_counts", "")
            if prefix == "rto":
                for column in [
                    "paired_delta_p10",
                    "paired_delta_p50",
                    "paired_delta_p90",
                    "paired_probability_gt_zero",
                    "paired_probability_gt_materiality",
                    "paired_probability_noninferior",
                    "paired_materiality_threshold",
                    "paired_noninferiority_floor",
                    "paired_draws_n",
                ]:
                    value = rec.get(column, np.nan)
                    if column.startswith("paired_delta") or column in {
                        "paired_materiality_threshold",
                        "paired_noninferiority_floor",
                    }:
                        value = _million(value) if pd.notna(value) else np.nan
                    row[column] = value
                row["roas_allocated_budget_p10"] = (
                    float(rec["total_effect_p10"]) / allocated_budget
                    if allocated_budget > 0
                    else np.nan
                )
                row["roas_allocated_budget_p50"] = (
                    float(rec["total_effect_p50"]) / allocated_budget
                    if allocated_budget > 0
                    else np.nan
                )
                row["roas_allocated_budget_p90"] = (
                    float(rec["total_effect_p90"]) / allocated_budget
                    if allocated_budget > 0
                    else np.nan
                )
                row["roas_requested_budget_p10"] = (
                    float(rec["total_effect_p10"]) / requested_budget
                    if requested_budget > 0
                    else np.nan
                )
                row["roas_requested_budget_p50"] = (
                    float(rec["total_effect_p50"]) / requested_budget
                    if requested_budget > 0
                    else np.nan
                )
                row["roas_requested_budget_p90"] = (
                    float(rec["total_effect_p90"]) / requested_budget
                    if requested_budget > 0
                    else np.nan
                )
                row["roas_denominator_kind"] = (
                    "allocated_budget"
                    if abs(requested_budget - allocated_budget) > 1.0
                    else "requested_budget"
                )
                row["roas_denominator_budget_rub"] = (
                    allocated_budget
                    if row["roas_denominator_kind"] == "allocated_budget"
                    else requested_budget
                )
        row["calculation_status"] = "Рассчитано частично" if unallocated_budget > 1.0 else "Рассчитано"
        row["cell_support_status"] = _cell_support_status(row)
        row["optimizer_status"] = _optimizer_status(row)
        row["quality_status"] = _human_quality(row)
        row["quality_explanation"] = _quality_explanation(row)
        rows.append(row)
    return pd.DataFrame(rows).sort_values(["campaign_name", "scenario_no"])


def _cell_support_status(row: dict[str, Any]) -> str:
    if int(row.get("hard_support_warnings_n") or 0) > 0:
        return "Вне надежной наблюдаемой зоны"
    if int(row.get("strong_support_warnings_n") or 0) > 0:
        return "Между p99 и robust upper"
    if int(row.get("elevated_support_warnings_n") or 0) > 0:
        return "Между p95 и p99"
    return "Внутри p95 support-zone"


def _optimizer_status(row: dict[str, Any]) -> str:
    if int(row.get("hard_support_warnings_n") or 0) > 0 or int(row.get("policy_violations_n") or 0) > 0:
        return "Только ручное распределение"
    if float(row.get("unallocated_budget_mln_rub") or 0.0) > 0.000001:
        return "Частичный безопасный план"
    objective_rows_n = int(row.get("objective_rows_n") or _objective_rows_count(row.get("rto_optimizer_use", "")))
    if objective_rows_n <= 0:
        return "Перераспределение недоступно по gate policy"
    if int(row.get("strong_support_warnings_n") or 0) > 0:
        return "Автоматический план доступен с оговоркой"
    return "Автоматический план доступен"


def _human_quality(row: dict[str, Any]) -> str:
    if str(row.get("calculation_status") or "").lower() == "blocked":
        return "Расчет невозможен"
    if int(row.get("hard_support_warnings_n") or 0) > 0 or int(row.get("policy_violations_n") or 0) > 0:
        return "Не использовать для автоматического перераспределения"
    if int(row.get("strong_support_warnings_n") or 0) > 0:
        return "Повышенная неопределенность"
    allowed = str(row.get("rto_allowed") or "")
    risk = str(row.get("rto_risk") or "")
    objective_rows_n = int(row.get("objective_rows_n") or _objective_rows_count(row.get("rto_optimizer_use", "")))
    if "diagnostic" in allowed or "high" in risk:
        return "Повышенная неопределенность" if objective_rows_n > 0 else "Требуется ручная проверка"
    if (
        int(row.get("elevated_support_warnings_n") or 0) > 0
        or float(row.get("unallocated_budget_mln_rub") or 0.0) > 0.000001
        or "caution" in allowed
        or "medium" in risk
    ):
        return "Повышенная неопределенность"
    return "Сопоставимо с историей"


def _quality_explanation(row: dict[str, Any]) -> str:
    if int(row.get("hard_support_warnings_n") or 0) > 0:
        return (
            "Прогноз рассчитан, но хотя бы одна geo x channel ячейка выше robust observed upper. "
            "Это не запрет кампании: автоматическое перераспределение блокируется, медиаплан проверяется вручную."
        )
    if int(row.get("policy_violations_n") or 0) > 0:
        return (
            "Сценарий меняет канал, который gate policy разрешает только сохранить или не увеличивать. "
            "Расчет остается сравнительным benchmark, но не рекомендацией."
        )
    if int(row.get("strong_support_warnings_n") or 0) > 0:
        return (
            "Часть бюджета находится выше p99, но не выше утвержденной robust upper. "
            "Это контролируемое расширение исторического диапазона: автоматическое распределение "
            "допустимо по policy, но требует явной оговорки."
        )
    if float(row.get("unallocated_budget_mln_rub") or 0.0) > 0.000001:
        return (
            "Система безопасно разместила только часть бюджета. Остаток не спрятан в экстраполяции и требует ручного решения."
        )
    if int(row.get("elevated_support_warnings_n") or 0) > 0:
        return "Есть ячейки между historical p95 и p99: расчет допустим, но неопределенность выше обычной."
    if row.get("quality_status") == "Повышенная неопределенность":
        return "Расчет выполнен, но часть channel/fit evidence имеет caution или diagnostic ограничения."
    if row.get("quality_status") == "Требуется ручная проверка":
        return "Эффект показан для прозрачности, но gate policy не разрешает использовать его как автоматический optimizer KPI."
    return "Сценарий находится внутри p95 support-zone и не нарушает model policy."


def _count_token(value: Any, token: str) -> int:
    total = 0
    for part in str(value or "").split(";"):
        if ":" not in part:
            continue
        key, raw = part.split(":", 1)
        if key.strip() != token:
            continue
        try:
            total += int(float(raw))
        except ValueError:
            pass
    return total


def _objective_rows_count(value: Any) -> int:
    return sum(
        _count_token(value, token)
        for token in ["optimize", "no_increase", "objective_allowed", "objective_allowed_with_penalty"]
    )


def _quality_rank(value: Any) -> int:
    return {
        "Сопоставимо с историей": 0,
        "Повышенная неопределенность": 1,
        "Требуется ручная проверка": 2,
        "Не использовать для автоматического перераспределения": 3,
        "Расчет невозможен": 4,
    }.get(str(value or ""), 9)


def _is_scenario6_candidate(value: Any) -> bool:
    return "__scenario6_" in str(value or "")


def _load_optimizer_tables(paths: ReportPaths) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    if paths.run_id:
        safe_run_id = str(paths.run_id).strip().replace("/", "_").replace("::", "__").replace(" ", "_")
        candidate_path = paths.optimizer_output_dir / f"{safe_run_id}_optimizer_candidate_scores.csv"
        finalist_path = paths.optimizer_output_dir / f"{safe_run_id}_optimizer_finalist_summary.csv"
        allocation_path = paths.optimizer_output_dir / f"{safe_run_id}_optimizer_recommended_allocations.csv"
    else:
        candidate_path = _first_existing(paths.optimizer_output_dir, "_optimizer_candidate_scores.csv")
        finalist_path = _first_existing(paths.optimizer_output_dir, "_optimizer_finalist_summary.csv")
        allocation_path = _first_existing(paths.optimizer_output_dir, "_optimizer_recommended_allocations.csv")
    if candidate_path is None or finalist_path is None or allocation_path is None:
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()
    if not candidate_path.exists() or not finalist_path.exists() or not allocation_path.exists():
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()
    return _read_csv(candidate_path), _read_csv(finalist_path), _read_csv(allocation_path)


def _load_paired_comparisons(paths: ReportPaths) -> pd.DataFrame:
    """Load optional paired posterior deltas produced by the optimizer."""
    if paths.run_id:
        paired_path = (
            paths.optimizer_output_dir
            / f"{_safe_run_id(paths.run_id)}_optimizer_paired_comparisons.csv"
        )
    else:
        paired_path = _first_existing(paths.optimizer_output_dir, "_optimizer_paired_comparisons.csv")
    if paired_path is None or not paired_path.exists():
        return pd.DataFrame()
    return _read_csv(paired_path)


def _attach_paired_comparisons(
    finalist_summary: pd.DataFrame,
    paired_comparisons: pd.DataFrame,
) -> pd.DataFrame:
    """Attach candidate-vs-source posterior evidence to cached finalist rows."""
    if finalist_summary.empty or paired_comparisons.empty:
        return finalist_summary
    paired_columns = [
        "candidate_name",
        "target",
        "paired_delta_p10",
        "paired_delta_p50",
        "paired_delta_p90",
        "paired_probability_gt_zero",
        "paired_probability_gt_materiality",
        "paired_probability_noninferior",
        "paired_materiality_threshold",
        "paired_noninferiority_floor",
        "paired_draws_n",
    ]
    available = [column for column in paired_columns if column in paired_comparisons.columns]
    if not {"candidate_name", "target"}.issubset(available):
        return finalist_summary
    paired = paired_comparisons[available].drop_duplicates(["candidate_name", "target"])
    stale = [
        column
        for column in available
        if column not in {"candidate_name", "target"} and column in finalist_summary.columns
    ]
    base = finalist_summary.drop(columns=stale, errors="ignore")
    return base.merge(
        paired,
        on=["candidate_name", "target"],
        how="left",
        validate="many_to_one",
    )


def _scenario6_summary(candidate_scores: pd.DataFrame, finalist_summary: pd.DataFrame) -> pd.DataFrame:
    if candidate_scores.empty:
        return pd.DataFrame()
    total = _campaign_total_rows(finalist_summary) if not finalist_summary.empty else pd.DataFrame()
    if not total.empty:
        total = total[total["candidate_name"].map(_is_scenario6_candidate)].copy()
    s6_scores = candidate_scores[candidate_scores["candidate_name"].map(_is_scenario6_candidate)].copy()
    rows: list[dict[str, Any]] = []
    total_groups = [] if total.empty else total.groupby(
        ["source_campaign_name", "candidate_name"], dropna=False
    )
    for (source_campaign, candidate_name), sub in total_groups:
        rto = sub[sub["target"].eq("turnover_per_user")]
        if rto.empty:
            continue
        rec = rto.iloc[0]
        tries = s6_scores[s6_scores["campaign_name"].eq(source_campaign)]
        score_rows = tries[tries["candidate_name"].eq(candidate_name)]
        strong_tries = pd.to_numeric(
            tries.get("strong_support_warnings_n", pd.Series(0, index=tries.index)),
            errors="coerce",
        ).fillna(0)
        hard_tries = pd.to_numeric(
            tries.get("hard_support_warnings_n", pd.Series(0, index=tries.index)),
            errors="coerce",
        ).fillna(0)
        statuses = tries.get("precheck_status", pd.Series("", index=tries.index)).fillna("")
        rejected_n = int(((strong_tries.gt(0) | hard_tries.gt(0)) | statuses.eq("rejected_infeasible")).sum())
        support_warnings = int(rec.get("spend_support_warnings_n") or rec.get("support_warnings_n") or 0)
        policy_violations = int(score_rows["policy_violations_n"].iloc[0]) if not score_rows.empty else 0
        if not score_rows.empty and "risk_policy_violations_n" in score_rows:
            policy_violations += int(score_rows["risk_policy_violations_n"].iloc[0])
        objective_rows_n = _objective_rows_count(rec.get("optimizer_use_counts", ""))
        hard_support_warnings = (
            int(score_rows["hard_support_warnings_n"].iloc[0])
            if not score_rows.empty and "hard_support_warnings_n" in score_rows
            else 0
        )
        elevated_support_warnings = (
            int(score_rows["elevated_support_warnings_n"].iloc[0])
            if not score_rows.empty and "elevated_support_warnings_n" in score_rows
            else int(rec.get("support_elevated_n") or 0)
        )
        strong_support_warnings = (
            int(score_rows["strong_support_warnings_n"].iloc[0])
            if not score_rows.empty and "strong_support_warnings_n" in score_rows
            else int(rec.get("support_strong_n") or 0)
        )
        requested_budget = (
            float(score_rows["requested_budget_rub"].iloc[0])
            if not score_rows.empty and "requested_budget_rub" in score_rows
            else float(rec.get("requested_budget_rub") or rec["spend_rub"])
        )
        allocated_budget = (
            float(score_rows["allocated_budget_rub"].iloc[0])
            if not score_rows.empty and "allocated_budget_rub" in score_rows
            else float(rec.get("allocated_budget_rub") or rec["spend_rub"])
        )
        unallocated_budget = max(requested_budget - allocated_budget, 0.0)
        quality_input = {
            "support_warnings_n": support_warnings,
            "elevated_support_warnings_n": elevated_support_warnings,
            "strong_support_warnings_n": strong_support_warnings,
            "hard_support_warnings_n": hard_support_warnings,
            "policy_violations_n": policy_violations,
            "objective_rows_n": objective_rows_n,
            "unallocated_budget_mln_rub": _million(unallocated_budget),
            "rto_allowed": rec.get("allowed_use_counts", ""),
            "rto_risk": rec.get("risk_level_counts", ""),
        }
        quality_status = _human_quality(quality_input)
        rows.append(
            {
                "campaign_name": source_campaign,
                "candidate_name": candidate_name,
                "scenario6_ran": True,
                "is_status_row": False,
                "attempts_total_n": int(
                    pd.to_numeric(
                        tries.get("search_attempts_evaluated_n", pd.Series(dtype=float)),
                        errors="coerce",
                    ).max()
                )
                if "search_attempts_evaluated_n" in tries
                and pd.to_numeric(tries["search_attempts_evaluated_n"], errors="coerce").notna().any()
                else int(tries[~statuses.eq("not_run_no_modifiable_cells")].shape[0]),
                "candidate_plans_n": int(tries[statuses.eq("scored")].shape[0]),
                "unique_allocations_n": int(
                    pd.to_numeric(score_rows["search_unique_allocations_n"], errors="coerce").iloc[0]
                )
                if not score_rows.empty and "search_unique_allocations_n" in score_rows
                else np.nan,
                "kernel_evaluations_n": int(
                    pd.to_numeric(score_rows["search_kernel_evaluations_n"], errors="coerce").iloc[0]
                )
                if not score_rows.empty and "search_kernel_evaluations_n" in score_rows
                else np.nan,
                "effective_dimension_n": int(
                    pd.to_numeric(score_rows["search_effective_dimension_n"], errors="coerce").iloc[0]
                )
                if not score_rows.empty and "search_effective_dimension_n" in score_rows
                else np.nan,
                "search_converged": bool(score_rows["search_converged"].iloc[0])
                if not score_rows.empty and "search_converged" in score_rows
                else False,
                "search_budget_exhausted": bool(score_rows["search_budget_exhausted"].iloc[0])
                if not score_rows.empty and "search_budget_exhausted" in score_rows
                else False,
                "search_posterior_samples": int(
                    pd.to_numeric(score_rows["search_posterior_samples"], errors="coerce").iloc[0]
                )
                if not score_rows.empty and "search_posterior_samples" in score_rows
                else np.nan,
                "smallest_transfer_mln_rub": _million(
                    score_rows["search_smallest_transfer_rub"].iloc[0]
                )
                if not score_rows.empty and "search_smallest_transfer_rub" in score_rows
                else np.nan,
                "attempts_rejected_by_support_n": rejected_n,
                "attempts_rejected_by_policy_n": int(
                    pd.to_numeric(tries.get("policy_violations_n"), errors="coerce").fillna(0).gt(0).sum()
                )
                if "policy_violations_n" in tries
                else 0,
                "attempts_rejected_precheck_n": int(statuses.eq("rejected_infeasible").sum()),
                "search_method": _full_list(tries["method"]) if not tries.empty else "none",
                "requested_budget_mln_rub": _million(requested_budget),
                "allocated_budget_mln_rub": _million(allocated_budget),
                "unallocated_budget_mln_rub": _million(unallocated_budget),
                "allocated_budget_share": allocated_budget / requested_budget if requested_budget > 0 else 0.0,
                "allocation_share": allocated_budget / requested_budget if requested_budget > 0 else np.nan,
                "support_limit_policy": str(score_rows["support_limit_policy"].iloc[0])
                if not score_rows.empty and "support_limit_policy" in score_rows
                else "unknown",
                "budget_mln_rub": _million(allocated_budget),
                "rto_p10_mln": _million(rec["total_effect_p10"]),
                "rto_p50_mln": _million(rec["total_effect_p50"]),
                "rto_p90_mln": _million(rec["total_effect_p90"]),
                "roas_p50": rec.get("roas_p50", ""),
                "roas_allocated_budget_p10": float(rec["total_effect_p10"]) / allocated_budget
                if allocated_budget > 0
                else np.nan,
                "roas_allocated_budget_p50": float(rec["total_effect_p50"]) / allocated_budget
                if allocated_budget > 0
                else np.nan,
                "roas_allocated_budget_p90": float(rec["total_effect_p90"]) / allocated_budget
                if allocated_budget > 0
                else np.nan,
                "roas_requested_budget_p10": float(rec["total_effect_p10"]) / requested_budget
                if requested_budget > 0
                else np.nan,
                "roas_requested_budget_p50": float(rec["total_effect_p50"]) / requested_budget
                if requested_budget > 0
                else np.nan,
                "roas_requested_budget_p90": float(rec["total_effect_p90"]) / requested_budget
                if requested_budget > 0
                else np.nan,
                "roas_denominator_kind": "requested_budget",
                "roas_denominator_budget_rub": requested_budget,
                "scenario_kind": "optimized_plan",
                "scenario_variant": "full_effect_maximizing",
                "scenario_feasibility_status": "feasible_full",
                "within_support_budget_rub": score_rows["within_support_budget_rub"].iloc[0]
                if not score_rows.empty and "within_support_budget_rub" in score_rows
                else np.nan,
                "within_support_share": score_rows["within_support_share"].iloc[0]
                if not score_rows.empty and "within_support_share" in score_rows
                else np.nan,
                "controlled_extrapolation_budget_rub": score_rows["controlled_extrapolation_budget_rub"].iloc[0]
                if not score_rows.empty and "controlled_extrapolation_budget_rub" in score_rows
                else np.nan,
                "controlled_extrapolation_share": score_rows["controlled_extrapolation_share"].iloc[0]
                if not score_rows.empty and "controlled_extrapolation_share" in score_rows
                else np.nan,
                "high_risk_budget_rub": score_rows["high_risk_budget_rub"].iloc[0]
                if not score_rows.empty and "high_risk_budget_rub" in score_rows
                else np.nan,
                "high_risk_share": score_rows["high_risk_share"].iloc[0]
                if not score_rows.empty and "high_risk_share" in score_rows
                else np.nan,
                "support_warnings_n": support_warnings,
                "elevated_support_warnings_n": elevated_support_warnings,
                "strong_support_warnings_n": strong_support_warnings,
                "hard_support_warnings_n": hard_support_warnings,
                "policy_violations_n": policy_violations,
                "policy_violation_codes": str(score_rows["policy_violation_codes"].iloc[0])
                if not score_rows.empty and "policy_violation_codes" in score_rows
                else "OK",
                "diagnostic_target_rows": _count_token(rec.get("allowed_use_counts", ""), "diagnostic")
                + _count_token(rec.get("risk_level_counts", ""), "high"),
                "caution_target_rows": _count_token(rec.get("allowed_use_counts", ""), "caution")
                + _count_token(rec.get("risk_level_counts", ""), "medium"),
                "objective_rows_n": objective_rows_n,
                "search_rank_raw": int(score_rows["search_rank_raw"].iloc[0])
                if not score_rows.empty and "search_rank_raw" in score_rows
                else np.nan,
                "search_rank_reliable": int(score_rows["search_rank_reliable"].iloc[0])
                if not score_rows.empty and "search_rank_reliable" in score_rows
                else np.nan,
                "quality_status": quality_status,
                "cell_support_status": _cell_support_status(quality_input),
                "optimizer_status": _optimizer_status(quality_input),
                "quality_explanation": _quality_explanation({**quality_input, "quality_status": quality_status}),
                "paired_delta_p10": _million(rec.get("paired_delta_p10"))
                if pd.notna(rec.get("paired_delta_p10"))
                else np.nan,
                "paired_delta_p50": _million(rec.get("paired_delta_p50"))
                if pd.notna(rec.get("paired_delta_p50"))
                else np.nan,
                "paired_delta_p90": _million(rec.get("paired_delta_p90"))
                if pd.notna(rec.get("paired_delta_p90"))
                else np.nan,
                "paired_probability_gt_zero": rec.get("paired_probability_gt_zero", np.nan),
                "paired_probability_gt_materiality": rec.get(
                    "paired_probability_gt_materiality", np.nan
                ),
                "paired_probability_noninferior": rec.get(
                    "paired_probability_noninferior", np.nan
                ),
                "paired_materiality_threshold": _million(rec.get("paired_materiality_threshold"))
                if pd.notna(rec.get("paired_materiality_threshold"))
                else np.nan,
                "paired_noninferiority_floor": _million(rec.get("paired_noninferiority_floor"))
                if pd.notna(rec.get("paired_noninferiority_floor"))
                else np.nan,
            }
        )
    represented_campaigns = {str(row["campaign_name"]) for row in rows}
    for campaign, tries in s6_scores.groupby("campaign_name", dropna=False):
        campaign = str(campaign)
        if campaign in represented_campaigns:
            continue
        statuses = tries.get("precheck_status", pd.Series(index=tries.index, dtype=str)).fillna("")
        no_modifiable = bool(statuses.eq("not_run_no_modifiable_cells").all())
        rejected_infeasible_n = int(statuses.eq("rejected_infeasible").sum())
        all_infeasible = rejected_infeasible_n == len(tries) and len(tries) > 0
        if not no_modifiable and not all_infeasible:
            continue
        status = tries.iloc[0]
        if all_infeasible:
            quality_status = "Требуется ручная проверка"
            quality_explanation = (
                "Scenario 6 не смог построить допустимый план внутри одновременно заданных "
                "support и business constraints. Это не означает, что исходную кампанию нельзя запускать."
            )
            support_decision = (
                "Автоматический план не сформирован; исходный медиаплан остается предметом ручного решения"
            )
            candidate_role = "S6 выполнен: допустимый вариант не найден"
            optimizer_status = "Только ручное распределение"
        else:
            quality_status = "Повышенная неопределенность"
            quality_explanation = (
                "Gate policy не разрешает менять доступные каналы. Прогноз исходного плана остается доступен, "
                "но Scenario 6 не имеет допустимой пары donor/receiver."
            )
            support_decision = "S6 не запускался: нет допустимой пары donor/receiver"
            candidate_role = "S6 недоступен для этой кампании"
            optimizer_status = "Перераспределение недоступно по gate policy"
        rows.append(
            {
                "campaign_name": campaign,
                "candidate_name": status["candidate_name"],
                "scenario6_ran": all_infeasible,
                "is_status_row": True,
                "attempts_total_n": int(len(tries)) if all_infeasible else 0,
                "attempts_rejected_by_support_n": rejected_infeasible_n,
                "attempts_rejected_by_policy_n": 0,
                "attempts_rejected_precheck_n": rejected_infeasible_n,
                "search_method": _full_list(tries["method"]) if all_infeasible else "Не запускался",
                "requested_budget_mln_rub": _million(status.get("total_budget_rub", 0.0)),
                "allocated_budget_mln_rub": np.nan,
                "unallocated_budget_mln_rub": np.nan,
                "allocated_budget_share": np.nan,
                "support_limit_policy": "precheck",
                "budget_mln_rub": _million(status.get("total_budget_rub", 0.0)),
                "rto_p10_mln": np.nan,
                "rto_p50_mln": np.nan,
                "roas_p50": np.nan,
                "roas_allocated_budget_p50": np.nan,
                "roas_requested_budget_p50": np.nan,
                "roas_denominator_kind": None,
                "roas_denominator_budget_rub": np.nan,
                "scenario_kind": "optimized_plan",
                "scenario_variant": "infeasible",
                "scenario_feasibility_status": "infeasible",
                "full_allocation_impossible_reason": str(status.get("precheck_reason") or ""),
                "support_warnings_n": rejected_infeasible_n,
                "elevated_support_warnings_n": 0,
                "strong_support_warnings_n": rejected_infeasible_n,
                "hard_support_warnings_n": rejected_infeasible_n,
                "policy_violations_n": 0,
                "policy_violation_codes": "OK",
                "diagnostic_target_rows": 1,
                "caution_target_rows": 0,
                "objective_rows_n": 0,
                "search_rank_raw": np.nan,
                "search_rank_reliable": np.nan,
                "quality_status": quality_status,
                "quality_explanation": quality_explanation,
                "cell_support_status": "Не оценено для Scenario 6",
                "optimizer_status": optimizer_status,
                "support_decision": support_decision,
                "candidate_role": candidate_role,
            }
        )
    df = pd.DataFrame(rows)
    if df.empty:
        return df
    df["quality_rank"] = df["quality_status"].map(_quality_rank)
    df["raw_rank"] = np.nan
    df["reliable_rank"] = np.nan
    ranked = df[df["scenario6_ran"].fillna(False) & ~df["is_status_row"].fillna(False)].copy()
    raw_order = ranked.sort_values(["campaign_name", "rto_p50_mln"], ascending=[True, False])
    df.loc[raw_order.index, "raw_rank"] = raw_order.groupby("campaign_name").cumcount() + 1
    reliable_order = ranked.sort_values(
            [
                "campaign_name",
                "hard_support_warnings_n",
                "policy_violations_n",
                "strong_support_warnings_n",
                "quality_rank",
                "unallocated_budget_mln_rub",
                "rto_p50_mln",
                "rto_p10_mln",
            ],
            ascending=[True, True, True, True, True, True, False, False],
    )
    df.loc[reliable_order.index, "reliable_rank"] = reliable_order.groupby("campaign_name").cumcount() + 1
    safe = df[
        df["scenario6_ran"].fillna(False)
        &
        df["hard_support_warnings_n"].eq(0)
        & df["policy_violations_n"].eq(0)
        & pd.to_numeric(
            df.get("high_risk_budget_rub", pd.Series(0.0, index=df.index)),
            errors="coerce",
        ).fillna(np.inf).le(0.000001)
        & df["objective_rows_n"].gt(0)
        & pd.to_numeric(df["unallocated_budget_mln_rub"], errors="coerce").fillna(np.inf).le(0.000001)
    ].copy()
    if not safe.empty:
        safe["safe_rank"] = (
            safe.sort_values(
                ["campaign_name", "rto_p50_mln", "rto_p10_mln", "controlled_extrapolation_budget_rub"],
                ascending=[True, False, False, True],
            )
            .groupby("campaign_name")
            .cumcount()
            + 1
        )
        df = df.merge(safe[["campaign_name", "candidate_name", "safe_rank"]], on=["campaign_name", "candidate_name"], how="left")
    else:
        df["safe_rank"] = np.nan
    df["is_best_raw_s6"] = df["raw_rank"].eq(1)
    df["is_best_safe_s6"] = df["safe_rank"].eq(1)
    def _support_decision(row: pd.Series) -> str:
        if int(row.get("hard_support_warnings_n") or 0) > 0:
            return "Вне robust observed upper: только ручное распределение"
        if int(row.get("policy_violations_n") or 0) > 0:
            return "Нарушает gate policy: не использовать для автоматического перераспределения"
        if int(row.get("strong_support_warnings_n") or 0) > 0:
            return "Контролируемое расширение между p99 и robust upper; допустимо с явной оговоркой"
        if float(row.get("unallocated_budget_mln_rub") or 0.0) > 0.000001:
            return "Безопасный частичный план; остаток бюджета требует ручного решения"
        if int(row.get("elevated_support_warnings_n") or 0) > 0:
            return "Автоматический план допустим с caveat: часть ячеек между p95 и p99"
        return "Автоматический план находится внутри p95 support-zone"

    default_support_decision = df.apply(_support_decision, axis=1)
    status_mask = df["is_status_row"].fillna(False)
    df["support_decision"] = df.get(
        "support_decision", pd.Series(pd.NA, index=df.index, dtype="object")
    ).where(status_mask, default_support_decision)
    df.loc[~df["scenario6_ran"].fillna(False) & ~status_mask, "support_decision"] = (
        "S6 не запускался: все доступные каналы зафиксированы gate policy"
    )

    def _role(row: pd.Series) -> str:
        if bool(row.get("is_status_row")):
            return str(row.get("candidate_role") or "Статус Scenario 6")
        if not bool(row.get("scenario6_ran")):
            return "S6 недоступен для этой кампании"
        roles: list[str] = []
        if bool(row.get("is_best_safe_s6")):
            roles.append("Лучший безопасный S6")
        if bool(row.get("is_best_raw_s6")):
            roles.append("Лучший raw по РТО")
        if roles:
            return " + ".join(roles)
        return "Дополнительная попытка"

    df["candidate_role"] = df.apply(_role, axis=1)
    return df.sort_values(
        ["campaign_name", "hard_support_warnings_n", "policy_violations_n", "strong_support_warnings_n", "quality_rank", "unallocated_budget_mln_rub", "rto_p50_mln", "rto_p10_mln"],
        ascending=[True, True, True, True, True, True, False, False],
    )


def _campaign_scale_assessment(
    *,
    model_input_budget_rub: float,
    active_dates: int,
    bounds: pd.Series | None,
) -> tuple[str, str]:
    if bounds is None or bounds.empty or active_dates <= 0:
        return (
            "Исторический benchmark недоступен",
            "Кампания рассчитана, но model package не содержит reviewed campaign-level support bounds.",
        )
    daily_intensity = model_input_budget_rub / active_dates
    budget_p95 = float(bounds.get("budget_p95") or 0.0)
    budget_p99 = float(bounds.get("budget_p99") or budget_p95)
    budget_upper = float(bounds.get("budget_robust_upper") or budget_p99)
    intensity_p95 = float(bounds.get("daily_intensity_p95") or 0.0)
    intensity_p99 = float(bounds.get("daily_intensity_p99") or intensity_p95)
    intensity_upper = float(bounds.get("daily_intensity_robust_upper") or intensity_p99)
    within_p95 = model_input_budget_rub <= budget_p95 and daily_intensity <= intensity_p95
    within_p99 = model_input_budget_rub <= budget_p99 and daily_intensity <= intensity_p99
    within_upper = model_input_budget_rub <= budget_upper and daily_intensity <= intensity_upper
    if within_p95:
        return (
            "Сопоставимо с историческими кампаниями",
            "И общий model-input бюджет, и средняя дневная интенсивность находятся внутри historical p95.",
        )
    if within_p99:
        return (
            "Крупная, но похожие кампании встречались",
            "Кампания выше обычного p95, но остается внутри historical p99 по бюджету и дневной интенсивности.",
        )
    if within_upper:
        return (
            "Очень крупная, нужна повышенная осторожность",
            "Масштаб выше p99, но не превышает robust observed upper. Прогноз допустим, автоувеличение ограничивается cell-level support.",
        )
    return (
        "Выше надежной наблюдаемой зоны",
        "Общий бюджет или дневная интенсивность выше robust observed campaign support. Это увеличивает неопределенность, но само по себе не является решением не запускать кампанию.",
    )


def _campaign_summary(
    flighting: pd.DataFrame,
    source_context: pd.DataFrame,
    support_bounds: pd.DataFrame,
) -> pd.DataFrame:
    df = flighting.copy()
    df["date"] = pd.to_datetime(df["date"]).dt.date
    out = (
        df.groupby("campaign_name", dropna=False)
        .agg(
            model_flighting_start=("date", "min"),
            model_flighting_end=("date", "max"),
            budget_mln_rub=("budget_rub", lambda s: float(s.sum()) / 1_000_000.0),
            directions=("segment", _full_list),
            channels=("channel", _full_list),
            geos_n=("geo", "nunique"),
            geos=("geo", _full_list),
            creatives=("creative_name", _full_list),
            rows_n=("campaign_name", "size"),
        )
        .reset_index()
    )
    if not source_context.empty and "campaign_name" in source_context:
        source = source_context.copy()
        for column in ["source_channels", "modeled_channels", "unmodeled_channels"]:
            if column in source:
                source[column] = source[column].map(
                    lambda value: ", ".join(str(item) for item in value)
                    if isinstance(value, list)
                    else str(value or "")
                )
        keep = [
            "campaign_name",
            "source_campaign_name",
            "source_sheet",
            "campaign_start",
            "campaign_end",
            "active_dates",
            "model_input_start",
            "model_input_end",
            "model_input_active_dates",
            "uploaded_budget_rub",
            "model_input_budget_rub",
            "unmodeled_budget_rub",
            "source_channels",
            "modeled_channels",
            "unmodeled_channels",
        ]
        source = source[[column for column in keep if column in source]].rename(
            columns={
                "campaign_start": "source_campaign_start",
                "campaign_end": "source_campaign_end",
                "active_dates": "source_active_dates",
            }
        )
        out = out.merge(source, on="campaign_name", how="left")
    out["campaign_start"] = pd.to_datetime(
        out.get("source_campaign_start", out["model_flighting_start"]),
        errors="coerce",
    ).dt.date
    out["campaign_start"] = out["campaign_start"].where(
        out["campaign_start"].notna(), out["model_flighting_start"]
    )
    out["campaign_end"] = pd.to_datetime(
        out.get("source_campaign_end", out["model_flighting_end"]),
        errors="coerce",
    ).dt.date
    out["campaign_end"] = out["campaign_end"].where(
        out["campaign_end"].notna(), out["model_flighting_end"]
    )
    if "model_input_start" not in out:
        out["model_input_start"] = out["model_flighting_start"]
    else:
        out["model_input_start"] = pd.to_datetime(
            out["model_input_start"], errors="coerce"
        ).dt.date.where(
            pd.to_datetime(out["model_input_start"], errors="coerce").notna(),
            out["model_flighting_start"],
        )
    if "model_input_end" not in out:
        out["model_input_end"] = out["model_flighting_end"]
    else:
        out["model_input_end"] = pd.to_datetime(
            out["model_input_end"], errors="coerce"
        ).dt.date.where(
            pd.to_datetime(out["model_input_end"], errors="coerce").notna(),
            out["model_flighting_end"],
        )
    for column, default in [
        ("uploaded_budget_rub", out["budget_mln_rub"] * 1_000_000.0),
        ("model_input_budget_rub", out["budget_mln_rub"] * 1_000_000.0),
        ("unmodeled_budget_rub", 0.0),
    ]:
        if column not in out:
            out[column] = default
        else:
            out[column] = pd.to_numeric(out[column], errors="coerce").fillna(default)
    if "source_active_dates" not in out:
        out["source_active_dates"] = (pd.to_datetime(out["campaign_end"]) - pd.to_datetime(out["campaign_start"])).dt.days + 1
    else:
        fallback_days = (pd.to_datetime(out["campaign_end"]) - pd.to_datetime(out["campaign_start"])).dt.days + 1
        out["source_active_dates"] = pd.to_numeric(out["source_active_dates"], errors="coerce").fillna(fallback_days).astype(int)
    out["uploaded_budget_mln_rub"] = out["uploaded_budget_rub"] / 1_000_000.0
    out["model_input_budget_mln_rub"] = out["model_input_budget_rub"] / 1_000_000.0
    out["unmodeled_budget_mln_rub"] = out["unmodeled_budget_rub"] / 1_000_000.0
    out["unmodeled_budget_share"] = np.where(
        out["uploaded_budget_rub"].gt(0),
        out["unmodeled_budget_rub"] / out["uploaded_budget_rub"],
        0.0,
    )
    out["calculation_status"] = np.where(
        out["unmodeled_budget_rub"].gt(1.0),
        "Рассчитано частично: часть каналов вне model package",
        "Рассчитано полностью",
    )
    bound_lookup = {
        str(row["segment"]): row
        for _, row in support_bounds.iterrows()
    } if not support_bounds.empty and "segment" in support_bounds else {}
    statuses: list[str] = []
    explanations: list[str] = []
    history_p95: list[float] = []
    history_p99: list[float] = []
    history_upper: list[float] = []
    for _, row in out.iterrows():
        bounds = bound_lookup.get(str(row["directions"]))
        status, explanation = _campaign_scale_assessment(
            model_input_budget_rub=float(row["model_input_budget_rub"]),
            active_dates=int(row["source_active_dates"]),
            bounds=bounds,
        )
        statuses.append(status)
        explanations.append(explanation)
        history_p95.append(_million(bounds.get("budget_p95")) if bounds is not None else np.nan)
        history_p99.append(_million(bounds.get("budget_p99")) if bounds is not None else np.nan)
        history_upper.append(_million(bounds.get("budget_robust_upper")) if bounds is not None else np.nan)
    out["campaign_scale_status"] = statuses
    out["campaign_scale_explanation"] = explanations
    out["historical_campaign_budget_p95_mln_rub"] = history_p95
    out["historical_campaign_budget_p99_mln_rub"] = history_p99
    out["historical_campaign_budget_robust_upper_mln_rub"] = history_upper
    return out


def _budget_detail(flighting: pd.DataFrame) -> pd.DataFrame:
    cells = _campaign_cells(flighting)
    cells["budget_mln_rub"] = cells["budget_rub"] / 1_000_000.0
    return cells[
        [
            "campaign_name",
            "segment",
            "start_date",
            "end_date",
            "creative_name",
            "channel",
            "geo",
            "budget_mln_rub",
            "active_days",
        ]
    ].sort_values(["campaign_name", "segment", "channel", "budget_mln_rub"], ascending=[True, True, True, False])


def _allocation_change_summary(allocation: pd.DataFrame) -> pd.DataFrame:
    if allocation.empty:
        return pd.DataFrame(columns=["campaign_name", "candidate_name", "moved_budget_mln_rub"])
    keys = ["segment", "geo", "channel"]
    rows: list[dict[str, Any]] = []
    for campaign, campaign_alloc in allocation.groupby("source_campaign_name", dropna=False):
        source = campaign_alloc[
            campaign_alloc["candidate_name"].astype(str).str.contains(
                "__scenario1_current_plan", na=False
            )
        ][keys + ["budget_rub"]].rename(columns={"budget_rub": "source_budget_rub"})
        if source.empty:
            continue
        for candidate_name, candidate in campaign_alloc.groupby("candidate_name", dropna=False):
            merged = source.merge(
                candidate[keys + ["budget_rub"]],
                on=keys,
                how="outer",
            ).fillna(0.0)
            moved = 0.5 * float(
                (merged["budget_rub"] - merged["source_budget_rub"]).abs().sum()
            )
            rows.append(
                {
                    "campaign_name": str(campaign),
                    "candidate_name": str(candidate_name),
                    "moved_budget_mln_rub": _million(moved),
                }
            )
    return pd.DataFrame(rows)


def _build_decision_pool(
    scenario_results: pd.DataFrame,
    scenario6: pd.DataFrame,
    campaign_summary: pd.DataFrame,
    allocation: pd.DataFrame,
    decision_policy: dict[str, Any] | None,
) -> pd.DataFrame:
    """Build one comparable row per benchmark plus the best-safe Scenario 6."""
    policy = _compile_decision_policy(decision_policy or {})
    rows: list[dict[str, Any]] = []
    for _, r in scenario_results.iterrows():
        rows.append(
            {
                "campaign_name": r["campaign_name"],
                "scenario_source": "Сценарии 1-5",
                "scenario_no": r["scenario_no"],
                "scenario_name": r["scenario_name"],
                "scenario_description": r.get("scenario_description", ""),
                "candidate_name": r["candidate_name"],
                "requested_budget_mln_rub": r.get("requested_budget_mln_rub", r.get("budget_mln_rub", np.nan)),
                "allocated_budget_mln_rub": r.get("allocated_budget_mln_rub", r.get("budget_mln_rub", np.nan)),
                "unallocated_budget_mln_rub": r.get("unallocated_budget_mln_rub", 0.0),
                "allocated_budget_share": r.get("allocated_budget_share", 1.0),
                "rto_p10_mln": r.get("rto_p10_mln", np.nan),
                "rto_p50_mln": r.get("rto_p50_mln", np.nan),
                "rto_p90_mln": r.get("rto_p90_mln", np.nan),
                "roas_p50": r.get(
                    "roas_requested_budget_p50",
                    r.get("rto_roas_p50", np.nan),
                ),
                "roas_allocated_budget_p10": r.get("roas_allocated_budget_p10", np.nan),
                "roas_allocated_budget_p50": r.get("roas_allocated_budget_p50", np.nan),
                "roas_allocated_budget_p90": r.get("roas_allocated_budget_p90", np.nan),
                "roas_requested_budget_p10": r.get("roas_requested_budget_p10", np.nan),
                "roas_requested_budget_p50": r.get("roas_requested_budget_p50", np.nan),
                "roas_requested_budget_p90": r.get("roas_requested_budget_p90", np.nan),
                "roas_denominator_kind": r.get("roas_denominator_kind"),
                "roas_denominator_budget_rub": r.get("roas_denominator_budget_rub", np.nan),
                "scenario_kind": r.get("scenario_kind", "benchmark_plan"),
                "scenario_variant": r.get("scenario_variant"),
                "scenario_feasibility_status": r.get("scenario_feasibility_status", "feasible_full"),
                "full_allocation_impossible_reason": r.get("full_allocation_impossible_reason", ""),
                "limiting_constraints": r.get("limiting_constraints", ""),
                "within_support_budget_rub": r.get("within_support_budget_rub", np.nan),
                "within_support_share": r.get("within_support_share", np.nan),
                "controlled_extrapolation_budget_rub": r.get("controlled_extrapolation_budget_rub", np.nan),
                "controlled_extrapolation_share": r.get("controlled_extrapolation_share", np.nan),
                "high_risk_budget_rub": r.get("high_risk_budget_rub", np.nan),
                "high_risk_share": r.get("high_risk_share", np.nan),
                "elevated_support_warnings_n": r.get("elevated_support_warnings_n", 0),
                "strong_support_warnings_n": r.get("strong_support_warnings_n", 0),
                "hard_support_warnings_n": r.get("hard_support_warnings_n", 0),
                "policy_violations_n": r.get("policy_violations_n", 0),
                "objective_rows_n": _objective_rows_count(r.get("rto_optimizer_use", "")),
                "calculation_status": r.get("calculation_status", "Рассчитано"),
                "cell_support_status": r.get("cell_support_status", ""),
                "optimizer_status": r.get("optimizer_status", ""),
                "quality_status": r.get("quality_status", ""),
                "quality_explanation": r.get("quality_explanation", ""),
                "paired_delta_p10": r.get("paired_delta_p10", np.nan),
                "paired_delta_p50": r.get("paired_delta_p50", np.nan),
                "paired_delta_p90": r.get("paired_delta_p90", np.nan),
                "paired_probability_gt_zero": r.get("paired_probability_gt_zero", np.nan),
                "paired_probability_noninferior": r.get("paired_probability_noninferior", np.nan),
            }
        )
    if not scenario6.empty and "is_best_safe_s6" in scenario6:
        for _, r in scenario6[scenario6["is_best_safe_s6"].fillna(False)].iterrows():
            rows.append(
                {
                    "campaign_name": r["campaign_name"],
                    "scenario_source": "Сценарий 6",
                    "scenario_no": "S06",
                    "scenario_name": "Сценарий 6. План максимального эффекта",
                    "scenario_description": "Лучший полный вариант по дополнительному обороту в пределах утвержденных ограничений риска.",
                    "candidate_name": r["candidate_name"],
                    "requested_budget_mln_rub": r.get("requested_budget_mln_rub", r.get("budget_mln_rub", np.nan)),
                    "allocated_budget_mln_rub": r.get("allocated_budget_mln_rub", r.get("budget_mln_rub", np.nan)),
                    "unallocated_budget_mln_rub": r.get("unallocated_budget_mln_rub", 0.0),
                    "allocated_budget_share": r.get("allocated_budget_share", 1.0),
                    "rto_p10_mln": r.get("rto_p10_mln", np.nan),
                    "rto_p50_mln": r.get("rto_p50_mln", np.nan),
                    "rto_p90_mln": r.get("rto_p90_mln", np.nan),
                    "roas_p50": r.get("roas_p50", np.nan),
                    "roas_allocated_budget_p10": r.get("roas_allocated_budget_p10", np.nan),
                    "roas_allocated_budget_p50": r.get("roas_allocated_budget_p50", np.nan),
                    "roas_allocated_budget_p90": r.get("roas_allocated_budget_p90", np.nan),
                    "roas_requested_budget_p10": r.get("roas_requested_budget_p10", np.nan),
                    "roas_requested_budget_p50": r.get("roas_requested_budget_p50", np.nan),
                    "roas_requested_budget_p90": r.get("roas_requested_budget_p90", np.nan),
                    "roas_denominator_kind": r.get("roas_denominator_kind"),
                    "roas_denominator_budget_rub": r.get("roas_denominator_budget_rub", np.nan),
                    "scenario_kind": "optimized_plan",
                    "scenario_variant": r.get("scenario_variant", "full_effect_maximizing"),
                    "scenario_feasibility_status": r.get("scenario_feasibility_status", "feasible_full"),
                    "within_support_budget_rub": r.get("within_support_budget_rub", np.nan),
                    "within_support_share": r.get("within_support_share", np.nan),
                    "controlled_extrapolation_budget_rub": r.get("controlled_extrapolation_budget_rub", np.nan),
                    "controlled_extrapolation_share": r.get("controlled_extrapolation_share", np.nan),
                    "high_risk_budget_rub": r.get("high_risk_budget_rub", np.nan),
                    "high_risk_share": r.get("high_risk_share", np.nan),
                    "elevated_support_warnings_n": r.get("elevated_support_warnings_n", 0),
                    "strong_support_warnings_n": r.get("strong_support_warnings_n", 0),
                    "hard_support_warnings_n": r.get("hard_support_warnings_n", 0),
                    "policy_violations_n": r.get("policy_violations_n", 0),
                    "objective_rows_n": r.get("objective_rows_n", 0),
                    "calculation_status": "Рассчитано частично"
                    if float(r.get("unallocated_budget_mln_rub") or 0.0) > 0.000001
                    else "Рассчитано",
                    "cell_support_status": r.get("cell_support_status", ""),
                    "optimizer_status": r.get("optimizer_status", ""),
                    "quality_status": r.get("quality_status", ""),
                    "quality_explanation": r.get("quality_explanation", ""),
                    "paired_delta_p10": r.get("paired_delta_p10", np.nan),
                    "paired_delta_p50": r.get("paired_delta_p50", np.nan),
                    "paired_delta_p90": r.get("paired_delta_p90", np.nan),
                    "paired_probability_gt_zero": r.get("paired_probability_gt_zero", np.nan),
                    "paired_probability_noninferior": r.get("paired_probability_noninferior", np.nan),
                    "search_converged": r.get("search_converged", False),
                    "search_budget_exhausted": r.get("search_budget_exhausted", False),
                }
            )
    pool = pd.DataFrame(rows)
    if pool.empty:
        return pool
    pool = pool.merge(
        _allocation_change_summary(allocation),
        on=["campaign_name", "candidate_name"],
        how="left",
    )
    pool["moved_budget_mln_rub"] = pd.to_numeric(
        pool.get("moved_budget_mln_rub"), errors="coerce"
    ).fillna(0.0)

    context_columns = [
        "campaign_name",
        "uploaded_budget_mln_rub",
        "model_input_budget_mln_rub",
        "unmodeled_budget_mln_rub",
        "unmodeled_budget_share",
        "unmodeled_channels",
        "campaign_scale_status",
        "campaign_scale_explanation",
    ]
    context = campaign_summary[[c for c in context_columns if c in campaign_summary]].copy()
    if not context.empty:
        pool = pool.merge(context, on="campaign_name", how="left", validate="many_to_one")
    pool["uploaded_budget_mln_rub"] = pd.to_numeric(
        pool.get("uploaded_budget_mln_rub", pool["requested_budget_mln_rub"]), errors="coerce"
    ).fillna(pool["requested_budget_mln_rub"])
    pool["model_input_budget_mln_rub"] = pd.to_numeric(
        pool.get("model_input_budget_mln_rub", pool["requested_budget_mln_rub"]), errors="coerce"
    ).fillna(pool["requested_budget_mln_rub"])
    pool["source_coverage_share"] = np.where(
        pool["uploaded_budget_mln_rub"].gt(0),
        pool["model_input_budget_mln_rub"] / pool["uploaded_budget_mln_rub"],
        0.0,
    )
    pool["effective_coverage_share"] = (
        pool["source_coverage_share"]
        * pd.to_numeric(pool["allocated_budget_share"], errors="coerce").fillna(0.0)
    )
    pool["quality_rank"] = pool["quality_status"].map(_quality_rank).fillna(9).astype(int)

    full_min = float(policy["reliability"]["full_coverage_min"])
    partial_min = float(policy["reliability"]["usable_partial_coverage_min"])

    def _reliability(row: pd.Series) -> tuple[int, str]:
        calculation = str(row.get("calculation_status") or "").lower()
        quality = str(row.get("quality_status") or "")
        coverage = float(row.get("effective_coverage_share") or 0.0)
        if "невозмож" in calculation or int(row.get("hard_support_warnings_n") or 0) > 0:
            return 5, "Не использовать автоматически"
        if int(row.get("policy_violations_n") or 0) > 0:
            return 5, "Нарушает ограничения модели"
        if (
            int(row.get("objective_rows_n") or 0) <= 0
            or quality == "Требуется ручная проверка"
        ):
            return 4, "Только ручная проверка"
        if coverage < partial_min:
            return 4, "Недостаточное покрытие бюджета"
        if coverage < full_min and int(row.get("elevated_support_warnings_n") or 0) == 0:
            return 1, "Надежный частичный план"
        if (
            int(row.get("elevated_support_warnings_n") or 0) > 0
            or int(row.get("strong_support_warnings_n") or 0) > 0
            or quality == "Повышенная неопределенность"
        ):
            return 2, "Допустимый план с оговорками"
        return 0, "Надежный полный план"

    reliability = pool.apply(_reliability, axis=1)
    pool["reliability_rank"] = [item[0] for item in reliability]
    pool["reliability_label"] = [item[1] for item in reliability]
    pool["uncertainty_width_share"] = np.where(
        pd.to_numeric(pool["rto_p50_mln"], errors="coerce").abs().gt(1e-9),
        (
            pd.to_numeric(pool["rto_p90_mln"], errors="coerce")
            - pd.to_numeric(pool["rto_p10_mln"], errors="coerce")
        )
        / pd.to_numeric(pool["rto_p50_mln"], errors="coerce").abs(),
        np.inf,
    )

    materiality = policy["materiality"]
    for campaign, index in pool.groupby("campaign_name").groups.items():
        group = pool.loc[index]
        current_rows = group[group["scenario_no"].eq("S01")]
        current = current_rows.iloc[0] if not current_rows.empty else group.iloc[0]
        baseline_p50 = float(current.get("rto_p50_mln") or 0.0)
        baseline_p10 = float(current.get("rto_p10_mln") or 0.0)
        min_gain = max(
            float(materiality["min_incremental_rto_gain_rub"]) / 1_000_000.0,
            abs(baseline_p50) * float(materiality["min_incremental_rto_gain_share"]),
        )
        min_moved = max(
            float(materiality["min_moved_budget_rub"]) / 1_000_000.0,
            float(current.get("model_input_budget_mln_rub") or 0.0)
            * float(materiality["min_moved_budget_share"]),
        )
        for row_index in index:
            delta = pool.at[row_index, "paired_delta_p50"]
            if pd.isna(delta):
                delta = float(pool.at[row_index, "rto_p50_mln"] or 0.0) - baseline_p50
                pool.at[row_index, "paired_delta_p50"] = delta
            relative = float(delta) / abs(baseline_p50) if abs(baseline_p50) > 1e-9 else np.nan
            probability = pool.at[row_index, "paired_probability_gt_zero"]
            probability = float(probability) if pd.notna(probability) else 0.0
            p10 = float(pool.at[row_index, "rto_p10_mln"] or 0.0)
            p10_floor = baseline_p10 - abs(baseline_p10) * float(
                materiality["max_p10_degradation_share"]
            )
            p10_noninferior = p10 >= p10_floor - 1e-12
            probability_noninferior = pool.at[row_index, "paired_probability_noninferior"]
            probability_noninferior = (
                float(probability_noninferior)
                if pd.notna(probability_noninferior)
                else float(p10_noninferior)
            )
            reliability_not_worse = int(pool.at[row_index, "reliability_rank"]) <= int(
                current["reliability_rank"]
            )
            is_current = str(pool.at[row_index, "scenario_no"]) == "S01"
            material = bool(
                not is_current
                and reliability_not_worse
                and float(delta) >= min_gain
                and pd.notna(relative)
                and float(relative) >= float(materiality["min_incremental_rto_gain_share"])
                and float(pool.at[row_index, "moved_budget_mln_rub"]) >= min_moved
                and probability >= float(materiality["min_positive_delta_probability"])
                and p10_noninferior
            )
            if is_current:
                status = "Базовый план"
            elif material:
                status = "Содержательная экономическая оптимизация"
            elif float(pool.at[row_index, "moved_budget_mln_rub"]) < min_moved:
                status = "План операционно не отличается от исходного"
            elif float(delta) < min_gain or pd.isna(relative) or float(relative) < float(
                materiality["min_incremental_rto_gain_share"]
            ):
                status = "Ожидаемый прирост слишком мал"
            elif probability < float(materiality["min_positive_delta_probability"]):
                status = "Преимущество недостаточно устойчиво по posterior"
            else:
                status = "Не проходит downside guardrail"
            pool.at[row_index, "relative_rto_gain"] = relative
            pool.at[row_index, "minimum_material_gain_mln_rub"] = min_gain
            pool.at[row_index, "minimum_moved_budget_mln_rub"] = min_moved
            pool.at[row_index, "p10_noninferior"] = p10_noninferior
            pool.at[row_index, "probability_noninferior"] = probability_noninferior
            pool.at[row_index, "economic_materiality_pass"] = material
            pool.at[row_index, "materiality_status"] = status

        champion_order = pool.loc[index].sort_values(
            [
                "reliability_rank",
                "hard_support_warnings_n",
                "policy_violations_n",
                "strong_support_warnings_n",
                "elevated_support_warnings_n",
                "effective_coverage_share",
                "uncertainty_width_share",
                "moved_budget_mln_rub",
                "scenario_no",
            ],
            ascending=[True, True, True, True, True, False, True, True, True],
        )
        pool.loc[index, "is_reliability_champion"] = False
        pool.at[champion_order.index[0], "is_reliability_champion"] = True
    pool["economic_materiality_pass"] = pool["economic_materiality_pass"].fillna(False).astype(bool)
    pool["is_reliability_champion"] = pool["is_reliability_champion"].fillna(False).astype(bool)
    return pool.sort_values(["campaign_name", "scenario_no"])


def _recommendations(
    scenario_results: pd.DataFrame,
    scenario6: pd.DataFrame,
    campaign_summary: pd.DataFrame,
    *,
    min_roas_p50: float | None,
    allocation: pd.DataFrame | None = None,
    decision_policy: dict[str, Any] | None = None,
    decision_pool: pd.DataFrame | None = None,
) -> pd.DataFrame:
    policy = _compile_decision_policy(decision_policy or {})
    pool = decision_pool if decision_pool is not None else _build_decision_pool(
        scenario_results,
        scenario6,
        campaign_summary,
        allocation if allocation is not None else pd.DataFrame(),
        policy,
    )
    if pool.empty:
        return pool
    scenario6_ran = (
        scenario6.groupby("campaign_name")["scenario6_ran"].any().to_dict()
        if not scenario6.empty and "scenario6_ran" in scenario6
        else {}
    )
    recommendations: list[pd.Series] = []
    full_min = float(policy["reliability"]["full_coverage_min"])
    partial_min = float(policy["reliability"]["usable_partial_coverage_min"])
    noninferiority_probability = float(
        policy["materiality"]["noninferiority_probability"]
    )
    for campaign, campaign_pool in pool.groupby("campaign_name", dropna=False):
        current_rows = campaign_pool[campaign_pool["scenario_no"].eq("S01")]
        current = (current_rows.iloc[0] if not current_rows.empty else campaign_pool.iloc[0]).copy()
        reliable = campaign_pool[campaign_pool["is_reliability_champion"]].iloc[0].copy()
        safe_s6_rows = campaign_pool[campaign_pool["scenario_no"].eq("S06")]
        safe_s6 = safe_s6_rows.iloc[0].copy() if not safe_s6_rows.empty else None
        chosen = current.copy()
        recommendation_type = "Оставить исходный план"
        allocation_decision = (
            "Автоматическое перераспределение не подтвердило надежного улучшения. "
            "Исходный план сохранен как точка отсчета и требует ручной проверки."
        )

        def _rank(value: Any) -> int:
            numeric = pd.to_numeric(pd.Series([value]), errors="coerce").iloc[0]
            return int(numeric) if pd.notna(numeric) else 9

        def _risk_improves(candidate: pd.Series) -> bool:
            return _rank(candidate.get("reliability_rank")) < _rank(
                current.get("reliability_rank")
            )

        def _no_material_downside(candidate: pd.Series) -> bool:
            return bool(candidate.get("p10_noninferior")) and float(
                candidate.get("probability_noninferior") or 0.0
            ) >= noninferiority_probability

        if (
            safe_s6 is not None
            and _risk_improves(safe_s6)
            and float(safe_s6.get("effective_coverage_share") or 0.0) >= full_min
            and _no_material_downside(safe_s6)
        ):
            chosen = safe_s6.copy()
            recommendation_type = "Перераспределить ради надежности"
            allocation_decision = (
                "S6 выбран не из-за микроскопического роста p50, а потому что уменьшает support/model risk "
                "без содержательного ухудшения downside."
            )
        elif safe_s6 is not None and bool(safe_s6.get("economic_materiality_pass")):
            chosen = safe_s6.copy()
            recommendation_type = "Перераспределить ради эффективности"
            allocation_decision = (
                "S6 проходит абсолютный, относительный, операционный, posterior и downside materiality gates."
            )
        elif (
            str(reliable.get("candidate_name")) != str(current.get("candidate_name"))
            and _risk_improves(reliable)
            and float(reliable.get("effective_coverage_share") or 0.0) >= full_min
            and _no_material_downside(reliable)
        ):
            chosen = reliable.copy()
            recommendation_type = "Перераспределить ради надежности"
            allocation_decision = (
                "Выбран наиболее надежный полный план: он лучше соответствует historical support "
                "и не создает содержательного downside против исходной схемы."
            )
        elif (
            _rank(current.get("reliability_rank")) >= 4
            and _risk_improves(reliable)
            and float(reliable.get("effective_coverage_share") or 0.0) >= partial_min
        ):
            chosen = reliable.copy()
            if float(reliable.get("effective_coverage_share") or 0.0) >= full_min:
                recommendation_type = "Надежный support-safe план"
                allocation_decision = (
                    "Исходный план выходит за надежную support-zone. Выбран полный медиаплан с меньшим "
                    "model/support risk; его более низкий прогноз нельзя трактовать как доказанную потерю "
                    "эффекта, потому что исходная оценка опирается на экстраполяцию."
                )
            else:
                recommendation_type = "Частичный безопасный план"
                allocation_decision = (
                    "Полного надежного плана нет. Показана безопасно рассчитываемая часть, остаток бюджета "
                    "возвращается маркетологу для ручного решения."
                )
        elif safe_s6 is not None:
            if float(safe_s6.get("paired_delta_p50") or 0.0) > 0.0:
                if bool(safe_s6.get("search_budget_exhausted")) or not bool(
                    safe_s6.get("search_converged")
                ):
                    allocation_decision = (
                        "S6 нашел математически более высокий вариант, но прирост не проходит materiality gate. "
                        "Поиск не подтвердил локальную сходимость в пределах заданного лимита; исходный план "
                        "сохранен как операционно эквивалентный, а не объявлен оптимальным."
                    )
                else:
                    allocation_decision = (
                        "S6 нашел математически более высокий вариант, но прирост не проходит materiality gate. "
                        "Исходный медиаплан сохранен как операционно эквивалентный, а не объявлен оптимальным."
                    )
            else:
                search_note = (
                    " Лимит поиска исчерпан, поэтому это не доказательство глобального optimum."
                    if bool(safe_s6.get("search_budget_exhausted"))
                    or not bool(safe_s6.get("search_converged"))
                    else " Поиск сошелся на минимальном заданном шаге, но это остается optimum внутри MMM и заданных ограничений."
                )
                allocation_decision = (
                    "Безопасный S6 не нашел ожидаемого улучшения против исходного плана; исходный медиаплан сохраняется."
                    + search_note
                )
        elif bool(scenario6_ran.get(campaign, False)):
            allocation_decision = (
                "Автоматический вариант не найден; исходный план сохранен для ручного решения."
            )
        else:
            allocation_decision = (
                "Scenario 6 недоступен: каналы зафиксированы gate policy. Это не является "
                "рекомендацией отменить кампанию."
            )

        chosen["recommendation_type"] = recommendation_type
        chosen["allocation_decision"] = allocation_decision
        chosen_scenario = str(chosen.get("scenario_no") or "")
        chosen_variant = str(chosen.get("scenario_variant") or "")
        if chosen_scenario == "S01":
            chosen["decision_status"] = "keep_uploaded_plan"
            chosen["review_status"] = "manual_review_required"
            if float(chosen.get("effective_coverage_share") or 0.0) < full_min:
                chosen["plan_status"] = "Полный медиаплан; частичное покрытие модели"
            else:
                chosen["plan_status"] = "Исходный план для ручной проверки"
        elif chosen_variant == "safe_partial" or float(
            chosen.get("unallocated_budget_mln_rub") or 0.0
        ) > 0.000001:
            chosen["decision_status"] = "no_safe_recommendation"
            chosen["review_status"] = "manual_review_required"
            chosen["plan_status"] = "Безопасно распределяемая часть; не рекомендация"
            chosen["allocation_decision"] = (
                "Весь бюджет нельзя распределить с приемлемой надежностью. "
                f"Безопасно распределить удалось только {float(chosen.get('allocated_budget_mln_rub') or 0.0):.2f} "
                f"из {float(chosen.get('requested_budget_mln_rub') or 0.0):.2f} млн рублей."
            )
        elif _rank(chosen.get("reliability_rank")) >= 4:
            chosen["decision_status"] = "manual_review_required"
            chosen["review_status"] = "manual_review_required"
            chosen["plan_status"] = "Требуется ручная проверка"
        elif float(chosen.get("effective_coverage_share") or 0.0) < full_min:
            chosen["decision_status"] = "manual_review_required"
            chosen["review_status"] = "manual_review_required"
            chosen["plan_status"] = "Полный медиаплан; частичное покрытие модели"
        else:
            chosen["decision_status"] = "recommended_reallocation"
            chosen["review_status"] = "not_required"
            chosen["plan_status"] = "Рекомендованный медиаплан"
        chosen["optimizer_available"] = safe_s6 is not None
        chosen["optimizer_status"] = (
            "Лучший безопасный S6 рассчитан"
            if safe_s6 is not None
            else "Автоматический план не найден"
            if bool(scenario6_ran.get(campaign, False))
            else "Перераспределение недоступно по gate policy"
        )
        chosen["reliable_scenario_no"] = reliable.get("scenario_no")
        chosen["reliable_scenario_name"] = reliable.get("scenario_name")
        chosen["reliable_candidate_name"] = reliable.get("candidate_name")
        chosen["reliable_rto_p10_mln"] = reliable.get("rto_p10_mln")
        chosen["reliable_rto_p50_mln"] = reliable.get("rto_p50_mln")
        chosen["reliable_rto_p90_mln"] = reliable.get("rto_p90_mln")
        chosen["reliable_coverage_share"] = reliable.get("effective_coverage_share")
        chosen["reliable_reliability_label"] = reliable.get("reliability_label")

        context_rows = campaign_summary[campaign_summary["campaign_name"].eq(campaign)]
        if not context_rows.empty:
            context = context_rows.iloc[0]
            for column in context.index:
                if column not in chosen or pd.isna(chosen.get(column)):
                    chosen[column] = context.get(column)
        roas = pd.to_numeric(
            pd.Series(
                [
                    chosen.get(
                        "roas_requested_budget_p50",
                        chosen.get("roas_p50"),
                    )
                ]
            ),
            errors="coerce",
        ).iloc[0]
        source_partial = (
            float(chosen.get("unmodeled_budget_mln_rub") or 0.0) > 0.000001
            or float(chosen.get("unallocated_budget_mln_rub") or 0.0) > 0.000001
        )
        optimizer_blocked = _rank(chosen.get("reliability_rank")) >= 4
        if min_roas_p50 is None:
            chosen["business_decision_status"] = "Не настроено: нужен бизнес-порог"
            chosen["campaign_decision"] = (
                "Система рекомендует распределение бюджета, но не решает запускать кампанию: "
                "бизнес-порог ROAS или contribution margin пока не утвержден."
            )
        elif optimizer_blocked or source_partial:
            chosen["business_decision_status"] = "Требуется ручное бизнес-решение"
            chosen["campaign_decision"] = (
                "ROAS-порог нельзя применить автоматически из-за model/support ограничений "
                "или неполного покрытия бюджета."
            )
        elif pd.isna(roas) or float(roas) < min_roas_p50:
            chosen["business_decision_status"] = "Ниже бизнес-порога"
            chosen["campaign_decision"] = f"Не проходит заданный порог ROAS {min_roas_p50:.2f}"
        else:
            chosen["business_decision_status"] = "Выше бизнес-порога"
            chosen["campaign_decision"] = f"Проходит заданный порог ROAS {min_roas_p50:.2f}"
        recommendations.append(chosen)
    return pd.DataFrame(recommendations).sort_values("campaign_name")


def _best_budget_plan(recommendations: pd.DataFrame, allocation: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for _, rec in recommendations.iterrows():
        campaign = rec["campaign_name"]
        candidate = rec["candidate_name"]
        sub = allocation[
            (allocation["source_campaign_name"] == campaign)
            & (allocation["candidate_name"] == candidate)
        ].copy()
        optimizer_available = bool(rec.get("optimizer_available"))
        uploaded_budget = float(
            rec.get("uploaded_budget_mln_rub")
            or rec.get("requested_budget_mln_rub")
            or rec.get("budget_mln_rub")
            or 0.0
        )
        for _, r in sub.iterrows():
            channel_policy = {
                "optimize": "Можно перераспределять",
                "no_increase": "Не увеличивать",
                "fixed_at_plan": "Сохранить как в исходном плане",
                "blocked": "Исключить",
            }.get(str(r.get("optimizer_policy") or ""), "Проверить вручную")
            if not optimizer_available:
                channel_policy = (
                    "Автооптимизация заблокирована; исходный бюджет показан для ручного решения"
                )
            rows.append(
                {
                    "campaign_name": campaign,
                    "plan_status": rec.get("plan_status", ""),
                    "chosen_scenario": rec["scenario_name"],
                    "candidate_name": candidate,
                    "direction": r["segment"],
                    "channel": r["channel"],
                    "geo": r["geo"],
                    "recommended_budget_mln_rub": _million(r["budget_rub"]),
                    "budget_share": _million(r["budget_rub"]) / uploaded_budget if uploaded_budget > 0 else 0.0,
                    "channel_policy": channel_policy,
                    "channel_gate_reasons": r.get("gate_reason_codes", ""),
                }
            )
        unallocated = float(rec.get("unallocated_budget_mln_rub") or 0.0)
        if unallocated > 0.000001:
            rows.append(
                {
                    "campaign_name": campaign,
                    "plan_status": rec.get("plan_status", ""),
                    "chosen_scenario": rec["scenario_name"],
                    "candidate_name": candidate,
                    "direction": "—",
                    "channel": "НЕРАСПРЕДЕЛЕННЫЙ ОСТАТОК",
                    "geo": "—",
                    "recommended_budget_mln_rub": unallocated,
                    "budget_share": unallocated / uploaded_budget if uploaded_budget > 0 else 0.0,
                    "channel_policy": "Вернуть маркетологу для ручного решения",
                    "channel_gate_reasons": "PARTIAL_SUPPORT_SAFE_PLAN",
                }
            )
        unmodeled = float(rec.get("unmodeled_budget_mln_rub") or 0.0)
        if unmodeled > 0.000001:
            unmodeled_channels = str(rec.get("unmodeled_channels") or "Неизвестный канал")
            rows.append(
                {
                    "campaign_name": campaign,
                    "plan_status": rec.get("plan_status", ""),
                    "chosen_scenario": rec["scenario_name"],
                    "candidate_name": candidate,
                    "direction": "—",
                    "channel": f"НЕ ПОКРЫТО МОДЕЛЬЮ: {unmodeled_channels}",
                    "geo": "—",
                    "recommended_budget_mln_rub": unmodeled,
                    "budget_share": unmodeled / uploaded_budget if uploaded_budget > 0 else 0.0,
                    "channel_policy": "Эффект не рассчитан; нужен отдельный ручной вывод",
                    "channel_gate_reasons": "SOURCE_CHANNEL_UNSUPPORTED_BY_MODEL_PACKAGE",
                }
            )
    out = pd.DataFrame(rows)
    if out.empty:
        return out
    return out.sort_values(
        ["campaign_name", "recommended_budget_mln_rub"], ascending=[True, False]
    )


def _optimizer_run_card(paths: ReportPaths) -> dict[str, Any]:
    if paths.run_id:
        safe_run_id = str(paths.run_id).strip().replace("/", "_").replace("::", "__").replace(" ", "_")
        card_path = paths.optimizer_output_dir / f"{safe_run_id}_optimizer_run_card.json"
    else:
        matches = sorted(paths.optimizer_output_dir.glob("*_optimizer_run_card.json"))
        card_path = matches[0] if matches else Path("__missing__")
    if not card_path.exists():
        return {}
    return json.loads(card_path.read_text(encoding="utf-8"))


def _sheet_name(index: int, campaign_name: str, used: set[str]) -> str:
    raw = str(campaign_name).strip()
    if " | " in raw:
        title, segment = raw.rsplit(" | ", 1)
        raw = f"{segment} {title}"
    clean = re.sub(r"[\\/*?:\[\]]+", " ", raw).strip()
    clean = re.sub(r"\s+", " ", clean)
    prefix = f"{index:02d}_"
    base = (prefix + clean)[:31] or f"{index:02d}_Кампания"
    value = base
    suffix = 2
    while value in used:
        tail = f"_{suffix}"
        value = base[: 31 - len(tail)] + tail
        suffix += 1
    used.add(value)
    return value


def _excel_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, np.generic):
        value = value.item()
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime()
    return value


def _write_title(ws: Any, row: int, title: str, max_col: int) -> None:
    max_col = max(max_col, 2)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    cell = ws.cell(row=row, column=1, value=title)
    cell.fill = PatternFill("solid", fgColor="17365D")
    cell.font = Font(color="FFFFFF", bold=True, size=12)
    cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[row].height = 24


def _write_table(
    ws: Any,
    frame: pd.DataFrame,
    *,
    start_row: int,
    title: str,
) -> tuple[int, int]:
    table = frame.copy()
    _write_title(ws, start_row, title, max(len(table.columns), 2))
    header_row = start_row + 1
    header_fill = PatternFill("solid", fgColor="2F6B6D")
    border = Border(bottom=Side(style="thin", color="B4C6E7"))
    for col_idx, column in enumerate(table.columns, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=str(column))
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[header_row].height = 34
    for row_offset, values in enumerate(table.itertuples(index=False, name=None), start=1):
        row_idx = header_row + row_offset
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=_excel_value(value))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=Side(style="hair", color="D9E2F3"))
            header = str(table.columns[col_idx - 1]).lower()
            if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                if "доля" in header or "вероят" in header or "%" in header:
                    cell.number_format = "0.0%"
                elif any(token in header for token in ["p10", "p50", "p90", "млн", "roas"]):
                    cell.number_format = "#,##0.000"
                elif any(token in header for token in ["кол-во", "попыт", "ячеек", "наруш"]):
                    cell.number_format = "#,##0"
                else:
                    cell.number_format = "#,##0.00"
            text = str(cell.value or "")
            if any(token in text for token in ["Надежный полный", "Сопоставимо с историей"]):
                cell.fill = PatternFill("solid", fgColor="E2F0D9")
            elif any(token in text for token in ["частичный", "оговорками", "Повышенная неопределенность"]):
                cell.fill = PatternFill("solid", fgColor="FFF2CC")
            elif any(token in text for token in ["ручная проверка", "Не использовать", "Расчет невозможен"]):
                cell.fill = PatternFill("solid", fgColor="FCE4D6")
        ws.row_dimensions[row_idx].height = 34
    return header_row + len(table) + 2, header_row


def _fit_wrapped_row_heights(ws: Any) -> None:
    """Give wrapped marketer-facing text enough vertical space to stay visible."""
    for row_idx in range(1, ws.max_row + 1):
        wrapped_lines = 1
        for cell in ws[row_idx]:
            if not cell.alignment.wrap_text or cell.value in (None, ""):
                continue
            column_width = ws.column_dimensions[get_column_letter(cell.column)].width or 12
            usable_characters = max(int(column_width * 0.9), 1)
            text_lines = sum(
                max(1, math.ceil(len(line) / usable_characters))
                for line in str(cell.value).splitlines() or [""]
            )
            wrapped_lines = max(wrapped_lines, text_lines)
        if wrapped_lines > 1:
            current_height = ws.row_dimensions[row_idx].height or 15
            estimated_height = min(18 * wrapped_lines + 6, 180)
            ws.row_dimensions[row_idx].height = max(current_height, estimated_height)


def _campaign_allocation_matrix(
    campaign_name: str,
    allocation: pd.DataFrame,
    decision_pool: pd.DataFrame,
    recommendation: pd.Series,
) -> pd.DataFrame:
    pool = decision_pool[decision_pool["campaign_name"].eq(campaign_name)].copy()
    candidate_to_scenario = dict(zip(pool["candidate_name"], pool["scenario_no"]))
    sub = allocation[
        allocation["source_campaign_name"].eq(campaign_name)
        & allocation["candidate_name"].isin(candidate_to_scenario)
    ].copy()
    if sub.empty:
        return pd.DataFrame()
    sub["scenario_no"] = sub["candidate_name"].map(candidate_to_scenario)
    sub["budget_mln_rub"] = pd.to_numeric(sub["budget_rub"], errors="coerce") / 1_000_000.0
    matrix = sub.pivot_table(
        index=["segment", "channel", "geo"],
        columns="scenario_no",
        values="budget_mln_rub",
        aggfunc="sum",
    ).reset_index()
    for code in ["S01", "S02", "S03", "S04", "S05", "S06"]:
        if code not in matrix:
            matrix[code] = np.nan
    chosen_code = str(recommendation.get("scenario_no") or "S01")
    matrix["Рекомендация"] = matrix.get(chosen_code, matrix["S01"])
    matrix["Δ к S01"] = matrix["Рекомендация"].fillna(0.0) - matrix["S01"].fillna(0.0)
    matrix = matrix.rename(
        columns={
            "segment": "Направление",
            "channel": "Канал",
            "geo": "Гео",
            "S01": "S01 Как загрузили",
            "S02": "S02 Ровно по связкам",
            "S03": "S03 Гео ровно",
            "S04": "S04 Каналы ровно",
            "S05": "S05 Support-safe",
            "S06": "S06 Best safe",
        }
    )
    scenario_columns = {
        "S01": "S01 Как загрузили",
        "S02": "S02 Ровно по связкам",
        "S03": "S03 Гео ровно",
        "S04": "S04 Каналы ровно",
        "S05": "S05 Support-safe",
        "S06": "S06 Best safe",
    }
    remainder = {"Направление": "Служебно", "Канал": "НЕРАСПРЕДЕЛЕНО", "Гео": "—"}
    for code, column in scenario_columns.items():
        rows = pool[pool["scenario_no"].eq(code)]
        remainder[column] = float(rows["unallocated_budget_mln_rub"].iloc[0]) if not rows.empty else np.nan
    chosen_rows = pool[pool["scenario_no"].eq(chosen_code)]
    remainder["Рекомендация"] = (
        float(chosen_rows["unallocated_budget_mln_rub"].iloc[0]) if not chosen_rows.empty else 0.0
    )
    remainder["Δ к S01"] = remainder["Рекомендация"] - float(remainder.get("S01 Как загрузили") or 0.0)
    matrix = pd.concat([matrix, pd.DataFrame([remainder])], ignore_index=True)
    unmodeled = float(recommendation.get("unmodeled_budget_mln_rub") or 0.0)
    if unmodeled > 0.000001:
        outside = {
            "Направление": "Служебно",
            "Канал": "ВНЕ MODEL PACKAGE",
            "Гео": str(recommendation.get("unmodeled_channels") or "—"),
            **{column: unmodeled for column in scenario_columns.values()},
            "Рекомендация": unmodeled,
            "Δ к S01": 0.0,
        }
        matrix = pd.concat([matrix, pd.DataFrame([outside])], ignore_index=True)
    return matrix[
        [
            "Направление",
            "Канал",
            "Гео",
            *scenario_columns.values(),
            "Рекомендация",
            "Δ к S01",
        ]
    ]


def _quality_matrix(decision_pool: pd.DataFrame, scenario6: pd.DataFrame) -> pd.DataFrame:
    rows = decision_pool.copy()
    if not scenario6.empty:
        present = set(zip(rows["campaign_name"], rows["scenario_no"]))
        for campaign, status in scenario6.groupby("campaign_name", dropna=False):
            if (campaign, "S06") in present:
                continue
            row = status.iloc[0]
            rows = pd.concat(
                [
                    rows,
                    pd.DataFrame(
                        [
                            {
                                "campaign_name": campaign,
                                "scenario_no": "S06",
                                "scenario_name": "Сценарий 6. Адаптивный поиск",
                                "effective_coverage_share": np.nan,
                                "rto_p10_mln": np.nan,
                                "rto_p50_mln": np.nan,
                                "rto_p90_mln": np.nan,
                                "elevated_support_warnings_n": 0,
                                "strong_support_warnings_n": row.get("strong_support_warnings_n", 0),
                                "hard_support_warnings_n": row.get("hard_support_warnings_n", 0),
                                "policy_violations_n": row.get("policy_violations_n", 0),
                                "reliability_label": "S6 недоступен",
                                "materiality_status": "Не оценено",
                                "quality_explanation": row.get("quality_explanation", ""),
                            }
                        ]
                    ),
                ],
                ignore_index=True,
            )
    keep = [
        "campaign_name",
        "scenario_no",
        "scenario_name",
        "effective_coverage_share",
        "rto_p10_mln",
        "rto_p50_mln",
        "rto_p90_mln",
        "elevated_support_warnings_n",
        "strong_support_warnings_n",
        "hard_support_warnings_n",
        "policy_violations_n",
        "reliability_label",
        "materiality_status",
        "quality_explanation",
    ]
    out = rows[[column for column in keep if column in rows]].copy()
    return out.rename(
        columns={
            "campaign_name": "Кампания",
            "scenario_no": "Сценарий",
            "scenario_name": "Название",
            "effective_coverage_share": "Покрытие бюджета",
            "rto_p10_mln": "Прирост РТО p10, млн руб.",
            "rto_p50_mln": "Прирост РТО p50, млн руб.",
            "rto_p90_mln": "Прирост РТО p90, млн руб.",
            "elevated_support_warnings_n": "Ячеек p95-p99",
            "strong_support_warnings_n": "Ячеек выше p99",
            "hard_support_warnings_n": "Ячеек вне robust upper",
            "policy_violations_n": "Нарушений model policy",
            "reliability_label": "Надежность",
            "materiality_status": "Содержательность изменения",
            "quality_explanation": "Почему",
        }
    ).sort_values(["Кампания", "Сценарий"])


def _write_dynamic_workbook(
    paths: ReportPaths,
    *,
    campaign_summary: pd.DataFrame,
    scenario6: pd.DataFrame,
    recommendations: pd.DataFrame,
    decision_pool: pd.DataFrame,
    allocation: pd.DataFrame,
    model_activation_status: str,
    production_blockers: list[str],
) -> None:
    paths.output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = "00_Итог_и_как_читать"
    summary_ws.sheet_view.showGridLines = False
    summary_ws.sheet_view.zoomScale = 85
    _write_title(summary_ws, 1, "MMM прогноз и рекомендация по медиаплану", 10)
    manual = pd.DataFrame(
        [
            ["Что считается", "Дополнительный РТО кампании против сценария без новой рекламы, а не полный товарооборот бизнеса."],
            ["p10 / p50 / p90", "p10 — осторожная граница, p50 — центральная оценка, p90 — верхняя граница при принятой модели."],
            ["Надежный сценарий", "Лучше всего соответствует historical support и model gates; он не обязан иметь максимальный p50."],
            ["Рекомендация", "Учитывает надежность, размер прироста, вероятность улучшения и операционный объем перелива бюджета."],
            ["Launch / cancel", "Не определяется автоматически, пока не утвержден ROAS или contribution-margin hurdle."],
            ["Статус модели", f"{model_activation_status}; blockers: {', '.join(production_blockers) if production_blockers else 'нет'}"],
        ],
        columns=["Понятие", "Как читать"],
    )
    next_row, _ = _write_table(summary_ws, manual, start_row=3, title="Короткая инструкция")

    decision_rows = []
    range_rows = []
    for _, rec in recommendations.iterrows():
        decision_rows.append(
            {
                "Кампания": rec["campaign_name"],
                "Загружено, млн руб.": rec.get("uploaded_budget_mln_rub"),
                "Покрыто моделью, млн руб.": rec.get("model_input_budget_mln_rub"),
                "Надежный ориентир": rec.get("reliable_scenario_name"),
                "Рекомендация": rec.get("scenario_name"),
                "Тип решения": rec.get("recommendation_type"),
                "Перемещено, млн руб.": rec.get("moved_budget_mln_rub"),
                "Почему": rec.get("allocation_decision"),
            }
        )
        range_rows.append(
            {
                "Кампания": rec["campaign_name"],
                "Надежный p10": rec.get("reliable_rto_p10_mln"),
                "Надежный p50": rec.get("reliable_rto_p50_mln"),
                "Надежный p90": rec.get("reliable_rto_p90_mln"),
                "Рекомендация p10": rec.get("rto_p10_mln"),
                "Рекомендация p50": rec.get("rto_p50_mln"),
                "Рекомендация p90": rec.get("rto_p90_mln"),
                "Δ p50 к S01, млн руб.": rec.get("paired_delta_p50"),
                "P(Δ > 0)": rec.get("paired_probability_gt_zero"),
                "Надежность": rec.get("reliability_label"),
                "Materiality": rec.get("materiality_status"),
            }
        )
    next_row, _ = _write_table(
        summary_ws,
        pd.DataFrame(decision_rows),
        start_row=next_row,
        title="Что выбрать",
    )
    _write_table(
        summary_ws,
        pd.DataFrame(range_rows),
        start_row=next_row,
        title="Диапазон прогноза и отличие от исходного плана",
    )

    used = {summary_ws.title, "99_Качество"}
    campaign_sheet_map: dict[str, str] = {}
    for index, campaign_name in enumerate(sorted(campaign_summary["campaign_name"].astype(str)), start=1):
        campaign_sheet_map[campaign_name] = _sheet_name(index, campaign_name, used)
    for campaign_name, sheet_name in campaign_sheet_map.items():
        ws = wb.create_sheet(sheet_name)
        ws.sheet_view.showGridLines = False
        ws.sheet_view.zoomScale = 80
        _write_title(ws, 1, campaign_name, 10)
        context = campaign_summary[campaign_summary["campaign_name"].eq(campaign_name)].iloc[0]
        geos_n = int(context.get("geos_n") or 0)
        passport = pd.DataFrame(
            [
                ["Период в источнике", f"{context.get('campaign_start')} — {context.get('campaign_end')}"],
                ["Период в модели", f"{context.get('model_input_start')} — {context.get('model_input_end')}"],
                ["Загруженный бюджет, млн руб.", context.get("uploaded_budget_mln_rub")],
                ["Покрыто моделью, млн руб.", context.get("model_input_budget_mln_rub")],
                ["Вне модели, млн руб.", context.get("unmodeled_budget_mln_rub")],
                ["Направление", context.get("directions")],
                ["Каналы", context.get("source_channels") or context.get("channels")],
                ["Гео", f"{geos_n}; полный список приведен в матрице медиапланов ниже"],
                ["Масштаб кампании", context.get("campaign_scale_status")],
                ["Статус расчета", context.get("calculation_status")],
            ],
            columns=["Параметр", "Значение"],
        )
        next_row, _ = _write_table(ws, passport, start_row=3, title="Описание кампании")
        campaign_pool = decision_pool[decision_pool["campaign_name"].eq(campaign_name)].copy()
        rec = recommendations[recommendations["campaign_name"].eq(campaign_name)].iloc[0]
        reliable = campaign_pool[campaign_pool["is_reliability_champion"]].iloc[0]
        safe_s6 = campaign_pool[campaign_pool["scenario_no"].eq("S06")]
        decisions = [
            {
                "Роль": "Наиболее надежный сценарий",
                "Сценарий": reliable.get("scenario_name"),
                "Размещено, млн руб.": reliable.get("allocated_budget_mln_rub"),
                "Остаток, млн руб.": reliable.get("unallocated_budget_mln_rub"),
                "p10": reliable.get("rto_p10_mln"),
                "p50": reliable.get("rto_p50_mln"),
                "p90": reliable.get("rto_p90_mln"),
                "ROAS p50": reliable.get("roas_p50"),
                "Надежность": reliable.get("reliability_label"),
                "Почему": reliable.get("quality_explanation"),
            }
        ]
        if not safe_s6.empty:
            s6 = safe_s6.iloc[0]
            decisions.append(
                {
                    "Роль": "Лучший безопасный S6",
                    "Сценарий": s6.get("scenario_name"),
                    "Размещено, млн руб.": s6.get("allocated_budget_mln_rub"),
                    "Остаток, млн руб.": s6.get("unallocated_budget_mln_rub"),
                    "p10": s6.get("rto_p10_mln"),
                    "p50": s6.get("rto_p50_mln"),
                    "p90": s6.get("rto_p90_mln"),
                    "ROAS p50": s6.get("roas_p50"),
                    "Надежность": s6.get("reliability_label"),
                    "Почему": s6.get("materiality_status"),
                }
            )
        decisions.append(
            {
                "Роль": "Что рекомендует система",
                "Сценарий": rec.get("scenario_name"),
                "Размещено, млн руб.": rec.get("allocated_budget_mln_rub"),
                "Остаток, млн руб.": rec.get("unallocated_budget_mln_rub"),
                "p10": rec.get("rto_p10_mln"),
                "p50": rec.get("rto_p50_mln"),
                "p90": rec.get("rto_p90_mln"),
                "ROAS p50": rec.get("roas_p50"),
                "Надежность": rec.get("reliability_label"),
                "Почему": rec.get("allocation_decision"),
            }
        )
        next_row, _ = _write_table(
            ws,
            pd.DataFrame(decisions),
            start_row=next_row,
            title="Главный вывод",
        )
        scenario_table = campaign_pool[
            [
                "scenario_no",
                "scenario_name",
                "allocated_budget_mln_rub",
                "unallocated_budget_mln_rub",
                "rto_p10_mln",
                "rto_p50_mln",
                "rto_p90_mln",
                "roas_p50",
                "paired_delta_p50",
                "paired_probability_gt_zero",
                "moved_budget_mln_rub",
                "effective_coverage_share",
                "reliability_label",
                "materiality_status",
                "quality_explanation",
            ]
        ].rename(
            columns={
                "scenario_no": "№",
                "scenario_name": "Сценарий",
                "allocated_budget_mln_rub": "Размещено, млн руб.",
                "unallocated_budget_mln_rub": "Остаток, млн руб.",
                "rto_p10_mln": "РТО p10, млн руб.",
                "rto_p50_mln": "РТО p50, млн руб.",
                "rto_p90_mln": "РТО p90, млн руб.",
                "roas_p50": "ROAS p50",
                "paired_delta_p50": "Δ p50 к S01, млн руб.",
                "paired_probability_gt_zero": "P(Δ > 0)",
                "moved_budget_mln_rub": "Перемещено, млн руб.",
                "effective_coverage_share": "Покрытие бюджета",
                "reliability_label": "Надежность",
                "materiality_status": "Содержательность",
                "quality_explanation": "Как читать",
            }
        )
        next_row, _ = _write_table(ws, scenario_table, start_row=next_row, title="Все сценарии")
        s6_rows = scenario6[scenario6["campaign_name"].eq(campaign_name)].copy()
        if not s6_rows.empty:
            safe_mask = s6_rows.get(
                "is_best_safe_s6", pd.Series(False, index=s6_rows.index)
            ).fillna(False)
            raw_mask = s6_rows.get(
                "is_best_raw_s6", pd.Series(False, index=s6_rows.index)
            ).fillna(False)
            status_mask = s6_rows.get(
                "is_status_row", pd.Series(False, index=s6_rows.index)
            ).fillna(False)
            visible = s6_rows[safe_mask | raw_mask | status_mask].copy()
            if visible.empty:
                visible = s6_rows.head(1).copy()
            s6_table = visible[
                [
                    c
                    for c in [
                        "candidate_role",
                        "attempts_total_n",
                        "candidate_plans_n",
                        "unique_allocations_n",
                        "effective_dimension_n",
                        "search_posterior_samples",
                        "smallest_transfer_mln_rub",
                        "search_converged",
                        "search_budget_exhausted",
                        "attempts_rejected_by_support_n",
                        "attempts_rejected_by_policy_n",
                        "rto_p10_mln",
                        "rto_p50_mln",
                        "rto_p90_mln",
                        "support_decision",
                        "quality_explanation",
                    ]
                    if c in visible
                ]
            ].rename(
                columns={
                    "candidate_role": "Результат S6",
                    "attempts_total_n": "Проверок переноса",
                    "candidate_plans_n": "Планов в финальном пуле",
                    "unique_allocations_n": "Уникальных планов найдено",
                    "effective_dimension_n": "Изменяемых geo x channel",
                    "search_posterior_samples": "Posterior draws поиска",
                    "smallest_transfer_mln_rub": "Мин. шаг переноса, млн руб.",
                    "search_converged": "Поиск сошелся",
                    "search_budget_exhausted": "Лимит поиска исчерпан",
                    "attempts_rejected_by_support_n": "Отброшено по support",
                    "attempts_rejected_by_policy_n": "Отброшено по model policy",
                    "rto_p10_mln": "РТО p10",
                    "rto_p50_mln": "РТО p50",
                    "rto_p90_mln": "РТО p90",
                    "support_decision": "Решение системы",
                    "quality_explanation": "Почему",
                }
            )
            next_row, _ = _write_table(ws, s6_table, start_row=next_row, title="Как работал Scenario 6")
        matrix = _campaign_allocation_matrix(campaign_name, allocation, decision_pool, rec)
        if not matrix.empty:
            _, allocation_header = _write_table(
                ws,
                matrix,
                start_row=next_row,
                title="Полный медиаплан geo x channel по всем сценариям",
            )
            ws.freeze_panes = f"D{allocation_header + 1}"
            ws.auto_filter.ref = (
                f"A{allocation_header}:{get_column_letter(len(matrix.columns))}{allocation_header + len(matrix)}"
            )

    quality_ws = wb.create_sheet("99_Качество")
    quality_ws.sheet_view.showGridLines = False
    quality_ws.sheet_view.zoomScale = 80
    _write_title(quality_ws, 1, "Качество расчетов по кампаниям и сценариям", 10)
    quality = _quality_matrix(decision_pool, scenario6)
    _, quality_header = _write_table(
        quality_ws,
        quality,
        start_row=3,
        title="Одна строка = одна кампания x сценарий",
    )
    quality_ws.freeze_panes = f"D{quality_header + 1}"
    quality_ws.auto_filter.ref = (
        f"A{quality_header}:{get_column_letter(len(quality.columns))}{quality_header + len(quality)}"
    )

    for ws in wb.worksheets:
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.orientation = "landscape"
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.page_margins.left = 0.25
        ws.page_margins.right = 0.25
        ws.page_margins.top = 0.4
        ws.page_margins.bottom = 0.4
        for column_index in range(1, ws.max_column + 1):
            letter = get_column_letter(column_index)
            max_length = max(
                (len(str(cell.value or "")) for cell in ws[letter][: min(ws.max_row, 160)]),
                default=10,
            )
            ws.column_dimensions[letter].width = min(max(max_length + 2, 12), 38)
        _fit_wrapped_row_heights(ws)
    wb.save(paths.output_xlsx)


def build_marketer_report(paths: ReportPaths) -> dict[str, Any]:
    """Build the workbook strictly from completed optimizer artifacts."""
    flighting = _read_csv(paths.flighting_path)
    candidate_scores, finalist_summary, allocation = _load_optimizer_tables(paths)
    if candidate_scores.empty or finalist_summary.empty or allocation.empty:
        raise FileNotFoundError(
            "Completed optimizer candidate/finalist/allocation artifacts are required before marketer report generation"
        )
    paired_comparisons = _load_paired_comparisons(paths)
    finalist_summary = _attach_paired_comparisons(finalist_summary, paired_comparisons)
    scenario_results = _result_rows_from_optimizer(finalist_summary, candidate_scores)
    scenario6 = _scenario6_summary(candidate_scores, finalist_summary)
    run_card = _optimizer_run_card(paths)
    min_roas_raw = ((run_card.get("business_guardrails") or {}).get("min_roas_p50"))
    min_roas_p50 = float(min_roas_raw) if min_roas_raw is not None else None
    decision_policy = _compile_decision_policy(run_card.get("decision_policy") or {})
    manifest_path = paths.model_run_dir / "model_manifest.json"
    manifest = json.loads(manifest_path.read_text(encoding="utf-8")) if manifest_path.exists() else {}

    source_context = _source_campaign_context(paths)
    support_bounds = _campaign_support_bounds(paths)
    campaign_summary = _campaign_summary(flighting, source_context, support_bounds)
    decision_pool = _build_decision_pool(
        scenario_results,
        scenario6,
        campaign_summary,
        allocation,
        decision_policy,
    )
    recommendations = _recommendations(
        scenario_results,
        scenario6,
        campaign_summary,
        min_roas_p50=min_roas_p50,
        allocation=allocation,
        decision_policy=decision_policy,
        decision_pool=decision_pool,
    )
    best_plan = _best_budget_plan(recommendations, allocation)
    _write_dynamic_workbook(
        paths,
        campaign_summary=campaign_summary,
        scenario6=scenario6,
        recommendations=recommendations,
        decision_pool=decision_pool,
        allocation=allocation,
        model_activation_status=str(manifest.get("activation_status") or "unknown"),
        production_blockers=list(manifest.get("production_blockers") or []),
    )

    scenario_results_path = paths.optimizer_output_dir / "marketer_report_scenario_results.csv"
    recommendations_path = paths.optimizer_output_dir / "marketer_report_recommendations.csv"
    best_plan_path = paths.optimizer_output_dir / "marketer_report_best_plan.csv"
    decision_pool_path = paths.optimizer_output_dir / "marketer_report_decision_pool.csv"
    scenario_results.to_csv(scenario_results_path, index=False)
    recommendations.to_csv(recommendations_path, index=False)
    best_plan.to_csv(best_plan_path, index=False)
    decision_pool.to_csv(decision_pool_path, index=False)
    workbook_check = load_workbook(paths.output_xlsx, read_only=True, data_only=True)
    workbook_sheet_names = list(workbook_check.sheetnames)
    workbook_check.close()

    card = {
        "output_xlsx": str(paths.output_xlsx),
        "campaigns_n": int(campaign_summary["campaign_name"].nunique()),
        "scenarios_1_5_rows": int(len(scenario_results)),
        "scenario6_rows": int(len(scenario6)),
        "recommendations_n": int(len(recommendations)),
        "decision_pool_rows": int(len(decision_pool)),
        "workbook_sheet_names": workbook_sheet_names,
        "forecast_recomputed_during_report": False,
        "calculation_source": "optimizer_finalist_artifacts",
        "source_adapter_campaigns_n": int(len(source_context)),
        "campaign_support_bounds_available": not support_bounds.empty,
        "model_activation_status": manifest.get("activation_status"),
        "production_blockers": list(manifest.get("production_blockers") or []),
        "scenario_results_csv": str(scenario_results_path),
        "recommendations_csv": str(recommendations_path),
        "best_plan_csv": str(best_plan_path),
        "decision_pool_csv": str(decision_pool_path),
        "decision_policy": decision_policy,
        "runtime_lineage": {
            "code_sha256": {
                "marketer_report.py": sha256_file(Path(__file__).resolve()),
            }
        },
        "source_artifact_sha256": {
            "flighting": sha256_file(paths.flighting_path),
            "optimizer_run_card": sha256_file(
                paths.optimizer_output_dir / f"{str(paths.run_id).strip().replace('/', '_').replace('::', '__').replace(' ', '_')}_optimizer_run_card.json"
            )
            if paths.run_id
            else None,
            "candidate_scores": sha256_file(
                paths.optimizer_output_dir / f"{str(paths.run_id).strip().replace('/', '_').replace('::', '__').replace(' ', '_')}_optimizer_candidate_scores.csv"
            )
            if paths.run_id
            else None,
            "finalist_summary": sha256_file(
                paths.optimizer_output_dir / f"{str(paths.run_id).strip().replace('/', '_').replace('::', '__').replace(' ', '_')}_optimizer_finalist_summary.csv"
            )
            if paths.run_id
            else None,
            "recommended_allocations": sha256_file(
                paths.optimizer_output_dir / f"{str(paths.run_id).strip().replace('/', '_').replace('::', '__').replace(' ', '_')}_optimizer_recommended_allocations.csv"
            )
            if paths.run_id
            else None,
            "paired_comparisons": sha256_file(
                paths.optimizer_output_dir / f"{str(paths.run_id).strip().replace('/', '_').replace('::', '__').replace(' ', '_')}_optimizer_paired_comparisons.csv"
            )
            if paths.run_id
            else None,
        },
        "output_sha256": {
            "xlsx": sha256_file(paths.output_xlsx),
            "scenario_results_csv": sha256_file(scenario_results_path),
            "recommendations_csv": sha256_file(recommendations_path),
            "best_plan_csv": sha256_file(best_plan_path),
            "decision_pool_csv": sha256_file(decision_pool_path),
        },
    }
    card_path = paths.optimizer_output_dir / "marketer_report_card.json"
    card_path.write_text(json.dumps(card, ensure_ascii=False, indent=2), encoding="utf-8")
    return card


def main() -> None:
    card = build_marketer_report(_parse_args())
    print(json.dumps(card, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
