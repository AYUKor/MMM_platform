"""Focused correctness tests for forecast, gates and optimizer contracts."""

from __future__ import annotations

import importlib.util
import json
import sys
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook

PYMC_CODE_DIR = Path(__file__).resolve().parents[1]
if str(PYMC_CODE_DIR) not in sys.path:
    sys.path.insert(0, str(PYMC_CODE_DIR))

from mmm_core.forecast_engine import (  # noqa: E402
    ForecastEngine,
    _CellResponseKernel,
    _adaptive_coordinate_refine,
    _assert_no_cross_campaign_overlap,
    _assert_modeled_spend_reconciles,
    _candidate_score,
    _candidate_from_allocation,
    _candidate_risk_budget_summary,
    _channel_balanced_candidate,
    _compile_decision_policy,
    _compile_optimizer_objective,
    _enforce_candidate_constraints,
    _geo_balanced_candidate,
    _generate_adaptive_scenario6_candidates,
    _generate_scenario5_candidates,
    _incremental_saturated_response,
    _incremental_saturated_response_draw_matrix,
    _greedy_marginal_allocation,
    _make_candidate_daily,
    _normalized_adstock,
    _normalized_adstock_draw_matrix,
    _objective_allowed_counts,
    _project_box_simplex,
    _project_proportional_box_simplex,
    _paired_candidate_comparisons,
    _reliable_candidate_sort_key,
    _risk_policy_violation_count,
    _support_weighted_candidate,
    CandidateFeasibilityError,
    summarize_forecast_detail,
)
from mmm_core.model_package import (  # noqa: E402
    DEFAULT_GATE_POLICY,
    evaluate_channel_gate,
    fit_runtime_provenance_issues,
    fit_risk_from_artifacts,
    sha256_file,
)
from mmm_core.model import run_model_refresh, validate_existing_model_run  # noqa: E402
from mmm_core.model_package_reader import StaleModelPackageError  # noqa: E402
from mmm_core.serving_semantics import (  # noqa: E402
    ServingSemanticsError,
    serving_model_inventory,
    validate_serving_model_inventory,
)

DATA_PIPELINE_PATH = Path(__file__).resolve().parents[3] / "00_Data" / "data_pipeline.py"
DATA_PIPELINE_SPEC = importlib.util.spec_from_file_location("x5_data_pipeline_contract", DATA_PIPELINE_PATH)
if DATA_PIPELINE_SPEC is None or DATA_PIPELINE_SPEC.loader is None:
    raise RuntimeError(f"Cannot load data pipeline module from {DATA_PIPELINE_PATH}")
DATA_PIPELINE = importlib.util.module_from_spec(DATA_PIPELINE_SPEC)
sys.modules[DATA_PIPELINE_SPEC.name] = DATA_PIPELINE
DATA_PIPELINE_SPEC.loader.exec_module(DATA_PIPELINE)

PANEL_PRIORS_PATH = Path(__file__).resolve().parents[1] / "01_panel_priors.py"
PANEL_PRIORS_SPEC = importlib.util.spec_from_file_location("x5_panel_priors_contract", PANEL_PRIORS_PATH)
if PANEL_PRIORS_SPEC is None or PANEL_PRIORS_SPEC.loader is None:
    raise RuntimeError(f"Cannot load panel-priors module from {PANEL_PRIORS_PATH}")
PANEL_PRIORS = importlib.util.module_from_spec(PANEL_PRIORS_SPEC)
sys.modules[PANEL_PRIORS_SPEC.name] = PANEL_PRIORS
PANEL_PRIORS_SPEC.loader.exec_module(PANEL_PRIORS)

MARKETER_REPORT_PATH = Path(__file__).resolve().parents[2] / "02_Budget_optimizer/marketer_report.py"
MARKETER_REPORT_SPEC = importlib.util.spec_from_file_location("x5_marketer_report_contract", MARKETER_REPORT_PATH)
if MARKETER_REPORT_SPEC is None or MARKETER_REPORT_SPEC.loader is None:
    raise RuntimeError(f"Cannot load marketer-report module from {MARKETER_REPORT_PATH}")
MARKETER_REPORT = importlib.util.module_from_spec(MARKETER_REPORT_SPEC)
sys.modules[MARKETER_REPORT_SPEC.name] = MARKETER_REPORT
MARKETER_REPORT_SPEC.loader.exec_module(MARKETER_REPORT)


class _PackageStub:
    activation_status = "preprod_restricted"
    manifest: dict = {}


def _engine_for_support() -> ForecastEngine:
    denominators = pd.DataFrame(
        [
            {
                "segment": "S",
                "geo_label": "G",
                "date": "2026-01-01",
                "population_k": 1.0,
                "unique_users": 1.0,
                "orders_cnt": 1.0,
                "market_size_tier": "small",
            }
        ]
    )
    support = pd.DataFrame(
        [
            {
                "fit_key": "S::turnover_per_user",
                "segment": "S",
                "target": "turnover_per_user",
                "channel": "C",
                "geo_label": "G",
                "daily_spend_p95_rub": 100.0,
                "daily_spend_p99_rub": 150.0,
                "daily_spend_max_rub": 220.0,
                "active_days": 100,
            }
        ]
    )
    capability = pd.DataFrame(
        [
            {
                "segment": "S",
                "target": "turnover_per_user",
                "channel": "C",
                "allowed_use": "primary",
                "optimizer_use": "optimize",
            }
        ]
    )
    return ForecastEngine(
        run_dir=Path("."),
        package=_PackageStub(),
        metadata={},
        media_scales=pd.DataFrame(),
        denominators=denominators,
        support_bounds=support,
        warm_start=pd.DataFrame(columns=["fit_key", "geo_label", "channel", "as_of_date", "lag", "scaled_spend"]),
        capability=capability,
    )


class GatePolicyTests(unittest.TestCase):
    def test_fit_provenance_uses_recorded_hash_not_mutable_source(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            run_dir = Path(tmp)
            recorded_hash = "a" * 64
            (run_dir / "fit_contract_S__T.json").write_text(
                json.dumps({"fit_code_sha256": recorded_hash}),
                encoding="utf-8",
            )
            config = {"fit_runtime_version": "1.0.0", "fit_code_sha256": recorded_hash}
            self.assertEqual(fit_runtime_provenance_issues(run_dir, config), [])

            (run_dir / "fit_contract_S__T.json").write_text(
                json.dumps({"fit_code_sha256": "b" * 64}),
                encoding="utf-8",
            )
            self.assertIn(
                "FIT_CONTRACT_CODE_HASH_MISMATCH",
                fit_runtime_provenance_issues(run_dir, config),
            )

    def test_upstream_diagnostic_can_never_be_optimizer_objective(self) -> None:
        gate = evaluate_channel_gate(
            target="turnover_per_user",
            fit_allowed="primary",
            upstream_row={
                "roas_use_case": "diagnostic_only",
                "quality_flags": "OK",
                "active_days": "200",
                "active_geos": "100",
                "pct_nonzero_rows": "30",
            },
            reliability_flags="OK",
            contraction={"rows_n": 1.0, "high_contraction_share": 1.0},
            gate_policy=DEFAULT_GATE_POLICY,
        )
        self.assertEqual(gate["allowed_use"], "diagnostic")
        self.assertEqual(gate["optimizer_policy"], "fixed_at_plan")
        self.assertEqual(gate["objective_role"], "side_metric_only")
        self.assertEqual(_objective_allowed_counts("fixed_at_plan:12"), 0)

    def test_posterior_expanded_reportable_effect_is_caution(self) -> None:
        gate = evaluate_channel_gate(
            target="turnover_per_user",
            fit_allowed="primary",
            upstream_row={
                "roas_use_case": "reportable",
                "quality_flags": "OK",
                "active_days": "200",
                "active_geos": "100",
                "pct_nonzero_rows": "30",
            },
            reliability_flags="OK",
            contraction={"rows_n": 1.0, "posterior_expanded_share": 1.0},
            gate_policy=DEFAULT_GATE_POLICY,
        )
        self.assertEqual(gate["allowed_use"], "caution")
        self.assertEqual(gate["optimizer_policy"], "no_increase")
        self.assertIn("POSTERIOR_EXPANDED_EFFECT", gate["gate_reason_codes"])

    def test_missing_contraction_evidence_fails_closed(self) -> None:
        gate = evaluate_channel_gate(
            target="turnover_per_user",
            fit_allowed="primary",
            upstream_row={
                "roas_use_case": "reportable",
                "quality_flags": "OK",
                "active_days": "200",
                "active_geos": "100",
                "pct_nonzero_rows": "30",
            },
            reliability_flags="OK",
            contraction={},
            gate_policy=DEFAULT_GATE_POLICY,
        )
        self.assertEqual(gate["allowed_use"], "diagnostic")
        self.assertIn("MISSING_CONTRACTION_EVIDENCE", gate["gate_reason_codes"])

    def test_fixed_saturation_shape_is_never_primary(self) -> None:
        gate = evaluate_channel_gate(
            target="turnover_per_user",
            fit_allowed="primary",
            upstream_row={
                "roas_use_case": "reportable",
                "quality_flags": "OK",
                "active_days": "200",
                "active_geos": "100",
                "pct_nonzero_rows": "30",
            },
            reliability_flags="OK",
            contraction={"rows_n": 1.0, "high_contraction_share": 1.0},
            gate_policy=DEFAULT_GATE_POLICY,
            fixed_response_shape=True,
        )
        self.assertEqual(gate["allowed_use"], "caution")
        self.assertEqual(gate["optimizer_policy"], "no_increase")
        self.assertTrue(gate["fixed_saturation_shape"])
        self.assertIn("FIXED_SATURATION_SHAPE", gate["gate_reason_codes"])

    def test_orders_are_always_diagnostic(self) -> None:
        gate = evaluate_channel_gate(
            target="orders_per_user",
            fit_allowed="primary",
            upstream_row={
                "roas_use_case": "reportable",
                "quality_flags": "OK",
                "active_days": "200",
                "active_geos": "100",
                "pct_nonzero_rows": "30",
            },
            reliability_flags="OK",
            contraction={"rows_n": 1.0, "high_contraction_share": 1.0},
            gate_policy=DEFAULT_GATE_POLICY,
        )
        self.assertEqual(gate["allowed_use"], "diagnostic")
        self.assertEqual(gate["optimizer_policy"], "fixed_at_plan")
        self.assertEqual(gate["objective_role"], "side_metric_only")
        self.assertIn("TARGET_POLICY_DIAGNOSTIC_ONLY", gate["gate_reason_codes"])

    def test_run_override_cannot_promote_orders(self) -> None:
        malicious_policy = {
            **DEFAULT_GATE_POLICY,
            "target_rules": {
                "orders_per_user": {
                    "minimum_allowed_use": "primary",
                    "reason_code": "IGNORE_MANDATORY_POLICY",
                }
            },
        }
        gate = evaluate_channel_gate(
            target="orders_per_user",
            fit_allowed="primary",
            upstream_row={
                "roas_use_case": "reportable",
                "quality_flags": "OK",
                "active_days": "200",
                "active_geos": "100",
                "pct_nonzero_rows": "30",
            },
            reliability_flags="OK",
            contraction={"rows_n": 1.0, "high_contraction_share": 1.0},
            gate_policy=malicious_policy,
        )
        self.assertEqual(gate["allowed_use"], "diagnostic")
        self.assertIn("TARGET_POLICY_DIAGNOSTIC_ONLY", gate["gate_reason_codes"])

    def test_run_override_cannot_upgrade_diagnostic_to_optimize(self) -> None:
        malicious_policy = {
            **DEFAULT_GATE_POLICY,
            "optimizer_actions": {**DEFAULT_GATE_POLICY["optimizer_actions"], "diagnostic": "optimize"},
            "forecast_actions": {**DEFAULT_GATE_POLICY["forecast_actions"], "diagnostic": "allowed"},
        }
        gate = evaluate_channel_gate(
            target="turnover_per_user",
            fit_allowed="primary",
            upstream_row={
                "roas_use_case": "diagnostic_only",
                "quality_flags": "OK",
                "active_days": "200",
                "active_geos": "100",
                "pct_nonzero_rows": "30",
            },
            reliability_flags="OK",
            contraction={"rows_n": 1.0, "high_contraction_share": 1.0},
            gate_policy=malicious_policy,
        )
        self.assertEqual(gate["optimizer_policy"], "fixed_at_plan")
        self.assertEqual(gate["forecast_policy"], "diagnostic_only")

    def test_missing_per_fit_evidence_fails_closed(self) -> None:
        risk_level, allowed_use, risks = fit_risk_from_artifacts(
            "S::turnover_per_user",
            diagnostics={},
            adequacy={},
            posterior_index={"S::turnover_per_user": {"file_name": "posterior.nc"}},
        )
        self.assertEqual(risk_level, "high")
        self.assertEqual(allowed_use, "diagnostic")
        self.assertIn("MISSING_FIT_DIAGNOSTICS", {row["risk_type"] for row in risks})
        self.assertIn("MISSING_FIT_ADEQUACY", {row["risk_type"] for row in risks})

    def test_ppc_r2_review_band_boundaries(self) -> None:
        diagnostics = {
            ("S::turnover_per_user",): {
                "status": "OK",
                "rhat_max": "1.0",
                "ess_bulk_min": "1000",
                "n_divergences": "0",
            }
        }
        posterior = {"S::turnover_per_user": {"file_name": "posterior.nc"}}
        expectations = [
            (0.1999, "diagnostic"),
            (0.20, "caution"),
            (0.2999, "caution"),
            (0.30, "primary"),
        ]
        for value, expected in expectations:
            _, allowed_use, risks = fit_risk_from_artifacts(
                "S::turnover_per_user",
                diagnostics,
                {"S::turnover_per_user": {"r2_mean": value}},
                posterior,
                DEFAULT_GATE_POLICY,
            )
            self.assertEqual(allowed_use, expected)
            if expected == "caution":
                self.assertIn("PPC_R2_MANUAL_REVIEW_BAND", {row["risk_type"] for row in risks})


class ForecastMathTests(unittest.TestCase):
    def test_vectorized_replay_adstock_matches_serving_scalar_path(self) -> None:
        x = np.array([10.0, 0.0, 30.0, 5.0, 0.0])
        alpha = np.array([0.0, 0.35, 0.8])
        actual = _normalized_adstock_draw_matrix(x, alpha, l_max=3)
        expected = np.vstack([_normalized_adstock(x, value, l_max=3) for value in alpha])
        np.testing.assert_allclose(actual, expected, rtol=0.0, atol=1e-12)

    def test_vectorized_adstock_matches_scalar_path_with_warm_start(self) -> None:
        x = np.array([10.0, 0.0, 30.0, 5.0, 0.0])
        alpha = np.array([0.0, 0.35, 0.8])
        warm_start = np.array([7.0, 11.0, 13.0])
        actual = _normalized_adstock_draw_matrix(x, alpha, l_max=3, warm_start=warm_start)
        expected = np.vstack(
            [_normalized_adstock(x, value, l_max=3, warm_start=warm_start) for value in alpha]
        )
        np.testing.assert_allclose(actual, expected, rtol=0.0, atol=1e-12)

    def test_vectorized_incremental_response_matches_scalar_path(self) -> None:
        x = np.array([0.1, 0.0, 0.3, 0.05])
        alpha = np.array([0.2, 0.5, 0.85])
        lam = np.array([0.7, 1.1, 1.8])
        warm_start = np.array([0.4, 0.2, 0.1])
        actual = _incremental_saturated_response_draw_matrix(
            x,
            alpha_values=alpha,
            lam_values=lam,
            l_max=3,
            warm_start=warm_start,
        )
        expected = np.vstack(
            [
                _incremental_saturated_response(
                    x,
                    alpha=float(alpha_value),
                    lam=float(lam_value),
                    l_max=3,
                    warm_start=warm_start,
                )
                for alpha_value, lam_value in zip(alpha, lam)
            ]
        )
        np.testing.assert_allclose(actual, expected, rtol=0.0, atol=1e-12)

    def test_campaign_quantiles_are_computed_after_summing_draws(self) -> None:
        common = {
            "campaign_name": "campaign",
            "segment": "S",
            "target": "turnover_per_user",
            "channel": "C",
            "spend_rub": 50.0,
            "effect_unit": "rub_per_user",
            "total_effect_unit": "incremental_turnover_rub",
            "allowed_use": "primary",
            "optimizer_use": "optimize",
            "risk_level": "low",
            "support_flags": "OK",
            "_effect_unit_weight": 1.0,
        }
        detail = pd.DataFrame(
            [
                {**common, "geo": "G1", "_total_effect_draws": np.array([0.0, 100.0]), "_effect_unit_draws": np.array([0.0, 100.0])},
                {**common, "geo": "G2", "_total_effect_draws": np.array([100.0, 0.0]), "_effect_unit_draws": np.array([100.0, 0.0])},
            ]
        )
        summary = summarize_forecast_detail(detail)
        total = summary[summary["channel"].eq("__TOTAL__")].iloc[0]
        self.assertAlmostEqual(float(total["total_effect_p10"]), 100.0)
        self.assertAlmostEqual(float(total["total_effect_p50"]), 100.0)
        self.assertAlmostEqual(float(total["total_effect_p90"]), 100.0)

    def test_support_boundary_uses_tolerance(self) -> None:
        engine = _engine_for_support()
        within = engine._support_assessment("S::turnover_per_user", "G", "C", np.array([100.0]))
        elevated = engine._support_assessment("S::turnover_per_user", "G", "C", np.array([140.0]))
        strong = engine._support_assessment("S::turnover_per_user", "G", "C", np.array([180.0]))
        near_boundary = engine._support_flags("S::turnover_per_user", "G", "C", np.array([220.001]))
        material_breach = engine._support_flags("S::turnover_per_user", "G", "C", np.array([220.02]))
        self.assertEqual(within.level, "within_support")
        self.assertEqual(elevated.level, "elevated_p95_p99")
        self.assertEqual(strong.level, "strong_p99_robust_upper")
        self.assertNotIn("FUTURE_DAILY_SPEND_GT_ROBUST_HIST_UPPER", near_boundary)
        self.assertIn("FUTURE_DAILY_SPEND_GT_ROBUST_HIST_UPPER", material_breach)

    def test_short_warm_start_is_left_padded_not_wrapped(self) -> None:
        actual = _normalized_adstock(np.array([0.0]), alpha=0.5, l_max=3, warm_start=np.array([4.0]))
        expected = (0.5 * 4.0) / (1.0 + 0.5 + 0.25 + 0.125)
        self.assertAlmostEqual(float(actual[0]), expected)

    def test_warm_history_is_present_in_both_sides_of_counterfactual(self) -> None:
        response = _incremental_saturated_response(
            np.zeros(3),
            alpha=0.7,
            lam=1.2,
            l_max=3,
            warm_start=np.array([4.0, 3.0, 2.0]),
        )
        np.testing.assert_allclose(response, np.zeros(3), atol=1e-12)

    def test_configured_analog_year_is_used_instead_of_previous_year(self) -> None:
        engine = _engine_for_support()
        rows = []
        for year, users in [(2025, 100.0), (2026, 250.0)]:
            rows.append(
                {
                    "segment": "S",
                    "geo_label": "G",
                    "date": f"{year}-01-01",
                    "population_k": 1.0,
                    "unique_users": users,
                    "orders_cnt": 1.0,
                    "market_size_tier": "small",
                }
            )
        engine.denominators = pd.DataFrame(rows)
        engine.__post_init__()
        configured = engine._denominator_for("S", "G", pd.Timestamp("2027-01-01").date(), analog_year=2025)
        previous_year = engine._denominator_for("S", "G", pd.Timestamp("2027-01-01").date())
        self.assertEqual(configured["unique_users"], 100.0)
        self.assertEqual(previous_year["unique_users"], 250.0)

    def test_new_geo_can_use_explicit_same_geo_nearest_year_policy(self) -> None:
        engine = _engine_for_support()
        value = engine._denominator_for(
            "S",
            "G",
            pd.Timestamp("2027-01-01").date(),
            analog_year=2025,
            missing_geo_policy="nearest_available_year_same_geo",
        )
        self.assertEqual(value["denominator_analog_year_used"], 2026)
        self.assertEqual(value["denominator_fallback_years"], 1)

    def test_overlapping_independent_campaigns_fail_closed(self) -> None:
        plan = pd.DataFrame(
            [
                {"campaign_name": "A", "segment": "S", "geo": "G", "channel": "C", "date": "2026-01-01", "budget_rub": 10.0},
                {"campaign_name": "B", "segment": "S", "geo": "G", "channel": "C", "date": "2026-01-01", "budget_rub": 20.0},
            ]
        )
        with self.assertRaises(ValueError):
            _assert_no_cross_campaign_overlap(plan)

    def test_multi_segment_total_sums_draws_before_quantiles(self) -> None:
        common = {
            "campaign_name": "campaign",
            "target": "turnover_per_user",
            "channel": "C",
            "geo": "G",
            "spend_rub": 50.0,
            "effect_unit": "rub_per_user",
            "total_effect_unit": "incremental_turnover_rub",
            "allowed_use": "primary",
            "optimizer_use": "optimize",
            "risk_level": "low",
            "support_flags": "OK",
            "_effect_unit_weight": 1.0,
        }
        detail = pd.DataFrame(
            [
                {**common, "segment": "S1", "_total_effect_draws": np.array([0.0, 100.0]), "_effect_unit_draws": np.array([0.0, 100.0])},
                {**common, "segment": "S2", "_total_effect_draws": np.array([100.0, 0.0]), "_effect_unit_draws": np.array([100.0, 0.0])},
            ]
        )
        summary = summarize_forecast_detail(detail)
        total = summary[(summary["segment"] == "__ALL__") & (summary["channel"] == "__TOTAL__")].iloc[0]
        self.assertAlmostEqual(float(total["total_effect_p50"]), 100.0)

    def test_silent_cell_loss_fails_reconciliation(self) -> None:
        plan = pd.DataFrame(
            [{"campaign_name": "A", "segment": "S", "geo": "G", "channel": "C", "budget_rub": 100.0}]
        )
        detail = pd.DataFrame(
            [
                {
                    "campaign_name": "A",
                    "segment": "S",
                    "geo": "OTHER",
                    "channel": "C",
                    "target": "turnover_per_user",
                    "spend_rub": 100.0,
                }
            ]
        )
        with self.assertRaises(ValueError):
            _assert_modeled_spend_reconciles(plan, detail)


class OptimizerConstraintTests(unittest.TestCase):
    @staticmethod
    def _response_kernel(cell_pos: int, slope: float) -> _CellResponseKernel:
        draws_n = 8
        base = np.zeros((draws_n, 1), dtype=float)
        return _CellResponseKernel(
            cell_pos=cell_pos,
            segment="S",
            geo=f"G{cell_pos}",
            channel="C",
            base_argument=base,
            unit_argument_per_rub=np.full((draws_n, 1), slope, dtype=float),
            counterfactual_tanh=np.zeros_like(base),
            effect_multiplier=np.full((draws_n, 1), 100.0, dtype=float),
        )

    def test_adaptive_greedy_preserves_exact_budget_without_line_item_rounding(self) -> None:
        kernels = [self._response_kernel(0, 0.02), self._response_kernel(1, 0.004)]
        allocation, diagnostics = _greedy_marginal_allocation(
            kernels,
            lower=np.zeros(2),
            upper=np.full(2, 1_000.0),
            total_budget=123.45,
            quantum_rub=10.0,
            statistic="p50",
        )
        self.assertAlmostEqual(float(allocation.sum()), 123.45, places=8)
        self.assertAlmostEqual(float(diagnostics["unallocated_budget_rub"]), 0.0, places=8)
        self.assertTrue(any(abs(value / 10.0 - round(value / 10.0)) > 1e-6 for value in allocation))

    def test_coordinate_refinement_finds_small_profitable_transfer(self) -> None:
        kernels = [self._response_kernel(0, 0.02), self._response_kernel(1, 0.002)]
        initial = np.array([50.0, 50.0])
        refined, _, trace, diagnostics = _adaptive_coordinate_refine(
            kernels,
            initial,
            lower=np.zeros(2),
            upper=np.full(2, 100.0),
            transfer_steps_rub=[10.0, 1.0, 0.1],
            beam_width=2,
            max_evaluations=200,
            statistic="p50",
        )
        self.assertAlmostEqual(float(refined.sum()), float(initial.sum()), places=8)
        self.assertGreater(float(refined[0]), float(initial[0]))
        self.assertTrue(any(bool(row["accepted"]) for row in trace))
        self.assertGreater(float(diagnostics["final_score"]), 0.0)

    def test_box_projection_preserves_budget_and_caps(self) -> None:
        projected = _project_box_simplex(
            preferred=np.array([90.0, 10.0]),
            lower=np.array([20.0, 20.0]),
            upper=np.array([60.0, 80.0]),
            total_budget=100.0,
        )
        self.assertAlmostEqual(float(projected.sum()), 100.0, places=6)
        self.assertTrue(np.all(projected >= np.array([20.0, 20.0]) - 1e-6))
        self.assertTrue(np.all(projected <= np.array([60.0, 80.0]) + 1e-6))

    def test_scenario_three_and_four_match_business_definitions(self) -> None:
        cells = pd.DataFrame(
            [
                {"channel": "A", "geo": "X", "budget_rub": 60.0},
                {"channel": "A", "geo": "Y", "budget_rub": 20.0},
                {"channel": "B", "geo": "X", "budget_rub": 10.0},
                {"channel": "B", "geo": "Y", "budget_rub": 10.0},
            ]
        )
        scenario_three = _channel_balanced_candidate(cells, 100.0)
        scenario_four = _geo_balanced_candidate(cells, 100.0)
        self.assertEqual(scenario_three.groupby("channel")["budget_rub"].sum().to_dict(), {"A": 80.0, "B": 20.0})
        self.assertEqual(scenario_three.groupby("channel")["budget_rub"].apply(list).to_dict(), {"A": [40.0, 40.0], "B": [10.0, 10.0]})
        self.assertEqual(scenario_four.groupby("geo")["budget_rub"].sum().to_dict(), {"X": 70.0, "Y": 30.0})
        self.assertEqual(scenario_four.groupby("geo")["budget_rub"].apply(list).to_dict(), {"X": [35.0, 35.0], "Y": [15.0, 15.0]})

    def test_scenario_one_preserves_nonuniform_daily_profile(self) -> None:
        cells = pd.DataFrame(
            [{"campaign_name": "A", "segment": "S", "geo": "G", "channel": "C", "budget_rub": 100.0}]
        )
        source = pd.DataFrame(
            [
                {"campaign_name": "A", "segment": "S", "geo": "G", "channel": "C", "date": "2026-01-01", "budget_rub": 90.0},
                {"campaign_name": "A", "segment": "S", "geo": "G", "channel": "C", "date": "2026-01-02", "budget_rub": 10.0},
            ]
        )
        daily = _make_candidate_daily(cells, source, "A__scenario1_current_plan")
        self.assertEqual([row["budget_rub"] for row in daily], [90.0, 10.0])

    def test_configured_p50_objective_beats_slightly_higher_p10(self) -> None:
        contract = _compile_optimizer_objective(
            {"primary": "maximize_incremental_turnover_p50", "model_risk_policy": "balanced"}
        )
        common = {
            "campaign_name": "A",
            "segment": "__ALL__",
            "target": "turnover_per_user",
            "channel": "__TOTAL__",
            "optimizer_use_counts": "optimize:1",
        }
        higher_p50 = pd.DataFrame([{**common, "total_effect_p10": 9.0, "total_effect_p50": 20.0}])
        higher_p10 = pd.DataFrame([{**common, "total_effect_p10": 10.0, "total_effect_p50": 19.0}])
        items = [
            {"candidate_name": "higher_p50", "score": _candidate_score(higher_p50, contract), "downside_score": 9.0},
            {"candidate_name": "higher_p10", "score": _candidate_score(higher_p10, contract), "downside_score": 10.0},
        ]
        ranked = sorted(items, key=_reliable_candidate_sort_key)
        self.assertEqual(ranked[0]["candidate_name"], "higher_p50")

    def test_strict_and_balanced_risk_policies_differ(self) -> None:
        balanced = _compile_optimizer_objective(
            {"primary": "maximize_incremental_turnover_p50", "model_risk_policy": "balanced"}
        )
        strict = _compile_optimizer_objective(
            {"primary": "maximize_incremental_turnover_p50", "model_risk_policy": "strict"}
        )
        self.assertEqual(_risk_policy_violation_count(1, 0, balanced), 0)
        self.assertEqual(_risk_policy_violation_count(1, 0, strict), 1)

    def test_fixed_diagnostic_cell_requires_supported_full_budget(self) -> None:
        engine = _engine_for_support()
        engine.capability.loc[:, "allowed_use"] = "diagnostic"
        engine.capability.loc[:, "optimizer_use"] = "fixed_at_plan"
        cells = pd.DataFrame(
            [{"campaign_name": "A", "segment": "S", "geo": "G", "channel": "C", "budget_rub": 10.0}]
        )
        source = pd.DataFrame(
            [
                {
                    "campaign_name": "A",
                    "segment": "S",
                    "geo": "G",
                    "channel": "C",
                    "date": "2026-01-01",
                    "budget_rub": 150.0,
                }
            ]
        )
        with self.assertRaisesRegex(CandidateFeasibilityError, "fixed-at-plan"):
            _enforce_candidate_constraints(
                cells,
                source,
                engine,
                150.0,
                support_limit="p95",
            )
        projected = _enforce_candidate_constraints(
            cells,
            source,
            engine,
            150.0,
            support_limit="p99",
        )
        self.assertAlmostEqual(float(projected["budget_rub"].iloc[0]), 150.0)

    def test_fixed_diagnostic_cell_above_robust_cap_becomes_safe_partial(self) -> None:
        engine = _engine_for_support()
        engine.capability.loc[:, "allowed_use"] = "diagnostic"
        engine.capability.loc[:, "optimizer_use"] = "fixed_at_plan"
        source = pd.DataFrame(
            [
                {
                    "campaign_name": "A",
                    "segment": "S",
                    "geo": "G",
                    "channel": "C",
                    "date": "2026-01-01",
                    "budget_rub": 250.0,
                }
            ]
        )
        cells = source.drop(columns=["date"]).copy()

        candidates = _generate_scenario5_candidates(
            cells,
            source,
            engine,
            250.0,
            {
                "support_expansion_levels": ["p95", "p99", "robust_upper"],
                "approved_maximum_risk_boundary": "robust_upper",
                "projection_modes": ["proportional"],
            },
            {},
        )

        self.assertEqual(len(candidates), 1)
        _, candidate = candidates[0]
        self.assertEqual(candidate["scenario_variant"].iloc[0], "safe_partial")
        self.assertAlmostEqual(float(candidate["allocated_budget_rub"].iloc[0]), 220.0)
        risk = _candidate_risk_budget_summary(candidate, source, engine)
        self.assertAlmostEqual(float(risk["high_risk_budget_rub"]), 0.0)

    def test_scenario_five_returns_partial_plan_instead_of_unsafe_fallback(self) -> None:
        engine = _engine_for_support()
        cells = pd.DataFrame(
            [{"campaign_name": "A", "segment": "S", "geo": "G", "channel": "C", "budget_rub": 250.0}]
        )
        source = pd.DataFrame(
            [
                {
                    "campaign_name": "A",
                    "segment": "S",
                    "geo": "G",
                    "channel": "C",
                    "date": "2026-01-01",
                    "budget_rub": 250.0,
                }
            ]
        )
        projected = _support_weighted_candidate(cells, source, engine, 250.0)
        self.assertAlmostEqual(float(projected["budget_rub"].sum()), 100.0)
        self.assertAlmostEqual(float(projected["allocated_budget_rub"].iloc[0]), 100.0)
        self.assertAlmostEqual(float(projected["unallocated_budget_rub"].iloc[0]), 150.0)
        self.assertEqual(str(projected["support_limit_policy"].iloc[0]), "p95")

    def test_scenario_five_uses_controlled_expansion_before_partial_fallback(self) -> None:
        engine = _engine_for_support()
        source = pd.DataFrame(
            [
                {
                    "campaign_name": "A",
                    "segment": "S",
                    "geo": "G",
                    "channel": "C",
                    "date": "2026-01-01",
                    "budget_rub": 180.0,
                }
            ]
        )
        cells = source.drop(columns=["date"]).copy()
        policy = {
            "support_expansion_levels": ["p95", "p99", "robust_upper"],
            "approved_maximum_risk_boundary": "robust_upper",
            "projection_modes": ["proportional", "additive"],
        }

        candidates = _generate_scenario5_candidates(
            cells, source, engine, 180.0, policy, {}
        )

        self.assertTrue(candidates)
        for suffix, candidate in candidates:
            self.assertIn("full_conservative_robust_upper", suffix)
            self.assertAlmostEqual(float(candidate["budget_rub"].sum()), 180.0)
            self.assertEqual(candidate["scenario_variant"].iloc[0], "full_conservative")
            self.assertAlmostEqual(float(candidate["unallocated_budget_rub"].iloc[0]), 0.0)
            risk = _candidate_risk_budget_summary(candidate, source, engine)
            self.assertAlmostEqual(float(risk["within_support_budget_rub"]), 100.0)
            self.assertAlmostEqual(float(risk["controlled_extrapolation_budget_rub"]), 80.0)
            self.assertAlmostEqual(float(risk["high_risk_budget_rub"]), 0.0)

    def test_scenario_five_partial_is_explicit_after_robust_capacity_is_exhausted(self) -> None:
        engine = _engine_for_support()
        source = pd.DataFrame(
            [
                {
                    "campaign_name": "A",
                    "segment": "S",
                    "geo": "G",
                    "channel": "C",
                    "date": "2026-01-01",
                    "budget_rub": 250.0,
                }
            ]
        )
        cells = source.drop(columns=["date"]).copy()
        candidates = _generate_scenario5_candidates(
            cells,
            source,
            engine,
            250.0,
            {
                "support_expansion_levels": ["p95", "p99", "robust_upper"],
                "approved_maximum_risk_boundary": "robust_upper",
                "projection_modes": ["proportional"],
            },
            {},
        )
        self.assertEqual(len(candidates), 1)
        _, candidate = candidates[0]
        self.assertEqual(candidate["scenario_variant"].iloc[0], "safe_partial")
        self.assertAlmostEqual(float(candidate["allocated_budget_rub"].iloc[0]), 220.0)
        self.assertAlmostEqual(float(candidate["unallocated_budget_rub"].iloc[0]), 30.0)
        self.assertTrue(str(candidate["full_allocation_impossible_reason"].iloc[0]).strip())
        self.assertTrue(str(candidate["limiting_constraints"].iloc[0]).strip())

    def test_scenario_six_rejects_silent_partial_allocation(self) -> None:
        cells = pd.DataFrame(
            [
                {"campaign_name": "A", "segment": "S", "geo": "G1", "channel": "C", "budget_rub": 50.0},
                {"campaign_name": "A", "segment": "S", "geo": "G2", "channel": "C", "budget_rub": 50.0},
            ]
        )
        with self.assertRaisesRegex(CandidateFeasibilityError, "full requested budget"):
            _candidate_from_allocation(
                cells,
                np.array([40.0, 40.0]),
                100.0,
                support_limit="robust_upper",
                diagnostics={},
            )

    def test_turnover_serving_inventory_is_measured_and_fails_closed(self) -> None:
        metadata = {
            "fits": {
                **{
                    f"segment_{index}::turnover_per_user": {
                        "segment": f"segment_{index}",
                        "target": "turnover_per_user",
                    }
                    for index in range(4)
                },
                **{
                    f"diagnostic_{index}": {
                        "segment": f"segment_{index % 4}",
                        "target": "orders_per_user" if index < 4 else "avg_basket",
                    }
                    for index in range(8)
                },
            }
        }
        inventory = validate_serving_model_inventory(metadata)
        self.assertEqual(inventory["research_models_in_package_n"], 12)
        self.assertEqual(inventory["active_serving_models_n"], 4)

        missing_fit = {"fits": dict(metadata["fits"])}
        missing_fit["fits"].pop("segment_3::turnover_per_user")
        self.assertEqual(serving_model_inventory(missing_fit)["active_serving_models_n"], 3)
        with self.assertRaises(ServingSemanticsError):
            validate_serving_model_inventory(missing_fit)

    def test_scenario_six_infeasible_capacity_skips_posterior_kernel_build(self) -> None:
        engine = _engine_for_support()
        support_row = engine.support_bounds.iloc[0].to_dict()
        engine.support_bounds = pd.DataFrame(
            [
                {**support_row, "geo_label": "G1"},
                {**support_row, "geo_label": "G2"},
            ]
        )
        source = pd.DataFrame(
            [
                {
                    "campaign_name": "A",
                    "segment": "S",
                    "geo": geo,
                    "channel": "C",
                    "date": "2026-01-01",
                    "budget_rub": 250.0,
                }
                for geo in ("G1", "G2")
            ]
        )
        cells = source.drop(columns=["date"]).copy()
        scenario_config = {
            "require_full_budget": True,
            "support_limits": ["p99", "robust_upper"],
            "approved_maximum_risk_boundary": "robust_upper",
            "infeasible_when_full_budget_cannot_be_allocated": True,
        }

        with patch(
            "mmm_core.forecast_engine._build_turnover_response_kernels"
        ) as kernel_builder:
            with self.assertRaisesRegex(CandidateFeasibilityError, "infeasible"):
                _generate_adaptive_scenario6_candidates(
                    cells,
                    source,
                    engine,
                    "A",
                    500.0,
                    search_samples=8,
                    seed=42,
                    max_evaluations=16,
                    finalists=2,
                    scenario_config=scenario_config,
                    business_constraints={},
                    analog_year=None,
                    analog_missing_geo_policy="nearest_available_year_same_geo",
                )
        kernel_builder.assert_not_called()

    def test_scenario_six_fixed_cell_above_robust_cap_is_infeasible(self) -> None:
        engine = _engine_for_support()
        engine.capability.loc[:, "allowed_use"] = "diagnostic"
        engine.capability.loc[:, "optimizer_use"] = "fixed_at_plan"
        source = pd.DataFrame(
            [
                {
                    "campaign_name": "A",
                    "segment": "S",
                    "geo": "G",
                    "channel": "C",
                    "date": "2026-01-01",
                    "budget_rub": 250.0,
                }
            ]
        )
        cells = source.drop(columns=["date"]).copy()

        with patch(
            "mmm_core.forecast_engine._build_turnover_response_kernels"
        ) as kernel_builder:
            with self.assertRaisesRegex(CandidateFeasibilityError, "infeasible"):
                _generate_adaptive_scenario6_candidates(
                    cells,
                    source,
                    engine,
                    "A",
                    250.0,
                    search_samples=8,
                    seed=42,
                    max_evaluations=16,
                    finalists=2,
                    scenario_config={
                        "require_full_budget": True,
                        "support_limits": ["p99", "robust_upper"],
                        "approved_maximum_risk_boundary": "robust_upper",
                    },
                    business_constraints={},
                    analog_year=None,
                    analog_missing_geo_policy="nearest_available_year_same_geo",
                )
        kernel_builder.assert_not_called()

    def test_scenario_five_proportional_projection_preserves_source_mix(self) -> None:
        projected = _project_proportional_box_simplex(
            preferred=np.array([60.0, 30.0, 10.0]),
            lower=np.zeros(3),
            upper=np.array([100.0, 100.0, 100.0]),
            total_budget=100.0,
        )
        np.testing.assert_allclose(projected, np.array([60.0, 30.0, 10.0]), atol=1e-6)

        capped = _project_proportional_box_simplex(
            preferred=np.array([60.0, 30.0, 10.0]),
            lower=np.zeros(3),
            upper=np.array([40.0, 100.0, 100.0]),
            total_budget=100.0,
        )
        self.assertAlmostEqual(float(capped.sum()), 100.0, places=6)
        self.assertAlmostEqual(float(capped[0]), 40.0, places=6)
        self.assertGreater(float(capped[1]), float(capped[2]))

    def test_paired_comparison_uses_identical_posterior_draws(self) -> None:
        draws = pd.DataFrame(
            [
                {"campaign_name": "A__scenario1_current_plan", "target": "turnover_per_user", "draw_index": i, "total_effect": value}
                for i, value in enumerate([10.0, 20.0, 30.0])
            ]
            + [
                {"campaign_name": "A__scenario6_test", "target": "turnover_per_user", "draw_index": i, "total_effect": value}
                for i, value in enumerate([11.0, 21.0, 29.0])
            ]
        )
        result = _paired_candidate_comparisons(
            draws,
            source_campaign_name="A",
            reference_candidate_name="A__scenario1_current_plan",
            decision_policy=_compile_decision_policy({}),
        )
        candidate = result[result["candidate_name"].eq("A__scenario6_test")].iloc[0]
        self.assertAlmostEqual(float(candidate["paired_delta_p50"]), 1.0)
        self.assertAlmostEqual(float(candidate["paired_probability_gt_zero"]), 2.0 / 3.0)


class DataRefreshControlTests(unittest.TestCase):
    def test_leading_control_backfill_is_forbidden(self) -> None:
        dates = pd.DataFrame({"date": pd.date_range("2025-01-01", "2025-01-03")})
        source = pd.DataFrame(
            {
                "date": [pd.Timestamp("2025-01-02")],
                "control": [10.0],
                "source_name": ["test"],
            }
        )
        with self.assertRaises(ValueError):
            DATA_PIPELINE._daily_backward_fill(
                dates,
                source,
                value_column="control",
                max_staleness_days=10,
            )

    def test_conflicting_control_overlap_fails_closed(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path_a = Path(tmp) / "a.csv"
            path_b = Path(tmp) / "b.csv"
            path_a.write_text("a", encoding="utf-8")
            path_b.write_text("b", encoding="utf-8")
            frame_a = pd.DataFrame({"date": ["2025-01-01"], "control": [10.0]})
            frame_b = pd.DataFrame({"date": ["2025-01-01"], "control": [11.0]})
            with self.assertRaises(ValueError):
                DATA_PIPELINE._combine_control_sources(
                    [("history", path_a, frame_a), ("tail", path_b, frame_b)],
                    value_column="control",
                )

    def test_invalid_target_bundle_uses_local_geo_segment_mean(self) -> None:
        dates = pd.date_range("2025-06-27", periods=7, freq="D")
        rows = []
        for index, date in enumerate(dates):
            users = 0 if index == 3 else 10 + index
            orders = 1 if index == 3 else 20 + index
            turnover = 100.0 if index == 3 else 2_000.0 + 100.0 * index
            rows.append(
                {
                    "date": date,
                    "geo_label": "КОНАКОВО",
                    "network": "ТС5",
                    "channel": "Онлайн",
                    "orders_cnt": orders,
                    "turnover_total": turnover,
                    "unique_users": users,
                    "turnover_per_user": np.nan if users == 0 else turnover / users,
                    "orders_per_user": np.nan if users == 0 else orders / users,
                    "avg_basket": turnover / orders,
                    "turnover_per_user_raw": np.nan if users == 0 else turnover / users,
                    "orders_per_user_raw": np.nan if users == 0 else orders / users,
                    "spend_Test": 140.0,
                    "spend_Test_pc": np.nan if users == 0 else 140.0 / users,
                }
            )
        rows.extend(
            {
                **row,
                "geo_label": "КОНАКОВО",
                "network": "ТСХ",
                "orders_cnt": 10_000,
                "turnover_total": 10_000_000.0,
                "unique_users": 10_000,
                "turnover_per_user": 1_000.0,
                "orders_per_user": 1.0,
                "avg_basket": 1_000.0,
            }
            for row in rows[:7]
        )
        panel = pd.DataFrame(rows)

        cleaned, audit = DATA_PIPELINE.impute_invalid_target_rows(panel)
        cleaned, auxiliary = DATA_PIPELINE.refresh_target_auxiliary(cleaned)
        patched = cleaned.iloc[3]

        self.assertEqual(int(patched["unique_users"]), 13)
        self.assertEqual(int(patched["orders_cnt"]), 23)
        self.assertAlmostEqual(float(patched["turnover_total"]), 2_300.0)
        self.assertEqual(audit["row_index"].nunique(), 1)
        self.assertEqual(set(audit["method"]), {"same_geo_segment_centered_window"})
        self.assertEqual(set(audit["peer_rows"]), {6})
        self.assertAlmostEqual(float(patched["spend_Test_pc"]), 140.0 / 13.0)
        self.assertEqual(auxiliary["spend_pc_missing_after"], 0)

    def test_new_geo_target_imputation_falls_back_to_same_segment_date(self) -> None:
        date = pd.Timestamp("2026-01-01")
        rows = [
            {
                "date": date,
                "geo_label": "ЯКУТСК",
                "network": "ТС5",
                "channel": "Онлайн",
                "orders_cnt": 1,
                "turnover_total": 3_361.0,
                "unique_users": 0,
                "turnover_per_user": np.nan,
                "orders_per_user": np.nan,
                "avg_basket": 3_361.0,
            }
        ]
        for index in range(6):
            rows.append(
                {
                    "date": date,
                    "geo_label": f"G{index}",
                    "network": "ТС5",
                    "channel": "Онлайн",
                    "orders_cnt": 100 + index,
                    "turnover_total": 150_000.0 + index * 1_000.0,
                    "unique_users": 90 + index,
                    "turnover_per_user": 1_500.0 + index,
                    "orders_per_user": 1.05 + index / 100.0,
                    "avg_basket": 1_450.0 + index,
                }
            )

        cleaned, audit = DATA_PIPELINE.impute_invalid_target_rows(pd.DataFrame(rows))

        self.assertGreater(int(cleaned.iloc[0]["unique_users"]), 0)
        self.assertEqual(set(audit["method"]), {"same_segment_same_date"})
        self.assertEqual(set(audit["peer_rows"]), {6})

    def test_shadow_comparison_ignores_serialization_noise(self) -> None:
        candidate = pd.DataFrame(
            {
                "date": pd.to_datetime(["2025-01-01"]),
                "geo_label": ["G"],
                "network": ["ТС5"],
                "channel": ["Онлайн"],
                "turnover_total": [100.0 + 5e-11],
                "ruonia_rate": [20.0],
            }
        )
        promoted = candidate.copy()
        promoted["turnover_total"] = 100.0
        promoted["ruonia_rate"] = 15.57
        with tempfile.TemporaryDirectory() as tmp:
            promoted_path = Path(tmp) / "promoted.parquet"
            audit_path = Path(tmp) / "diff.csv"
            promoted.to_parquet(promoted_path, index=False)
            result = DATA_PIPELINE.compare_candidate_to_promoted(candidate, promoted_path, audit_path)
            audit = pd.read_csv(audit_path)

        self.assertEqual(result["columns_with_differences"], 1)
        self.assertEqual(audit["column"].tolist(), ["ruonia_rate"])
        self.assertEqual(audit["difference_category"].tolist(), ["expected_control_history_correction"])

    def test_reviewed_panel_promotion_is_hash_bound_and_immutable(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            candidate_path = root / "candidate.parquet"
            baseline_path = root / "panel_final_v2.parquet"
            target_path = root / "panel_final_v3.parquet"
            build_manifest_path = root / "build_manifest.json"
            preflight_path = root / "panel_preflight.json"
            decision_path = root / "panel_final_v3_promotion_decision.json"
            pd.DataFrame({"value": [1.0]}).to_parquet(candidate_path, index=False)
            pd.DataFrame({"value": [0.0]}).to_parquet(baseline_path, index=False)
            candidate_hash = DATA_PIPELINE._sha256_path(candidate_path)
            baseline_hash = DATA_PIPELINE._sha256_path(baseline_path)
            build_manifest_path.write_text(
                json.dumps(
                    {
                        "run_id": "test_refresh",
                        "status": "candidate_built_not_promoted",
                        "promotion_blockers": [],
                        "candidate_panel": str(candidate_path),
                        "candidate_panel_sha256": candidate_hash,
                        "baseline_promoted_panel_sha256": baseline_hash,
                        "summary": {"promotion_status": "candidate_ready_for_dq", "promotion_blockers": []},
                    }
                ),
                encoding="utf-8",
            )
            preflight_path.write_text(
                json.dumps(
                    {
                        "preflight_status": "passed",
                        "panel_path": str(candidate_path),
                        "panel_sha256": candidate_hash,
                    }
                ),
                encoding="utf-8",
            )

            decision = DATA_PIPELINE.promote_reviewed_candidate(
                candidate_path=candidate_path,
                baseline_path=baseline_path,
                target_path=target_path,
                build_manifest_path=build_manifest_path,
                preflight_summary_path=preflight_path,
                decision_path=decision_path,
                reviewed_by="project_owner",
                reason="reviewed DQ passed",
            )

            self.assertEqual(DATA_PIPELINE._sha256_path(target_path), candidate_hash)
            self.assertEqual(decision["status"], "reviewed_promoted")
            self.assertTrue(decision_path.exists())
            with self.assertRaises(FileExistsError):
                DATA_PIPELINE.promote_reviewed_candidate(
                    candidate_path=candidate_path,
                    baseline_path=baseline_path,
                    target_path=target_path,
                    build_manifest_path=build_manifest_path,
                    preflight_summary_path=preflight_path,
                    decision_path=decision_path,
                    reviewed_by="project_owner",
                    reason="second promotion must fail",
                )


class PanelPreflightTests(unittest.TestCase):
    @staticmethod
    def _summary() -> dict:
        return {
            "train": {"rows": 365},
            "dq": {
                "duplicate_grain_rows": 0,
                "target_na_cells_train": 0,
                "target_nonpositive_cells_train": 0,
                "rf_like_geo_rows": 0,
                "control_columns_missing": [],
                "control_variation_by_year": {
                    "2025": {
                        column: {
                            "days": 365,
                            "missing_days": 0,
                            "unique_values": 365,
                            "std": 1.0,
                        }
                        for column in PANEL_PRIORS.CONTROL_VARIATION_COLS
                    }
                },
            },
        }

    def test_nonpositive_model_target_blocks_preflight(self) -> None:
        summary = self._summary()
        summary["dq"]["target_nonpositive_cells_train"] = 1
        with self.assertRaisesRegex(ValueError, "NONPOSITIVE_TARGETS_IN_TRAIN_CUT"):
            PANEL_PRIORS.assert_panel_preflight(summary)

    def test_constant_yearly_control_blocks_preflight(self) -> None:
        summary = self._summary()
        summary["dq"]["control_variation_by_year"]["2025"]["ruonia_rate"].update(
            {"unique_values": 1, "std": 0.0}
        )
        with self.assertRaisesRegex(ValueError, "CONSTANT_TEMPORAL_CONTROL:2025:ruonia_rate"):
            PANEL_PRIORS.assert_panel_preflight(summary)

    def test_varied_complete_controls_pass_preflight(self) -> None:
        PANEL_PRIORS.assert_panel_preflight(self._summary())


class MarketerReportContractTests(unittest.TestCase):
    def test_wrapped_marketer_explanation_expands_row_height(self) -> None:
        workbook = Workbook()
        worksheet = workbook.active
        long_explanation = (
            "Система сохраняет исходный план, потому что найденное улучшение меньше "
            "порога содержательности и поиск не подтвердил устойчивое преимущество."
        )
        row_after_table, _ = MARKETER_REPORT._write_table(
            worksheet,
            pd.DataFrame([{"Почему": long_explanation}]),
            start_row=1,
            title="Проверка",
        )
        worksheet.column_dimensions["A"].width = 38

        MARKETER_REPORT._fit_wrapped_row_heights(worksheet)

        data_row = row_after_table - 2
        self.assertGreater(worksheet.row_dimensions[data_row].height, 34)

    def test_cached_paired_comparisons_are_attached_to_finalists(self) -> None:
        finalist = pd.DataFrame(
            [
                {
                    "candidate_name": "campaign__scenario6_support_aware_hybrid_001",
                    "target": "turnover_per_user",
                    "total_effect_p50": 101.0,
                }
            ]
        )
        paired = pd.DataFrame(
            [
                {
                    "candidate_name": "campaign__scenario6_support_aware_hybrid_001",
                    "target": "turnover_per_user",
                    "paired_delta_p10": -0.1,
                    "paired_delta_p50": 1.0,
                    "paired_delta_p90": 2.0,
                    "paired_probability_gt_zero": 0.91,
                    "paired_probability_noninferior": 0.98,
                    "paired_draws_n": 300,
                }
            ]
        )
        attached = MARKETER_REPORT._attach_paired_comparisons(finalist, paired)
        self.assertAlmostEqual(float(attached.iloc[0]["paired_delta_p50"]), 1.0)
        self.assertAlmostEqual(float(attached.iloc[0]["paired_probability_gt_zero"]), 0.91)

    @staticmethod
    def _decision_inputs(*, source_strong: int, s6_gain_mln: float) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
        scenario_results = pd.DataFrame(
            [
                {
                    "campaign_name": "campaign",
                    "scenario_no": "S01",
                    "scenario_name": "Сценарий 1. Как загрузили",
                    "candidate_name": "campaign__scenario1_current_plan",
                    "budget_mln_rub": 10.0,
                    "requested_budget_mln_rub": 10.0,
                    "allocated_budget_mln_rub": 10.0,
                    "allocated_budget_share": 1.0,
                    "rto_p10_mln": 90.0,
                    "rto_p50_mln": 100.0,
                    "rto_p90_mln": 110.0,
                    "rto_roas_p50": 10.0,
                    "strong_support_warnings_n": source_strong,
                    "hard_support_warnings_n": 0,
                    "policy_violations_n": 0,
                    "rto_optimizer_use": "optimize:1",
                    "quality_status": "Требуется ручная проверка" if source_strong else "Сопоставимо с историей",
                    "quality_explanation": "source",
                }
            ]
        )
        scenario6 = pd.DataFrame(
            [
                {
                    "campaign_name": "campaign",
                    "candidate_name": "campaign__scenario6_support_aware_hybrid_001",
                    "scenario6_ran": True,
                    "is_best_safe_s6": True,
                    "rto_p10_mln": 90.1,
                    "rto_p50_mln": 100.0 + s6_gain_mln,
                    "rto_p90_mln": 110.1,
                    "roas_p50": 10.0,
                    "requested_budget_mln_rub": 10.0,
                    "allocated_budget_mln_rub": 10.0,
                    "unallocated_budget_mln_rub": 0.0,
                    "allocated_budget_share": 1.0,
                    "elevated_support_warnings_n": 0,
                    "strong_support_warnings_n": 0,
                    "hard_support_warnings_n": 0,
                    "policy_violations_n": 0,
                    "objective_rows_n": 1,
                    "quality_status": "Сопоставимо с историей",
                    "quality_explanation": "safe",
                    "paired_delta_p10": s6_gain_mln,
                    "paired_delta_p50": s6_gain_mln,
                    "paired_delta_p90": s6_gain_mln,
                    "paired_probability_gt_zero": 1.0,
                    "paired_probability_noninferior": 1.0,
                }
            ]
        )
        campaign_summary = pd.DataFrame(
            [
                {
                    "campaign_name": "campaign",
                    "uploaded_budget_mln_rub": 10.0,
                    "model_input_budget_mln_rub": 10.0,
                    "unmodeled_budget_mln_rub": 0.0,
                }
            ]
        )
        allocation = pd.DataFrame(
            [
                {"source_campaign_name": "campaign", "candidate_name": "campaign__scenario1_current_plan", "segment": "S", "geo": "G1", "channel": "C", "budget_rub": 5_000_000.0},
                {"source_campaign_name": "campaign", "candidate_name": "campaign__scenario1_current_plan", "segment": "S", "geo": "G2", "channel": "C", "budget_rub": 5_000_000.0},
                {"source_campaign_name": "campaign", "candidate_name": "campaign__scenario6_support_aware_hybrid_001", "segment": "S", "geo": "G1", "channel": "C", "budget_rub": 4_000_000.0},
                {"source_campaign_name": "campaign", "candidate_name": "campaign__scenario6_support_aware_hybrid_001", "segment": "S", "geo": "G2", "channel": "C", "budget_rub": 6_000_000.0},
            ]
        )
        return scenario_results, scenario6, campaign_summary, allocation

    def test_infeasible_scenario6_is_reported_as_attempted_but_unavailable(self) -> None:
        candidate_scores = pd.DataFrame(
            [
                {
                    "campaign_name": "campaign",
                    "candidate_name": f"campaign__scenario6_support_aware_hybrid_{index:03d}",
                    "method": "scenario6_search",
                    "precheck_status": "rejected_infeasible",
                    "precheck_reason": "Campaign budget exceeds aggregate model/support capacity",
                    "total_budget_rub": 10_000_000.0,
                }
                for index in range(1, 4)
            ]
        )
        scenario6 = MARKETER_REPORT._scenario6_summary(candidate_scores, pd.DataFrame())
        self.assertEqual(len(scenario6), 1)
        self.assertTrue(bool(scenario6.iloc[0]["scenario6_ran"]))
        self.assertEqual(int(scenario6.iloc[0]["attempts_total_n"]), 3)
        self.assertEqual(int(scenario6.iloc[0]["attempts_rejected_by_support_n"]), 3)
        self.assertIn("допустимый вариант не найден", scenario6.iloc[0]["candidate_role"].lower())
        scenario_results = pd.DataFrame(
            [
                {
                    "campaign_name": "campaign",
                    "scenario_no": "S01",
                    "scenario_name": "Сценарий 1. Как загрузили",
                    "candidate_name": "campaign__scenario1_current_plan",
                    "budget_mln_rub": 10.0,
                    "rto_p10_mln": 1.0,
                    "rto_p50_mln": 2.0,
                    "rto_roas_p50": 0.2,
                    "orders_p50_mln": 0.0,
                    "basket_p50_mln": 0.0,
                    "support_warnings_n": 1,
                    "strong_support_warnings_n": 1,
                    "hard_support_warnings_n": 1,
                    "policy_violations_n": 0,
                    "rto_optimizer_use": "optimize:1",
                    "quality_status": "Не использовать для автоматического перераспределения",
                    "quality_explanation": "Outside support.",
                }
            ]
        )
        recommendation = MARKETER_REPORT._recommendations(
            scenario_results,
            scenario6,
            pd.DataFrame([{"campaign_name": "campaign"}]),
            min_roas_p50=None,
        ).iloc[0]
        self.assertFalse(bool(recommendation["optimizer_available"]))
        self.assertEqual(recommendation["scenario_name"], "Сценарий 1. Как загрузили")
        self.assertIn("Автоматический вариант не найден", recommendation["allocation_decision"])

    def test_no_modifiable_scenario6_cannot_be_reported_as_available(self) -> None:
        scenario_results = pd.DataFrame(
            [
                {
                    "campaign_name": "campaign",
                    "scenario_no": "S01",
                    "scenario_name": "Сценарий 1. Как загрузили",
                    "candidate_name": "campaign__scenario1_current_plan",
                    "budget_mln_rub": 10.0,
                    "rto_p10_mln": 1.0,
                    "rto_p50_mln": 2.0,
                    "rto_roas_p50": 0.2,
                    "orders_p50_mln": 0.0,
                    "basket_p50_mln": 0.0,
                    "support_warnings_n": 0,
                    "strong_support_warnings_n": 0,
                    "hard_support_warnings_n": 0,
                    "policy_violations_n": 0,
                    "rto_optimizer_use": "no_increase:4",
                    "quality_status": "Повышенная неопределенность",
                    "quality_explanation": "Only caution rows are present.",
                }
            ]
        )
        scenario6 = pd.DataFrame(
            [
                {
                    "campaign_name": "campaign",
                    "scenario6_ran": False,
                    "is_best_safe_s6": False,
                }
            ]
        )
        result = MARKETER_REPORT._recommendations(
            scenario_results,
            scenario6,
            pd.DataFrame([{"campaign_name": "campaign"}]),
            min_roas_p50=None,
        )
        self.assertFalse(bool(result.iloc[0]["optimizer_available"]))
        self.assertEqual(result.iloc[0]["scenario_no"], "S01")
        self.assertEqual(result.iloc[0]["optimizer_status"], "Перераспределение недоступно по gate policy")
        self.assertIn("Scenario 6 недоступен", result.iloc[0]["allocation_decision"])

    def test_tiny_safe_s6_gain_keeps_source_plan(self) -> None:
        scenario_results, scenario6, campaign_summary, allocation = self._decision_inputs(
            source_strong=0,
            s6_gain_mln=0.1,
        )
        result = MARKETER_REPORT._recommendations(
            scenario_results,
            scenario6,
            campaign_summary,
            min_roas_p50=None,
            allocation=allocation,
        ).iloc[0]
        self.assertEqual(result["scenario_no"], "S01")
        self.assertEqual(result["recommendation_type"], "Оставить исходный план")
        self.assertIn("не проходит materiality", result["allocation_decision"])
        self.assertEqual(result["decision_status"], "keep_uploaded_plan")
        self.assertEqual(result["review_status"], "manual_review_required")
        self.assertEqual(result["plan_status"], "Исходный план для ручной проверки")

    def test_reliability_improvement_can_override_tiny_gain(self) -> None:
        scenario_results, scenario6, campaign_summary, allocation = self._decision_inputs(
            source_strong=1,
            s6_gain_mln=0.1,
        )
        result = MARKETER_REPORT._recommendations(
            scenario_results,
            scenario6,
            campaign_summary,
            min_roas_p50=None,
            allocation=allocation,
        ).iloc[0]
        self.assertEqual(result["scenario_no"], "S06")
        self.assertEqual(result["recommendation_type"], "Перераспределить ради надежности")

    def test_full_safe_fallback_is_not_labeled_partial(self) -> None:
        scenario_results, scenario6, campaign_summary, allocation = self._decision_inputs(
            source_strong=1,
            s6_gain_mln=-5.0,
        )
        scenario6.loc[:, "rto_p10_mln"] = 80.0
        scenario6.loc[:, "paired_probability_noninferior"] = 0.0
        result = MARKETER_REPORT._recommendations(
            scenario_results,
            scenario6,
            campaign_summary,
            min_roas_p50=None,
            allocation=allocation,
        ).iloc[0]
        self.assertEqual(result["scenario_no"], "S06")
        self.assertEqual(result["recommendation_type"], "Надежный support-safe план")
        self.assertNotIn("частич", result["allocation_decision"].lower())

    def test_unmodeled_budget_means_partial_coverage_not_partial_plan(self) -> None:
        scenario_results, scenario6, campaign_summary, allocation = self._decision_inputs(
            source_strong=0,
            s6_gain_mln=0.1,
        )
        campaign_summary.loc[:, "uploaded_budget_mln_rub"] = 11.0
        campaign_summary.loc[:, "model_input_budget_mln_rub"] = 10.0
        campaign_summary.loc[:, "unmodeled_budget_mln_rub"] = 1.0
        result = MARKETER_REPORT._recommendations(
            scenario_results,
            scenario6,
            campaign_summary,
            min_roas_p50=None,
            allocation=allocation,
        ).iloc[0]
        self.assertEqual(result["scenario_no"], "S01")
        self.assertEqual(
            result["plan_status"],
            "Полный медиаплан; частичное покрытие модели",
        )


class SupportedArtifactsContractTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls.project_root = Path(__file__).resolve().parents[3]
        cls.optimizer_dir = cls.project_root / (
            "03_Outputs/02_Budget_optimizer_outputs/"
            "18_Budget_optimizer_14072026_agency_gender_boost_contract_v1"
        )
        cls.forecast_dir = cls.project_root / (
            "03_Outputs/03_AC_forecast_outputs/"
            "20_AC_forecast_14072026_agency_gender_boost_serving_v3"
        )
        cls.run_id = "optimizer_agency_gender_boost_contract_v1_14072026"
        cls.forecast_run_id = "forecast_agency_gender_boost_serving_v3_14072026"
        cls.flighting_path = cls.project_root / (
            "00_Data/00_Future_Campaigns/03_Flighting/"
            "optimizer_agency_gender_boost_contract_v1_14072026_campaign_flighting_daily.csv"
        )
        cls.model_run_dir = cls.project_root / (
            "03_Outputs/01_PyMC_outputs/09_PyMC_14072026_panel_v3_serving_policy_v3/"
            "production_panel_v3_q1_2026_guarded_serving_v3"
        )
        cls.legacy_model_run_dir = cls.project_root / (
            "03_Outputs/01_PyMC_outputs/04_PyMC_05072026_Q1_2026_refit/"
            "production_q1_2026_tc5_specific_indoor_separate_rf_prorata_"
            "tc5_online_basket_pooled_tc5_offline_turnover_nat_tv_tier_pool"
        )
        required_external_artifacts = (
            cls.optimizer_dir,
            cls.forecast_dir,
            cls.flighting_path,
            cls.model_run_dir,
            cls.legacy_model_run_dir,
        )
        missing = [
            str(path.relative_to(cls.project_root))
            for path in required_external_artifacts
            if not path.exists()
        ]
        if missing:
            raise unittest.SkipTest(
                "External supported-artifact fixture is not included in the source checkout: "
                + ", ".join(missing)
            )

    def _require(self, path: Path) -> None:
        if not path.exists():
            self.fail(f"Supported E2E artifact is not present: {path}")

    def test_supported_optimizer_s6_is_policy_and_hard_support_safe(self) -> None:
        candidate_path = self.optimizer_dir / f"{self.run_id}_optimizer_candidate_scores.csv"
        self._require(candidate_path)
        candidates = pd.read_csv(candidate_path)
        s6 = candidates[
            candidates["candidate_name"].astype(str).str.contains("__scenario6_", na=False)
            & candidates["precheck_status"].eq("scored")
        ]
        self.assertFalse(s6.empty)
        self.assertEqual(int(pd.to_numeric(s6["hard_support_warnings_n"]).fillna(0).sum()), 0)
        self.assertEqual(int(pd.to_numeric(s6["policy_violations_n"]).fillna(0).sum()), 0)

    def test_previous_supported_model_is_stale_after_policy_upgrade(self) -> None:
        with self.assertRaises(StaleModelPackageError):
            validate_existing_model_run(self.legacy_model_run_dir)

    def test_fresh_sampling_requires_guarded_fit_config(self) -> None:
        with self.assertRaisesRegex(ValueError, "Guarded fit requires --fit-config"):
            run_model_refresh(self.model_run_dir, mode="fit")

    def test_supported_allocations_preserve_budget_and_gate_actions(self) -> None:
        allocation_path = self.optimizer_dir / f"{self.run_id}_optimizer_recommended_allocations.csv"
        self._require(allocation_path)
        self._require(self.flighting_path)
        allocations = pd.read_csv(allocation_path)
        allocations = allocations[allocations["candidate_name"].astype(str).str.contains("__scenario6_", na=False)]
        source = pd.read_csv(self.flighting_path)
        keys = ["campaign_name", "segment", "geo", "channel"]
        current = source.groupby(keys, as_index=False)["budget_rub"].sum().rename(columns={"budget_rub": "current_budget_rub"})
        merged = allocations.merge(
            current,
            left_on=["source_campaign_name", "segment", "geo", "channel"],
            right_on=keys,
            how="left",
        )
        fixed = merged[merged["optimizer_policy"].eq("fixed_at_plan")]
        no_increase = merged[merged["optimizer_policy"].eq("no_increase")]
        self.assertTrue((fixed["budget_rub"] - fixed["current_budget_rub"]).abs().le(1.0).all())
        self.assertTrue((no_increase["budget_rub"] - no_increase["current_budget_rub"]).le(1.0).all())
        candidate_path = self.optimizer_dir / f"{self.run_id}_optimizer_candidate_scores.csv"
        self._require(candidate_path)
        candidates = pd.read_csv(candidate_path)
        candidates = candidates[
            candidates["candidate_name"].astype(str).str.contains("__scenario6_", na=False)
            & candidates["precheck_status"].eq("scored")
        ].set_index(["campaign_name", "candidate_name"])
        actual_totals = allocations.groupby(["source_campaign_name", "candidate_name"])["budget_rub"].sum()
        for key, value in actual_totals.items():
            candidate = candidates.loc[key]
            self.assertAlmostEqual(float(value), float(candidate["allocated_budget_rub"]), delta=1.0)
            self.assertAlmostEqual(
                float(candidate["allocated_budget_rub"] + candidate["unallocated_budget_rub"]),
                float(candidate["requested_budget_rub"]),
                delta=1.0,
            )

    def test_supported_forecast_uses_draw_level_quantiles(self) -> None:
        summary_path = self.forecast_dir / f"{self.forecast_run_id}_forecast_summary.csv"
        self._require(summary_path)
        summary = pd.read_csv(summary_path)
        self.assertEqual(set(summary["quantile_aggregation"]), {"sum_draws_then_quantile"})

    def test_scenario_one_matches_standalone_forecast_serving_sample(self) -> None:
        forecast_path = self.forecast_dir / f"{self.forecast_run_id}_forecast_summary.csv"
        finalist_path = self.optimizer_dir / f"{self.run_id}_optimizer_finalist_summary.csv"
        self._require(forecast_path)
        self._require(finalist_path)
        forecast = pd.read_csv(forecast_path)
        finalists = pd.read_csv(finalist_path)
        forecast = forecast[
            forecast["segment"].eq("__ALL__")
            & forecast["channel"].eq("__TOTAL__")
            & forecast["target"].eq("turnover_per_user")
        ]
        scenario_one = finalists[
            finalists["segment"].eq("__ALL__")
            & finalists["channel"].eq("__TOTAL__")
            & finalists["target"].eq("turnover_per_user")
            & finalists["candidate_name"].astype(str).str.contains("__scenario1_current_plan", na=False)
        ]
        merged = forecast.merge(
            scenario_one,
            left_on="campaign_name",
            right_on="source_campaign_name",
            suffixes=("_forecast", "_optimizer"),
            validate="one_to_one",
        )
        self.assertEqual(len(merged), 2)
        for metric in ["total_effect_p10", "total_effect_p50", "total_effect_p90", "roas_p50"]:
            np.testing.assert_allclose(
                merged[f"{metric}_forecast"],
                merged[f"{metric}_optimizer"],
                rtol=0.0,
                atol=1e-6,
            )

    def test_supported_artifact_lineage_matches_current_files(self) -> None:
        manifest_path = self.model_run_dir / "model_manifest.json"
        forecast_card_path = self.forecast_dir / f"{self.forecast_run_id}_forecast_run_card.json"
        optimizer_card_path = self.optimizer_dir / f"{self.run_id}_optimizer_run_card.json"
        report_card_path = self.optimizer_dir / "marketer_report_card.json"
        for path in [manifest_path, forecast_card_path, optimizer_card_path, report_card_path]:
            self._require(path)

        forecast_card = json.loads(forecast_card_path.read_text(encoding="utf-8"))
        optimizer_card = json.loads(optimizer_card_path.read_text(encoding="utf-8"))
        report_card = json.loads(report_card_path.read_text(encoding="utf-8"))
        manifest_hash = sha256_file(manifest_path)
        self.assertEqual(forecast_card["model_manifest_sha256"], manifest_hash)
        self.assertEqual(optimizer_card["model_manifest_sha256"], manifest_hash)
        self.assertEqual(forecast_card["n_samples"], optimizer_card["final_samples"])
        self.assertEqual(forecast_card["seed"], optimizer_card["final_seed"])

        for card in [forecast_card, optimizer_card]:
            for artifact_name, artifact_hash in card["runtime_lineage"]["model_artifact_sha256"].items():
                self.assertEqual(artifact_hash, sha256_file(self.model_run_dir / artifact_name))
            for output_name, output_path in card["outputs"].items():
                self.assertEqual(card["output_sha256"][output_name], sha256_file(Path(output_path)))

        code_paths = {
            "forecast_engine.py": PYMC_CODE_DIR / "mmm_core/forecast_engine.py",
            "campaign_plan.py": PYMC_CODE_DIR / "mmm_core/campaign_plan.py",
            "model_package_reader.py": PYMC_CODE_DIR / "mmm_core/model_package_reader.py",
            "ac_forecast.py": self.project_root / "02_Code/03_AC_forecast/ac_forecast.py",
            "budget_optimizer.py": self.project_root / "02_Code/02_Budget_optimizer/budget_optimizer.py",
            "marketer_report.py": self.project_root / "02_Code/02_Budget_optimizer/marketer_report.py",
        }
        for card in [forecast_card, optimizer_card]:
            for code_name, code_hash in card["runtime_lineage"]["code_sha256"].items():
                if code_name == "marketer_report.py":
                    continue
                self.assertEqual(code_hash, sha256_file(code_paths[code_name]))
        self.assertEqual(
            report_card["runtime_lineage"]["code_sha256"]["marketer_report.py"],
            sha256_file(code_paths["marketer_report.py"]),
        )

        report_sources = {
            "flighting": self.flighting_path,
            "optimizer_run_card": optimizer_card_path,
            "candidate_scores": self.optimizer_dir / f"{self.run_id}_optimizer_candidate_scores.csv",
            "finalist_summary": self.optimizer_dir / f"{self.run_id}_optimizer_finalist_summary.csv",
            "recommended_allocations": self.optimizer_dir / f"{self.run_id}_optimizer_recommended_allocations.csv",
            "paired_comparisons": self.optimizer_dir / f"{self.run_id}_optimizer_paired_comparisons.csv",
        }
        for source_name, source_path in report_sources.items():
            self.assertEqual(report_card["source_artifact_sha256"][source_name], sha256_file(source_path))
        report_outputs = {
            "xlsx": Path(report_card["output_xlsx"]),
            "scenario_results_csv": Path(report_card["scenario_results_csv"]),
            "recommendations_csv": Path(report_card["recommendations_csv"]),
            "best_plan_csv": Path(report_card["best_plan_csv"]),
            "decision_pool_csv": Path(report_card["decision_pool_csv"]),
        }
        for output_name, output_path in report_outputs.items():
            self.assertEqual(report_card["output_sha256"][output_name], sha256_file(output_path))

    def test_marketer_workbook_hides_backend_identifiers(self) -> None:
        workbook_path = self.optimizer_dir / "marketer_preprod_forecast_optimizer_report.xlsx"
        self._require(workbook_path)
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        self.assertEqual(workbook.sheetnames[0], "00_Итог_и_как_читать")
        self.assertEqual(workbook.sheetnames[-1], "99_Качество")
        self.assertEqual(len(workbook.sheetnames), 4)
        self.assertTrue(workbook.sheetnames[1].startswith("01_"))
        self.assertTrue(workbook.sheetnames[2].startswith("02_"))
        text = "\n".join(
            str(cell.value)
            for worksheet in workbook.worksheets
            for row in worksheet.iter_rows()
            for cell in row
            if cell.value is not None
        )
        self.assertNotIn("__scenario6_", text)
        self.assertNotIn("S01_AS_IS", text)
        self.assertNotIn("UPSTREAM_DIAGNOSTIC_ONLY", text)
        self.assertNotIn("rhat_max", text)
        workbook.close()


if __name__ == "__main__":
    unittest.main()
