"""Build the active-package media-plan dictionary served by the Product API."""

from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import Final

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from mmm_core.federal_geo_allocator import (
    FEDERAL_GEO_ALIASES,
    FederalGeoAllocationContext,
)
from mmm_core.serving_semantics import CHANNEL_DISPLAY_NAMES
from services.geo_catalog import CanonicalGeoCatalog


DICTIONARY_FILENAME: Final = "Словарь_для_медиаплана.xlsx"
_DIRECTION_COLUMNS: Final = (
    ("ТС5/Онлайн", "ТС5 Онлайн"),
    ("ТС5/Оффлайн", "ТС5 Офлайн"),
    ("ТСХ/Онлайн", "ТСХ Онлайн"),
    ("ТСХ/Оффлайн", "ТСХ Офлайн"),
)
_CHANNEL_DESCRIPTIONS: Final = {
    "Digital_Performance": "Performance-размещения в цифровых каналах.",
    "OOH_Total": "Наружная реклама.",
    "Indoor": "Реклама внутри помещений и торговых объектов.",
    "Радио": "Размещения на радио.",
    "Нац_ТВ": "Размещения на национальном телевидении.",
    "Рег_ТВ": "Размещения на региональном телевидении.",
}


def _format_table_sheet(
    sheet: Worksheet,
    *,
    widths: tuple[int, ...],
    header_row: int = 1,
) -> None:
    sheet.freeze_panes = f"A{header_row + 1}"
    last_column = sheet.cell(header_row, len(widths)).column_letter
    sheet.auto_filter.ref = f"A{header_row}:{last_column}{sheet.max_row}"
    for cell in sheet[header_row]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="214E63")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[
            sheet.cell(header_row, index).column_letter
        ].width = width
    for row in sheet.iter_rows(min_row=header_row + 1):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def build_media_plan_dictionary(
    context: FederalGeoAllocationContext,
    catalog: CanonicalGeoCatalog,
) -> bytes:
    """Return an XLSX derived only from one verified package context and catalog."""

    workbook = Workbook()
    workbook.properties.title = "Словарь для медиаплана"
    workbook.properties.creator = "MMM Platform"
    workbook.properties.created = datetime(2026, 1, 1)
    workbook.properties.modified = datetime(2026, 1, 1)

    channels = workbook.active
    channels.title = "Каналы"
    channels.append(("Канал", "Описание", "Федеральное размещение"))
    active_channels = {channel for _, channel in context.capability_pairs}
    for channel_id, display_name in CHANNEL_DISPLAY_NAMES.items():
        channels.append(
            (
                display_name,
                _CHANNEL_DESCRIPTIONS[channel_id],
                "Да" if channel_id in active_channels else "Нет",
            )
        )
    _format_table_sheet(channels, widths=(28, 58, 28))

    geographies = workbook.create_sheet("Географии")
    geographies.append(
        (
            "Список показывает географии, заявленные текущей моделью. "
            "Фактическая доступность расчета зависит от дат кампании и "
            "проверяется при загрузке медиаплана.",
        )
    )
    geographies.merge_cells("A1:E1")
    geographies["A1"].font = Font(bold=True, color="214E63")
    geographies["A1"].fill = PatternFill("solid", fgColor="DDEBF2")
    geographies["A1"].alignment = Alignment(wrap_text=True, vertical="center")
    geographies.row_dimensions[1].height = 34
    geographies.append(
        ("География", *(display for _, display in _DIRECTION_COLUMNS))
    )
    eligible_ids = {
        direction: {geo.geo_id for geo in context.eligible_by_direction.get(direction, ())}
        for direction, _ in _DIRECTION_COLUMNS
    }
    for entry in sorted(
        catalog.entries,
        key=lambda row: str(row["geo_display_name"]).casefold(),
    ):
        geo_id = str(entry["geo_id"])
        geographies.append(
            (
                str(entry["geo_display_name"]),
                *("Да" if geo_id in eligible_ids[direction] else "Нет" for direction, _ in _DIRECTION_COLUMNS),
            )
        )
    _format_table_sheet(
        geographies,
        widths=(34, 18, 18, 18, 18),
        header_row=2,
    )

    russia = workbook.create_sheet("Как указать всю Россию")
    russia.append(("Раздел", "Содержание"))
    russia.append(
        (
            "Правило",
            "Для федерального размещения укажите одно из точных значений ниже. "
            "Регистр букв и пробелы по краям не влияют на распознавание.",
        )
    )
    for alias in FEDERAL_GEO_ALIASES:
        russia.append(("Допустимое значение", alias))
    russia.append(
        (
            "Что произойдет",
            "Федеральный бюджет будет распределен между географиями, для которых "
            "модель поддерживает расчет на выбранный период, пропорционально "
            "населению.",
        )
    )
    russia.append(
        (
            "Пример",
            "Autumn TV | Национальное ТВ | РФ | 01.09.2026–01.10.2026 | 100 000 000 ₽",
        )
    )
    russia.append(
        (
            "Важно",
            "Если в одном плане указаны РФ и отдельная локальная география, "
            "локальный бюджет добавляется поверх федерального распределения.",
        )
    )
    _format_table_sheet(russia, widths=(24, 92))

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
