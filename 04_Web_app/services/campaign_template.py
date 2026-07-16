"""Build the synthetic campaign-plan XLSX served by the Product API."""

from __future__ import annotations

from datetime import datetime
from functools import lru_cache
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet


TEMPLATE_FILENAME = "campaign-plan-template.xlsx"
DAILY_COLUMNS = (
    "campaign_name",
    "date",
    "segment",
    "geo",
    "channel",
    "budget_rub",
)
INTERVAL_COLUMNS = (
    "campaign_name",
    "start_date",
    "end_date",
    "segment",
    "geo",
    "channel",
    "budget_rub",
)


def _format_table_sheet(sheet: Worksheet) -> None:
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="214E63")
        cell.alignment = Alignment(horizontal="center")
    for column in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column)
        sheet.column_dimensions[column[0].column_letter].width = min(
            max(max_length + 2, 14),
            36,
        )


@lru_cache(maxsize=1)
def build_campaign_plan_template() -> bytes:
    """Return a deterministic workbook with two supported input layouts."""

    workbook = Workbook()
    workbook.properties.title = "Synthetic MMM campaign-plan template"
    workbook.properties.creator = "MMM Platform"
    workbook.properties.created = datetime(2026, 1, 1)
    workbook.properties.modified = datetime(2026, 1, 1)

    instructions = workbook.active
    instructions.title = "00_Инструкция"
    instruction_rows = (
        ("Правило", "Один файл должен описывать ровно одну будущую кампанию."),
        ("Формат", "Заполните либо 01_Daily, либо 02_Interval."),
        (
            "Daily",
            "Одна строка задает бюджет одного channel x geo на конкретную дату.",
        ),
        (
            "Interval",
            "Бюджет строки равномерно распределяется от start_date до end_date.",
        ),
        (
            "Пример",
            "Все строки в шаблоне синтетические; удалите их перед загрузкой.",
        ),
    )
    for row in instruction_rows:
        instructions.append(row)
    instructions.column_dimensions["A"].width = 16
    instructions.column_dimensions["B"].width = 82
    for cell in instructions[1]:
        cell.font = Font(bold=True)

    daily = workbook.create_sheet("01_Daily")
    daily.append(DAILY_COLUMNS)
    daily.append(
        (
            "SYNTHETIC_DAILY_CAMPAIGN",
            "2026-01-15",
            "SYNTHETIC_SEGMENT",
            "SYNTHETIC_GEO_A",
            "SYNTHETIC_CHANNEL_A",
            1000,
        )
    )
    _format_table_sheet(daily)

    interval = workbook.create_sheet("02_Interval")
    interval.append(INTERVAL_COLUMNS)
    interval.append(
        (
            "SYNTHETIC_INTERVAL_CAMPAIGN",
            "2026-02-01",
            "2026-02-07",
            "SYNTHETIC_SEGMENT",
            "SYNTHETIC_GEO_A",
            "SYNTHETIC_CHANNEL_A",
            7000,
        )
    )
    _format_table_sheet(interval)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
