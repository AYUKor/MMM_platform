from __future__ import annotations

import io
import sys
import unittest
from pathlib import Path

from openpyxl import load_workbook


WEB_APP_DIR = Path(__file__).resolve().parents[1]
PYMC_CODE_DIR = WEB_APP_DIR.parent / "02_Code" / "01_PyMC"
for entry in (WEB_APP_DIR, PYMC_CODE_DIR):
    if str(entry) not in sys.path:
        sys.path.insert(0, str(entry))

from mmm_core.federal_geo_allocator import (  # noqa: E402
    EligibleGeo,
    FederalGeoAllocationContext,
    SUPPORTED_CHANNELS,
)
from services.geo_catalog import load_canonical_geo_catalog  # noqa: E402
from services.media_plan_dictionary import build_media_plan_dictionary  # noqa: E402


class MediaPlanDictionaryTest(unittest.TestCase):
    def test_workbook_is_generated_from_context_and_canonical_catalog(self) -> None:
        catalog = load_canonical_geo_catalog()
        self.assertEqual(catalog.geographies_n, 220)
        eligible = tuple(
            EligibleGeo(
                geo_id=str(row["geo_id"]),
                geo_label=str(row["geo_normalized_name"]),
                geo_display_name=str(row["geo_display_name"]),
                population_k=float(index + 1),
            )
            for index, row in enumerate(catalog.entries)
        )
        counts = {
            "ТС5/Онлайн": 211,
            "ТС5/Оффлайн": 220,
            "ТСХ/Онлайн": 114,
            "ТСХ/Оффлайн": 117,
        }
        context = FederalGeoAllocationContext(
            policy_version="FEDERAL_GEO_ALLOCATION_V1",
            package_id="pkg_1234567890abcdef_1234567890abcdef",
            package_pointer_sha256="a" * 64,
            registration_content_sha256="b" * 64,
            support_source_sha256="c" * 64,
            denominator_source_sha256="d" * 64,
            capability_source_sha256="e" * 64,
            population_source_sha256="f" * 64,
            geo_catalog_sha256="0" * 64,
            eligible_by_direction={
                direction: eligible[:count] for direction, count in counts.items()
            },
            capability_pairs=frozenset(
                (direction, channel)
                for direction in counts
                for channel in SUPPORTED_CHANNELS
            ),
            geo_by_label={geo.geo_label: geo for geo in eligible},
            eligible_geo_set_checksums={direction: "1" * 64 for direction in counts},
        )

        content = build_media_plan_dictionary(context, catalog)
        self.assertTrue(content.startswith(b"PK"))
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        self.assertEqual(
            workbook.sheetnames,
            ["Каналы", "Географии", "Как указать всю Россию"],
        )

        channels = list(workbook["Каналы"].iter_rows(values_only=True))
        self.assertEqual(
            channels[0],
            ("Канал", "Описание", "Федеральное размещение"),
        )
        self.assertEqual(len(channels) - 1, 6)
        self.assertTrue(all(row[2] == "Да" for row in channels[1:]))

        geography_rows = list(
            workbook["Географии"].iter_rows(values_only=True)
        )
        self.assertEqual(len(geography_rows) - 1, 220)
        for index, expected in enumerate(counts.values(), start=1):
            self.assertEqual(
                sum(row[index] == "Да" for row in geography_rows[1:]),
                expected,
            )

        russia_values = {
            str(value)
            for row in workbook["Как указать всю Россию"].iter_rows(values_only=True)
            for value in row
            if value is not None
        }
        for alias in ("РФ", "Россия", "Российская Федерация"):
            self.assertIn(alias, russia_values)


if __name__ == "__main__":
    unittest.main()
