from __future__ import annotations

import ast
import json
import re
import sys
import tempfile
import unittest
import zipfile
from dataclasses import replace
from io import BytesIO
from pathlib import Path
from typing import Any, Callable

from openpyxl import load_workbook


WEB_APP_DIR = Path(__file__).resolve().parents[1]
PROJECT_ROOT = WEB_APP_DIR.parent
if str(WEB_APP_DIR) not in sys.path:
    sys.path.insert(0, str(WEB_APP_DIR))

from api.http_smoke import LocalApiState  # noqa: E402
from contracts.application_lifecycle_v1 import (  # noqa: E402
    BudgetByChannelPreview,
    BudgetByGeoPreview,
    ChannelFlightingPreview,
    LifecycleStatus,
    SamplingProfile,
    ValidationPreview,
    ValidationPreviewCheck,
    ValidationResultV1,
    parse_lifecycle_contract,
)
from services.campaign_template import build_campaign_plan_template  # noqa: E402
from services.local_campaign_service import (  # noqa: E402
    MAX_XLSX_WORKBOOK_XML_BYTES,
    LocalCampaignService,
    LocalCampaignServiceSettings,
    _xlsx_sheet_names,
)


HAPPY_FIXTURE = (
    WEB_APP_DIR
    / "tests"
    / "fixtures"
    / "application_lifecycle_v1_happy_path_synthetic.json"
)
FORBIDDEN_MARKETER_TERMS = (
    "backend",
    "model package",
    "model-aware validation",
    "campaign x geo x channel x target",
    "daily flighting",
    "calculation profile недоступен без campaign service",
    "calculation profile недоступен без паспорта активной модели",
)


def _literal_text(node: ast.AST | None) -> str:
    if node is None:
        return ""
    if isinstance(node, ast.Constant) and isinstance(node.value, str):
        return node.value
    if isinstance(node, ast.JoinedStr):
        return "".join(_literal_text(value) for value in node.values)
    if isinstance(node, ast.FormattedValue):
        return ""
    if isinstance(node, ast.BinOp) and isinstance(node.op, ast.Add):
        return _literal_text(node.left) + _literal_text(node.right)
    return "".join(_literal_text(child) for child in ast.iter_child_nodes(node))


class _DeferredExecutor:
    def __init__(self) -> None:
        self.calls: list[tuple[Callable[..., Any], tuple[Any, ...]]] = []

    def submit(self, function: Callable[..., Any], *args: Any) -> None:
        self.calls.append((function, args))

    def run_next(self) -> None:
        function, args = self.calls.pop(0)
        function(*args)


class NewCalculationPreviewV1Test(unittest.TestCase):
    def setUp(self) -> None:
        self.temporary = tempfile.TemporaryDirectory()
        root = Path(self.temporary.name)
        self.executor = _DeferredExecutor()
        self.state = LocalApiState(root / "state")
        self.settings = LocalCampaignServiceSettings(
            project_root=PROJECT_ROOT,
            artifact_root=root / "artifacts",
            validation_runtime_root=root / "runtime" / "validations",
            registry_root=root / "registry",
            registry_channel="preprod",
            expected_package_id="pkg_synthetic_preview",
            optimizer_policy_path=(
                PROJECT_ROOT
                / "02_Code"
                / "02_Budget_optimizer"
                / "optimizer_decision_policy_v2.yaml"
            ),
            business_policy_path=(
                PROJECT_ROOT
                / "02_Code"
                / "02_Budget_optimizer"
                / "business_threshold_policy_v1.yaml"
            ),
            default_sampling=SamplingProfile(64, 16, 32, 42, 10042),
        )
        self.service = LocalCampaignService(
            self.settings,
            self.state,
            self.executor,
            lambda _: (_ for _ in ()).throw(AssertionError("job must not be submitted")),
        )

    def tearDown(self) -> None:
        self.temporary.cleanup()

    def _upload_and_parse(self, filename: str, content: bytes) -> dict[str, Any]:
        upload, _ = self.service.create_upload(
            filename=filename,
            content=content,
            idempotency_key=f"preview-upload-{filename}-0001",
            actor_id="actor_11111111111111111111",
        )
        self.executor.run_next()
        return self.state.read_upload(upload["upload_id"])

    def _validate_count(self, upload: dict[str, Any]) -> dict[str, Any]:
        validation, _ = self.service.request_validation(
            upload["upload_id"],
            f"preview-validation-{upload['upload_id']}",
        )
        self.executor.run_next()
        return self.state.read_validation(validation["validation_id"])

    def _assert_actionable_issue(
        self,
        issue: dict[str, Any],
        expected_code: str,
    ) -> None:
        self.assertEqual(issue["code"], expected_code)
        for field_name in ("what", "why", "recommended_action"):
            self.assertIsInstance(issue[field_name], str)
            self.assertTrue(issue[field_name].strip())
        display_text = issue["display_text"].lower()
        for forbidden in FORBIDDEN_MARKETER_TERMS:
            self.assertNotIn(forbidden, display_text)

    def test_csv_and_xlsx_are_accepted_while_xls_and_tsv_are_rejected(self) -> None:
        csv_content = (
            "campaign_name,segment,geo,channel,start_date,end_date,budget_rub\n"
            "Synthetic A,SYNTHETIC_SEGMENT,SYNTHETIC_GEO,SYNTHETIC_CHANNEL,"
            "2026-01-01,2026-01-02,1000\n"
        ).encode("utf-8")
        csv_upload, _ = self.service.create_upload(
            filename="campaign.csv",
            content=csv_content,
            idempotency_key="preview-csv-upload-0001",
            actor_id="actor_11111111111111111111",
        )
        xlsx_upload, _ = self.service.create_upload(
            filename="campaign.xlsx",
            content=build_campaign_plan_template(),
            idempotency_key="preview-xlsx-upload-0001",
            actor_id="actor_11111111111111111111",
        )
        self.executor.run_next()
        self.executor.run_next()
        parsed_csv = self.state.read_upload(csv_upload["upload_id"])
        parsed_xlsx = self.state.read_upload(xlsx_upload["upload_id"])
        self.assertEqual(parsed_csv["status"]["code"], "parsed")
        self.assertEqual(parsed_xlsx["status"]["code"], "parsed")
        for filename in ("campaign.xls", "campaign.tsv"):
            with self.subTest(filename=filename), self.assertRaisesRegex(
                ValueError,
                "CSV и XLSX",
            ):
                self.service.create_upload(
                    filename=filename,
                    content=b"synthetic",
                    idempotency_key=f"preview-rejected-{filename}-0001",
                    actor_id="actor_11111111111111111111",
                )

    def test_blank_campaign_name_is_rejected_instead_of_collapsing_to_fallback(self) -> None:
        content = (
            "campaign_name,segment,geo,channel,start_date,end_date,budget_rub\n"
            ",SYNTHETIC_SEGMENT,SYNTHETIC_GEO,SYNTHETIC_CHANNEL,"
            "2026-01-01,2026-01-02,1000\n"
        ).encode("utf-8")
        upload = self._upload_and_parse("blank-name.csv", content)
        self.assertEqual(upload["status"]["code"], "rejected")
        self.assertIsNone(upload["detected_campaigns_n"])

    def test_zero_campaigns_block_validation_and_job_creation(self) -> None:
        header_only = (
            "campaign_name,segment,geo,channel,start_date,end_date,budget_rub\n"
        ).encode("utf-8")
        upload = self._upload_and_parse("zero.csv", header_only)
        self.assertEqual(upload["detected_campaigns_n"], 0)
        validation = self._validate_count(upload)
        self.assertEqual(validation["status"]["code"], "invalid")
        self.assertFalse(validation["job_creation_allowed"])
        self.assertEqual(
            validation["blocking_errors"][0]["code"],
            "CAMPAIGN_COUNT_NOT_ONE",
        )
        issue = validation["blocking_errors"][0]
        self._assert_actionable_issue(issue, "CAMPAIGN_COUNT_NOT_ONE")
        self.assertIn("campaign_name", issue["recommended_action"])
        with self.assertRaises(ValueError):
            self.service.create_job(
                validation["validation_id"],
                "preview-zero-job-0001",
            )

    def test_multiple_campaigns_are_not_auto_split_or_sent_to_model_validation(self) -> None:
        content = (
            "campaign_name,segment,geo,channel,start_date,end_date,budget_rub\n"
            "Synthetic A,SYNTHETIC_SEGMENT,SYNTHETIC_GEO,SYNTHETIC_CHANNEL,2026-01-01,2026-01-02,1000\n"
            "Synthetic B,SYNTHETIC_SEGMENT,SYNTHETIC_GEO,SYNTHETIC_CHANNEL,2026-01-01,2026-01-02,2000\n"
        ).encode("utf-8")
        upload = self._upload_and_parse("multiple.csv", content)
        self.assertEqual(upload["detected_campaigns_n"], 2)
        validation = self._validate_count(upload)
        self.assertEqual(validation["status"]["code"], "invalid")
        self.assertFalse(validation["job_creation_allowed"])
        issue = validation["blocking_errors"][0]
        self._assert_actionable_issue(issue, "CAMPAIGN_COUNT_NOT_ONE")
        self.assertIn("Разделите файл", issue["recommended_action"])
        self.assertEqual(self.executor.calls, [])

    def test_one_campaign_continues_to_normal_model_aware_validation(self) -> None:
        content = (
            "campaign_name,segment,geo,channel,start_date,end_date,budget_rub\n"
            "Synthetic A,SYNTHETIC_SEGMENT,SYNTHETIC_GEO,SYNTHETIC_CHANNEL,2026-01-01,2026-01-02,1000\n"
        ).encode("utf-8")
        upload = self._upload_and_parse("single.csv", content)
        validation, _ = self.service.request_validation(
            upload["upload_id"],
            "preview-single-validation-0001",
        )
        self.assertEqual(upload["detected_campaigns_n"], 1)
        self.assertEqual(validation["status"]["code"], "running")
        self.assertEqual(len(self.executor.calls), 1)
        function, args = self.executor.calls[0]
        self.assertEqual(function.__name__, "_validate_campaign")
        self.assertEqual(args[0].detected_campaigns_n, 1)

    def test_preview_aggregates_reconcile_and_contain_no_model_metrics(self) -> None:
        normalized = [
            {"channel": "A", "geo": "G1", "budget_rub": "100"},
            {"channel": "A", "geo": "G2", "budget_rub": "200"},
            {"channel": "B", "geo": "G1", "budget_rub": "300"},
        ]
        daily = [
            {"channel": "A", "geo": "G1", "date": "2026-01-01", "budget_rub": "50"},
            {"channel": "A", "geo": "G2", "date": "2026-01-01", "budget_rub": "100"},
            {"channel": "A", "geo": "G1", "date": "2026-01-02", "budget_rub": "50"},
            {"channel": "A", "geo": "G2", "date": "2026-01-02", "budget_rub": "100"},
            {"channel": "B", "geo": "G1", "date": "2026-01-01", "budget_rub": "150"},
            {"channel": "B", "geo": "G1", "date": "2026-01-02", "budget_rub": "150"},
        ]
        preview = self.service._validation_preview(normalized, daily, ())
        self.assertEqual(
            sum(row.total_budget_rub for row in preview.budget_by_channel or ()),
            600,
        )
        self.assertEqual(
            sum(row.total_budget_rub for row in preview.budget_by_geo or ()),
            600,
        )
        self.assertEqual(
            sum(row.daily_budget_rub for row in preview.channel_flighting or ()),
            600,
        )
        self.assertEqual(
            {row.channel: row.max_daily_budget_rub for row in preview.budget_by_channel or ()},
            {"A": 150, "B": 150},
        )
        serialized = json.dumps(preview, default=lambda value: value.__dict__)
        for forbidden in ("roas", "posterior", "incremental", "optimizer", "forecast"):
            self.assertNotIn(forbidden, serialized.lower())
        for check in preview.checks or ():
            self.assertIsNone(re.search(r"[A-Za-z]", check.display_text))

    def test_actionable_guidance_covers_unsupported_and_generic_failures(self) -> None:
        validation_id = "validation_aaaaaaaaaaaa"
        preparation = (
            self.settings.validation_runtime_root / validation_id / "preparation"
        )
        preparation.mkdir(parents=True)
        validation_file = preparation / "synthetic_campaign_model_validation.csv"
        validation_file.write_text(
            "supported_by_model,campaign_name,segment,geo,channel,target\n"
            "false,Synthetic A,Segment A,Geo A,Channel A,turnover_per_user\n",
            encoding="utf-8",
        )
        unsupported_issue = self.service._validation_failure_issue(validation_id)
        unsupported_payload = {
            **unsupported_issue.__dict__,
            "affected_cells": [cell.__dict__ for cell in unsupported_issue.affected_cells],
        }
        self._assert_actionable_issue(
            unsupported_payload,
            "UNSUPPORTED_MODEL_CELLS",
        )
        self.assertEqual(len(unsupported_payload["affected_cells"]), 1)

        generic_issue = self.service._validation_failure_issue(
            "validation_bbbbbbbbbbbb"
        )
        self._assert_actionable_issue(
            dict(generic_issue.__dict__),
            "CAMPAIGN_VALIDATION_FAILED",
        )

    def test_actionable_guidance_covers_caution_and_diagnostic_rows(self) -> None:
        fixture = json.loads(HAPPY_FIXTURE.read_text(encoding="utf-8"))
        record = parse_lifecycle_contract(fixture["validations"][0])
        self.assertIsInstance(record, ValidationResultV1)
        campaign = record.campaigns[0]
        validation_rows = [
            {
                "campaign_name": campaign.campaign_name,
                "segment": "Segment A",
                "geo": "Geo A",
                "channel": "Channel A",
                "target": "turnover_per_user",
                "allowed_use": "caution",
            },
            {
                "campaign_name": campaign.campaign_name,
                "segment": "Segment A",
                "geo": "Geo A",
                "channel": "Channel B",
                "target": "orders_per_user",
                "allowed_use": "diagnostic",
            },
        ]
        warnings = self.service._validation_warnings(
            validation_rows,
            record.campaigns,
        )
        by_code = {warning.code: dict(warning.__dict__) for warning in warnings}
        self._assert_actionable_issue(
            by_code["MODEL_CAUTION_CELLS"],
            "MODEL_CAUTION_CELLS",
        )
        self._assert_actionable_issue(
            by_code["MODEL_DIAGNOSTIC_CELLS"],
            "MODEL_DIAGNOSTIC_CELLS",
        )

    def test_marketer_facing_display_text_literals_avoid_internal_terms(self) -> None:
        paths = (
            WEB_APP_DIR / "api" / "http_smoke.py",
            WEB_APP_DIR / "services" / "local_campaign_service.py",
            WEB_APP_DIR / "services" / "job_progress_view.py",
            WEB_APP_DIR / "services" / "product_api_service.py",
            WEB_APP_DIR / "worker" / "execution_worker.py",
        )
        display_literals: list[tuple[Path, int, str]] = []
        for path in paths:
            tree = ast.parse(path.read_text(encoding="utf-8"), filename=str(path))
            for node in ast.walk(tree):
                if isinstance(node, ast.Call):
                    for keyword in node.keywords:
                        if keyword.arg == "display_text":
                            display_literals.append(
                                (path, node.lineno, _literal_text(keyword.value))
                            )
                if isinstance(node, ast.Dict):
                    for key, value in zip(node.keys, node.values):
                        if _literal_text(key) == "display_text":
                            display_literals.append(
                                (path, node.lineno, _literal_text(value))
                            )
        for path, line, display_text in display_literals:
            for forbidden in FORBIDDEN_MARKETER_TERMS:
                with self.subTest(path=path.name, line=line, term=forbidden):
                    self.assertNotIn(forbidden, display_text.lower())

    def test_preview_is_optional_and_geo_points_are_omitted_without_reference(self) -> None:
        fixture = json.loads(HAPPY_FIXTURE.read_text(encoding="utf-8"))
        old_payload = fixture["validations"][0]
        old_record = parse_lifecycle_contract(old_payload)
        self.assertIsInstance(old_record, ValidationResultV1)
        self.assertEqual(old_record.to_dict(), old_payload)
        for collection in ("blocking_errors", "warnings"):
            for issue in old_payload[collection]:
                self.assertNotIn("what", issue)
                self.assertNotIn("why", issue)
                self.assertNotIn("recommended_action", issue)

        total = old_record.totals.model_input_budget_rub  # type: ignore[union-attr]
        daily_total = old_record.totals.daily_budget_rub  # type: ignore[union-attr]
        preview = ValidationPreview(
            budget_by_channel=(
                BudgetByChannelPreview("SYNTHETIC_CHANNEL", total, daily_total),
            ),
            budget_by_geo=(
                BudgetByGeoPreview("SYNTHETIC_GEO", total, daily_total),
            ),
            channel_flighting=(
                ChannelFlightingPreview(
                    "SYNTHETIC_CHANNEL",
                    old_record.campaigns[0].start_date,
                    daily_total,
                ),
            ),
            geo_points=None,
            checks=(
                ValidationPreviewCheck(
                    "CAMPAIGN_COUNT",
                    "passed",
                    "В файле найдена ровно одна кампания.",
                ),
            ),
        )
        current = replace(old_record, preview=preview)
        current.validate()
        payload = current.to_dict()
        self.assertIn("preview", payload)
        self.assertNotIn("geo_points", payload["preview"])

    def test_template_is_valid_openxml_with_required_synthetic_sheets(self) -> None:
        content = build_campaign_plan_template()
        self.assertTrue(content.startswith(b"PK"))
        with zipfile.ZipFile(BytesIO(content)) as archive:
            self.assertIsNone(archive.testzip())
        workbook = load_workbook(BytesIO(content), data_only=True)
        self.assertIn("01_Daily", workbook.sheetnames)
        self.assertIn("02_Interval", workbook.sheetnames)
        cell_text = "\n".join(
            str(cell.value or "")
            for sheet in workbook.worksheets
            for row in sheet.iter_rows()
            for cell in row
        )
        self.assertIn("SYNTHETIC_DAILY_CAMPAIGN", cell_text)
        self.assertIn("SYNTHETIC_INTERVAL_CAMPAIGN", cell_text)
        for forbidden in ("МОСКВА", "ТС5", "РЕГ_ТВ", "X5"):
            self.assertNotIn(forbidden, cell_text.upper())

    def test_xlsx_archive_expansion_is_bounded_before_xml_parsing(self) -> None:
        path = Path(self.temporary.name) / "oversized.xlsx"
        with zipfile.ZipFile(path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
            archive.writestr(
                "xl/workbook.xml",
                b"x" * (MAX_XLSX_WORKBOOK_XML_BYTES + 1),
            )
        with self.assertRaisesRegex(ValueError, "metadata exceeds"):
            _xlsx_sheet_names(path)


if __name__ == "__main__":
    unittest.main()
